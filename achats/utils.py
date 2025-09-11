# achats/utils.py
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from django.conf import settings
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import tempfile
from datetime import datetime
from decimal import Decimal
# achats/utils.py
from django.db import transaction
from django.utils import timezone
from decimal import Decimal
from django.contrib import messages
import logging
from comptabilite.models import PlanComptableOHADA, JournalComptable, EcritureComptable, LigneEcriture
from achats.models import FactureFournisseur


def generate_bon_reception_pdf(bon_reception):
    """Génère un PDF pour un bon de réception avec coordonnées de l'entreprise"""
    # Vérification que reportlab est installé
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.units import mm
    except ImportError:
        raise ImportError("Le module reportlab n'est pas installé. Exécutez: pip install reportlab")
    
    buffer = io.BytesIO()
    
    # Création du document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm
    )
    
    styles = getSampleStyleSheet()
    elements = []
    
    # En-tête avec coordonnées de l'entreprise
    entreprise = bon_reception.entreprise
    
    # Style pour l'en-tête
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Tableau d'en-tête avec coordonnées
    header_data = []
    
    # Logo de l'entreprise (si disponible)
    try:
        if entreprise.logo and os.path.exists(entreprise.logo.path):
            logo = Image(entreprise.logo.path, width=80, height=40)
            header_data.append([logo, ''])
        else:
            header_data.append([Paragraph(f"<b>{entreprise.nom}</b>", styles['Heading2']), ''])
    except:
        header_data.append([Paragraph(f"<b>{entreprise.nom}</b>", styles['Heading2']), ''])
    
    # Coordonnées de l'entreprise
    coordonnees = []
    if entreprise.adresse:
        coordonnees.append(entreprise.adresse)

  
    if entreprise.telephone:
        coordonnees.append(f"Tél: {entreprise.telephone}")
    if entreprise.email:
        coordonnees.append(f"Email: {entreprise.email}")
    if entreprise.site_web:
        coordonnees.append(f"Site: {entreprise.site_web}")
    
    if coordonnees:
        header_data.append([Paragraph("<br/>".join(coordonnees), header_style), ''])
    
    header_table = Table(header_data, colWidths=[100*mm, 80*mm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 15))
    
    # Ligne séparatrice
    elements.append(Spacer(1, 2))
    elements.append(Table([['']], colWidths=[180*mm], style=[
        ('LINEABOVE', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(Spacer(1, 15))
    
    # Titre du document
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centered
    )
    elements.append(Paragraph("BON DE RÉCEPTION", title_style))
    
    # Informations du bon
    info_data = [
        ['Numéro du bon:', bon_reception.numero_bon],
        ['Date de réception:', bon_reception.date_reception.strftime('%d/%m/%Y')],
        ['Commande associée:', bon_reception.commande.numero_commande],
        ['Fournisseur:', bon_reception.commande.fournisseur.nom],
    ]
    
    info_table = Table(info_data, colWidths=[60*mm, 120*mm])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 15))
    
    ligne_data = [['Produit', 'Code', 'Quantité', 'Prix Unitaire', 'TVA %', 'Total HT', 'Montant TVA']]
    
    total_ht = Decimal('0.00')
    total_tva = Decimal('0.00')
    
    for ligne in bon_reception.lignes.all():
        # Calcul cohérent avec le modèle
        remise = ligne.ligne_commande.remise or Decimal('0.00')
        remise_factor = Decimal('1') - (remise / Decimal('100'))
        prix_apres_remise = ligne.prix_unitaire_ht * remise_factor
        
        ligne_ht = ligne.quantite * prix_apres_remise
        ligne_tva = ligne_ht * (ligne.taux_tva / Decimal('100'))
        
        total_ht += ligne_ht
        total_tva += ligne_tva
        
        # Récupération du code produit
        code_produit = getattr(ligne.ligne_commande.produit, 'code_barre_numero', 
                              getattr(ligne.ligne_commande.produit, 'reference', 
                                     getattr(ligne.ligne_commande.produit, 'code', '-')))
        
        ligne_data.append([
            ligne.ligne_commande.produit.nom,
            code_produit or '-',
            str(ligne.quantite),
            f"{ligne.prix_unitaire_ht:.2f}",
            f"{ligne.taux_tva:.2f}%",
            f"{ligne_ht:.2f}",
            f"{ligne_tva:.2f}"
        ])
    
    # Lignes de totaux - CORRECTION : Ajouter les totaux TVA et TTC
    ligne_data.append(['', '', '', '', '', 'TOTAL HT:', f"{total_ht:.2f}"])
    ligne_data.append(['', '', '', '', '', 'TOTAL TVA:', f"{total_tva:.2f}"])
    ligne_data.append(['', '', '', '', '', 'TOTAL TTC:', f"{total_ht + total_tva:.2f}"])
    
    # Ajuster les largeurs de colonnes
    ligne_table = Table(ligne_data, colWidths=[50*mm, 25*mm, 20*mm, 25*mm, 20*mm, 25*mm, 25*mm])
    ligne_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (-2, -1), (-1, -1), colors.lightblue),
        ('FONT', (-2, -1), (-1, -1), 'Helvetica-Bold', 10),
    ]))
    
    elements.append(ligne_table)
    elements.append(Spacer(1, 20))
    
    # Notes
    if bon_reception.notes:
        elements.append(Paragraph("Notes:", styles['Heading3']))
        elements.append(Paragraph(bon_reception.notes, styles['Normal']))
    
    # Pied de page avec signature
    elements.append(Spacer(1, 40))
    
    signature_data = [
        ['', ''],
        ['Le Responsable', 'Le Fournisseur'],
        ['', ''],
        ['_________________________', '_________________________'],
        ['Nom et signature', 'Nom et signature']
    ]
    
    signature_table = Table(signature_data, colWidths=[90*mm, 90*mm])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONT', (0, 1), (-1, 1), 'Helvetica-Bold', 10),
        ('FONT', (0, 3), (-1, 3), 'Helvetica', 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    elements.append(signature_table)
    elements.append(Spacer(1, 20))
    
    # Information de génération
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Italic'],
        fontSize=8,
        alignment=1  # Centered
    )
    
    footer_text = f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} par {bon_reception.created_by.get_full_name()}"
    elements.append(Paragraph(footer_text, footer_style))
    
    # Construction du PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

def generate_bon_reception_excel(bon_reception):
    """Génère un fichier Excel pour un bon de réception avec coordonnées de l'entreprise"""
    # Vérification que openpyxl est installé
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Le module openpyxl n'est pas installé. Exécutez: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Bon {bon_reception.numero_bon}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-tête avec coordonnées de l'entreprise
    entreprise = bon_reception.entreprise
    
    ws.merge_cells('A1:E1')
    ws['A1'] = entreprise.nom
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Coordonnées de l'entreprise
    coordonnees = []
    if entreprise.adresse:
        coordonnees.append(entreprise.adresse)
  
    if entreprise.code_postal:
        coordonnees.append(entreprise.code_postal)
    if entreprise.pays:
        coordonnees.append(str(entreprise.pays))
    if entreprise.telephone:
        coordonnees.append(f"Tél: {entreprise.telephone}")
    if entreprise.email:
        coordonnees.append(f"Email: {entreprise.email}")
    
    if coordonnees:
        ws.merge_cells('A2:E2')
        ws['A2'] = " - ".join(coordonnees)
        ws['A2'].alignment = Alignment(horizontal='center')
    
    # Titre du document
    ws.merge_cells('A3:E3')
    ws['A3'] = "BON DE RÉCEPTION"
    ws['A3'].font = Font(bold=True, size=16)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Informations du bon
    info_data = [
        ['Numéro du bon:', bon_reception.numero_bon],
        ['Date de réception:', bon_reception.date_reception.strftime('%d/%m/%Y')],
        ['Commande associée:', bon_reception.commande.numero_commande],
        ['Fournisseur:', bon_reception.commande.fournisseur.nom],
        ['Créé par:', bon_reception.created_by.get_full_name()],
    ]
    
    for i, (label, value) in enumerate(info_data, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Lignes de réception - CORRECTION : Ajouter les colonnes TVA
    headers = ['Produit', 'Code', 'Quantité', 'Prix Unitaire', 'TVA %', 'Total HT', 'Montant TVA']
    start_row = 10
    
    total_ht = Decimal('0.00')
    total_tva = Decimal('0.00')
    
    for row, ligne in enumerate(bon_reception.lignes.all(), start=start_row + 1):
        # Calcul cohérent avec le modèle
        remise = ligne.ligne_commande.remise or Decimal('0.00')
        remise_factor = Decimal('1') - (remise / Decimal('100'))
        prix_apres_remise = ligne.prix_unitaire_ht * remise_factor
        
        ligne_ht = ligne.quantite * prix_apres_remise
        ligne_tva = ligne_ht * (ligne.taux_tva / Decimal('100'))
        
        total_ht += ligne_ht
        total_tva += ligne_tva
        
        # Récupération du code produit
        code_produit = getattr(ligne.ligne_commande.produit, 'code_barre_numero', 
                              getattr(ligne.ligne_commande.produit, 'reference', 
                                     getattr(ligne.ligne_commande.produit, 'code', '-')))
        
        ws.cell(row=row, column=1, value=ligne.ligne_commande.produit.nom).border = thin_border
        ws.cell(row=row, column=2, value=code_produit or '-').border = thin_border
        ws.cell(row=row, column=3, value=float(ligne.quantite)).border = thin_border
        ws.cell(row=row, column=4, value=float(ligne.prix_unitaire_ht)).border = thin_border
        ws.cell(row=row, column=5, value=float(ligne.taux_tva)).border = thin_border
        ws.cell(row=row, column=6, value=float(ligne_ht)).border = thin_border
        ws.cell(row=row, column=7, value=float(ligne_tva)).border = thin_border
    
    # Lignes de totaux - CORRECTION : Ajouter les totaux TVA et TTC
    total_row = start_row + 1 + bon_reception.lignes.count()
    
    ws.cell(row=total_row, column=5, value="TOTAL HT:").font = total_font
    ws.cell(row=total_row, column=5).fill = total_fill
    ws.cell(row=total_row, column=5).border = thin_border
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=6, value=float(total_ht)).font = total_font
    ws.cell(row=total_row, column=6).fill = total_fill
    ws.cell(row=total_row, column=6).border = thin_border
    
    ws.cell(row=total_row + 1, column=5, value="TOTAL TVA:").font = total_font
    ws.cell(row=total_row + 1, column=5).fill = total_fill
    ws.cell(row=total_row + 1, column=5).border = thin_border
    ws.cell(row=total_row + 1, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row + 1, column=6, value=float(total_tva)).font = total_font
    ws.cell(row=total_row + 1, column=6).fill = total_fill
    ws.cell(row=total_row + 1, column=6).border = thin_border
    
    ws.cell(row=total_row + 2, column=5, value="TOTAL TTC:").font = total_font
    ws.cell(row=total_row + 2, column=5).fill = total_fill
    ws.cell(row=total_row + 2, column=5).border = thin_border
    ws.cell(row=total_row + 2, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row + 2, column=6, value=float(total_ht + total_tva)).font = total_font
    ws.cell(row=total_row + 2, column=6).fill = total_fill
    ws.cell(row=total_row + 2, column=6).border = thin_border
    
    # Ajustement des largeurs de colonnes
    column_widths = [40, 20, 15, 15, 12, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Sauvegarde dans un fichier temporaire
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    return tmp_path

logger = logging.getLogger(__name__)
def generer_facture_depuis_bon(bon_reception, request):
    """
    Génère une facture fournisseur avec écritures comptables à partir d'un bon de réception
    VERSION CORRIGÉE - Utilise les taux TVA réels du bon
    """
    try:
        with transaction.atomic():
            # Vérifier si une facture existe déjà pour ce bon
            facture_existante = FactureFournisseur.objects.filter(
                bon_reception=bon_reception,
                entreprise=request.entreprise
            ).first()
            
            if facture_existante:
                messages.warning(request, f"Une facture existe déjà pour ce bon: {facture_existante.numero_facture}")
                return facture_existante
            
            # FORCER le recalcul des totaux du bon AVANT de créer la facture
            bon_reception.calculer_totaux()
            bon_reception.refresh_from_db()
            
            # VÉRIFICATION CRITIQUE : S'assurer que les totaux sont valides
            if bon_reception.total_ht <= 0 or bon_reception.total_ttc <= 0:
                messages.error(request, "Les totaux du bon de réception sont invalides")
                return None
            
            # Créer la facture avec les totaux RECALCULÉS du bon
            facture = FactureFournisseur(
                entreprise=request.entreprise,
                fournisseur=bon_reception.commande.fournisseur,
                bon_reception=bon_reception,
                date_facture=timezone.now().date(),
                date_echeance=timezone.now().date() + timezone.timedelta(days=30),
                montant_ht=bon_reception.total_ht,      # UTILISER LES VALEURS DU BON
                montant_tva=bon_reception.total_tva,    # UTILISER LES VALEURS DU BON
                montant_ttc=bon_reception.total_ttc,    # UTILISER LES VALEURS DU BON
                statut='brouillon',
                created_by=request.user
            )
            facture.save()
            
            # Créer les écritures comptables
            creer_ecritures_comptables_facture(facture, request)
            
            messages.success(request, f"Facture {facture.numero_facture} créée avec succès avec les écritures comptables.")
            return facture
            
    except Exception as e:
        logger.error(f"Erreur génération facture depuis bon: {e}", exc_info=True)
        messages.error(request, f"Erreur lors de la génération de la facture: {str(e)}")
        return None

def creer_ecritures_comptables_facture(facture, request):
    """
    Crée les écritures comptables pour une facture fournisseur
    Utilise les comptes OHADA standards
    """
    try:
        entreprise = request.entreprise
        
        # Récupérer ou créer les comptes OHADA standards
        compte_achats, created = PlanComptableOHADA.objects.get_or_create(
            numero='607',
            entreprise=entreprise,
            defaults={
                'classe': '6',
                'intitule': 'Achats de marchandises',
                'type_compte': 'charge',
                'description': 'Compte pour enregistrer les achats de marchandises'
            }
        )
        
        # Utiliser le compte TVA OHADA standard (4455)
        compte_tva, created = PlanComptableOHADA.objects.get_or_create(
            numero='4455',
            entreprise=entreprise,
            defaults={
                'classe': '4',
                'intitule': 'TVA à décaisser',
                'type_compte': 'passif',
                'description': 'TVA collectée à reverser à l\'état'
            }
        )
        
        compte_fournisseurs, created = PlanComptableOHADA.objects.get_or_create(
            numero='401',
            entreprise=entreprise,
            defaults={
                'classe': '4',
                'intitule': 'Fournisseurs',
                'type_compte': 'passif',
                'description': 'Compte des dettes envers les fournisseurs'
            }
        )
        
        # Récupérer ou créer le journal des achats
        journal_achats, created = JournalComptable.objects.get_or_create(
            code='ACH',
            entreprise=entreprise,
            defaults={
                'intitule': 'Journal des Achats',
                'type_journal': 'achat'
            }
        )
        
        # Créer l'écriture comptable principale
        ecriture = EcritureComptable(
            journal=journal_achats,
            date_ecriture=timezone.now(),
            date_comptable=facture.date_facture,
            libelle=f"Facture {facture.numero_facture} - {facture.fournisseur.nom}",
            piece_justificative=facture.numero_facture,
            entreprise=entreprise,
            created_by=request.user,
            facture_fournisseur_liee=facture
        )
        ecriture.save()
        
        # Ligne 1: Débit des achats (HT)
        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=compte_achats,
            libelle=f"Achats HT - {facture.fournisseur.nom}",
            debit=facture.montant_ht,
            credit=Decimal('0.00'),
            entreprise=entreprise
        )
        
        # Ligne 2: Débit de la TVA
        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=compte_tva,
            libelle=f"TVA {facture.montant_tva} - {facture.fournisseur.nom}",
            debit=facture.montant_tva,
            credit=Decimal('0.00'),
            entreprise=entreprise
        )
        
        # Ligne 3: Crédit des fournisseurs (TTC)
        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=compte_fournisseurs,
            libelle=f"Dette fournisseur - {facture.fournisseur.nom}",
            debit=Decimal('0.00'),
            credit=facture.montant_ttc,
            entreprise=entreprise
        )
        
        # Vérifier l'équilibre de l'écriture
        total_debit = sum(float(ligne.debit) for ligne in ecriture.lignes.all())
        total_credit = sum(float(ligne.credit) for ligne in ecriture.lignes.all())
        
        if abs(total_debit - total_credit) > Decimal('0.01'):
            logger.warning(f"Écriture déséquilibrée: débit={total_debit}, crédit={total_credit}")
        
        return ecriture
        
    except Exception as e:
        logger.error(f"Erreur création écritures comptables: {e}", exc_info=True)
        raise Exception(f"Erreur lors de la création des écritures comptables: {str(e)}")