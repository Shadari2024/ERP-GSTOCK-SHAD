from django.shortcuts import render,redirect,get_object_or_404
from django.views import View
from .models import *

from django.db.models import F
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.utils.dateparse import parse_date
from django.utils import timezone
from io import BytesIO
import csv
from xhtml2pdf import pisa
from weasyprint import HTML
from django.template.loader import get_template
from achats.models import *
from ventes.models import *
from parametres.mixins import EntrepriseAccessMixin
from django.db.models.functions import TruncMonth
from django.views.decorators.csrf import csrf_exempt
import json
from security.models import * # Import ProfilUtilisateur
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.views.generic import ListView, DetailView, UpdateView, DeleteView, CreateView
from django.db import transaction
from django.template.loader import render_to_string
from django.core.mail import EmailMessage, send_mail
from django.conf import settings
from io import BytesIO
from xhtml2pdf import pisa
from decimal import Decimal
from django.utils.timezone import now, timedelta
from django.db.models import Sum
from django.utils import formats
from datetime import date
from security.models import UtilisateurPersonnalise
from django.db.models.functions import Abs
from django.db.models import Case, When, F, Sum
from django.http import FileResponse
from django.views.decorators.http import require_POST
from django.core.mail import send_mail
from django.contrib.auth.hashers import make_password
from django.utils.crypto import get_random_string
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from collections import defaultdict
import tempfile
from weasyprint import HTML
from decimal import Decimal
from django.core.mail import mail_admins
from io import BytesIO
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import Group
from .utils import envoyer_mail_bienvenue 
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required,user_passes_test
from django.db import transaction
from decimal import Decimal
import csv
import tempfile
from django.utils.timezone import now
from django.db.models import Q
from django.db.models import Sum,F,Count
from .utils import bilan_caisse_du_jour
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime, timedelta
from weasyprint import HTML
from django.template.loader import get_template
from django.http import HttpResponse
from weasyprint import HTML
import tempfile
from .models import InventairePhysique
from django.contrib.auth.decorators import login_required
import json
import base64
from django.views.generic.edit import DeleteView
from django.urls import reverse_lazy
from STOCK.notifications import *
from STOCK.ml.prophet_utils import *
# Create your views here.
from weasyprint import HTML
from django.http import HttpResponse
from django.template.loader import get_template
from parametres.models import ConfigurationSAAS, TauxChange
# Ajoutez ceci au début de votre fichier views.py ou settings.py
import matplotlib
matplotlib.use('Agg')  # Mode non-interactif
import matplotlib.pyplot as plt



from django.http import JsonResponse
from .models import TauxChange, Parametre  # ajuste l'import selon ton app
from django.contrib.auth.decorators import login_required

@login_required
def taux_change_api(request):
    devise_cible = request.GET.get('devise', 'USD').upper()  # sécurité : normalise la saisie
    try:
        parametres = Parametre.objects.get(user=request.user)
        devise_principale = parametres.devise_principale.upper()
    except Parametre.DoesNotExist:
        devise_principale = 'FC'  # valeur par défaut

    # Si la devise demandée est identique à la devise principale
    if devise_cible == devise_principale:
        return JsonResponse({'taux': 1.0})

    try:
        # Chercher un taux direct
        taux = TauxChange.objects.get(
            devise_source=devise_principale,
            devise_cible=devise_cible
        ).taux
        return JsonResponse({'taux': float(taux)})

    except TauxChange.DoesNotExist:
        try:
            # Chercher un taux inverse et l'inverser
            taux_inverse = TauxChange.objects.get(
                devise_source=devise_cible,
                devise_cible=devise_principale
            ).taux
            return JsonResponse({'taux': round(1.0 / float(taux_inverse), 6)})

        except TauxChange.DoesNotExist:
            return JsonResponse({'error': 'Taux non disponible entre %s et %s' % (
                devise_principale, devise_cible)}, status=404)

from django.views import View
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator

from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.core.exceptions import PermissionDenied
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.template.loader import get_template
from io import BytesIO
from xhtml2pdf import pisa
from decimal import Decimal
from .models import Client, Devise
from parametres.mixins import EntrepriseAccessMixin
# STOCK/views.py
from django_countries import countries
from django.views.generic import ListView

@method_decorator([
    login_required,
    permission_required('STOCK.view_client', raise_exception=True)
], name='dispatch')
class ClientListView(EntrepriseAccessMixin, ListView):
    model = Client
    template_name = 'clients/liste_clients.html'
    context_object_name = 'Clients'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().filter(entreprise=self.request.entreprise)

        # Filtre de recherche générale
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(nom__icontains=query) |
                Q(email__icontains=query) |
                Q(telephone__icontains=query) |
                Q(adresse__icontains=query) |
                Q(ville__icontains=query) |
                Q(code_postal__icontains=query) |
                Q(numero_fiscal__icontains=query) |
                Q(numero_tva__icontains=query)
            ).distinct()

        # Filtres avancés
        statut = self.request.GET.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)

        type_client = self.request.GET.get('type_client')
        if type_client:
            queryset = queryset.filter(type_client=type_client)

        pays = self.request.GET.get('pays')
        if pays:
            queryset = queryset.filter(pays=pays)

        date_creation = self.request.GET.get('date_creation')
        today = timezone.now().date()
        if date_creation == 'today':
            queryset = queryset.filter(date_creation__date=today)
        elif date_creation == 'week':
            start_of_week = today - timedelta(days=today.weekday())
            queryset = queryset.filter(date_creation__date__gte=start_of_week)
        elif date_creation == 'month':
            queryset = queryset.filter(date_creation__year=today.year, date_creation__month=today.month)
        elif date_creation == 'year':
            queryset = queryset.filter(date_creation__year=today.year)

        # Tri par défaut
        queryset = queryset.order_by('-cree_le')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_entreprise'] = self.request.entreprise
        context['is_paginated'] = context['page_obj'].has_other_pages
        
        # Statistiques pour le tableau de bord
        clients = Client.objects.filter(entreprise=self.request.entreprise)
        context['stats'] = {
            'actifs': clients.filter(statut='ACTIF').count(),
            'inactifs': clients.filter(statut='INACTIF').count(),
            'entreprises': clients.filter(type_client='ENT').count(),
            'particuliers': clients.filter(type_client='PART').count(),
        }
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
            # ... autres contextes ...
        context['pays_options'] = list(countries)
        return context
    
# STOCK/views.py

from django.views.generic import ListView, View
from django.db.models import Q, Case, When
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from decimal import Decimal
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives# Assurez-vous que le chemin est correct

@method_decorator([
    login_required,
    permission_required('STOCK.add_client', raise_exception=True)
], name='dispatch')
class FormulaireViewClient(EntrepriseAccessMixin, View):
    template_name = "clients/formulaire_client.html"
    
    def get(self, request):
        config_saas, created = ConfigurationSAAS.objects.get_or_create(
            entreprise=request.entreprise,
            defaults={
                'fuseau_horaire': 'UTC',
                'langue': 'fr',
                'expiration_session': 30,
                'complexite_mdp': ConfigurationSAAS.ComplexiteMDP.MOYEN
            }
        )
        
        devises = Devise.objects.filter(active=True)
        if config_saas.devise_principale:
            devises = devises.order_by(
                Case(
                    When(id=config_saas.devise_principale_id, then=0),
                    default=1
                )
            )
        
        context = {
            'devises': devises,
            'current_entreprise': request.entreprise,
            'pays_liste': self.get_pays_liste()
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        form_data = request.POST.dict()
        form_data['entreprise'] = request.entreprise
        form_data['cree_par'] = request.user
        
        try:
            self.validate_client_data(form_data)
            client = Client.objects.create(**self.prepare_client_data(form_data))
            
            if 'photo' in request.FILES:
                client.photo = request.FILES['photo']
                client.save()
            
            self.send_welcome_email(client)
            
            messages.success(request, f"Client {client.nom} créé avec succès")
            # --- C'EST ICI QUE LA CORRECTION EST FAITE ---
            return redirect('liste_client') # Utilisez le namespace et le nom corrects
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
            return self.get(request)
    
    def validate_client_data(self, data):
        """Valide les données du client avant création"""
        if Client.objects.filter(entreprise=data['entreprise'], email=data['email']).exists():
            raise ValueError("Un client avec cet email existe déjà")
        
        if not data.get('nom'):
            raise ValueError("Le nom est obligatoire")
    
    def prepare_client_data(self, data):
        """Prépare les données pour la création"""
        return {
            'entreprise': data['entreprise'],
            'type_client': data.get('type_client', 'PART'),
            'nom': data['nom'],
            'statut': data.get('statut', 'ACT'),
            'telephone': data['telephone'],
            'telephone_secondaire': data.get('telephone_secondaire'),
            'email': data.get('email'),
            'website': data.get('website'),
            'adresse': data.get('adresse'),
            'ville': data.get('ville'),
            'code_postal': data.get('code_postal'),
            'pays': data.get('pays'),
            'devise_preferee_id': data.get('devise_preferee'),
            'limite_credit': Decimal(data.get('limite_credit', 0)),
            'delai_paiement': int(data.get('delai_paiement', 30)),
            'taux_remise': Decimal(data.get('taux_remise', 0)),
            'numero_tva': data.get('numero_tva'),
            'numero_fiscal': data.get('numero_fiscal'),
            'exonere_tva': data.get('exonere_tva') == 'on',
            'notes': data.get('notes'),
            'cree_par': data['cree_par']
        }
    
    def send_welcome_email(self, client):
        """Envoie l'email de bienvenue au client"""
        subject = f"Bienvenue chez {client.entreprise.nom}"
        context = {
            'client': client,
            'entreprise': client.entreprise
        }
        html_content = render_to_string('emails/bienvenue_client.html', context)
        
        msg = EmailMultiAlternatives(
            subject,
            '',
            settings.DEFAULT_FROM_EMAIL,
            [client.email],
            reply_to=[client.entreprise.email]
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
    
    def get_pays_liste(self):
        """Retourne une liste de pays pour le select"""
        from django_countries import countries
        return list(countries)




@method_decorator([
    login_required,
    permission_required('STOCK.view_client', raise_exception=True)
], name='dispatch')
class Detailsclient(EntrepriseAccessMixin, View):
    template_name = 'clients/clientdetails.html'

    def get(self, request, my_id): # On utilise my_id comme discuté précédemment pour l'ID entier
        client = get_object_or_404(
            Client,
            pk=my_id, # Recherche par clé primaire (id)
            entreprise=request.entreprise
        )

        context = {
            'client': client,
            'current_entreprise': request.entreprise,
            'historique_commandes': self.get_commandes(client)
        }
        return render(request, self.template_name, context)

    def get_commandes(self, client):
        """Récupère l'historique des commandes du client"""
        # Assurez-vous que votre modèle Commande a un champ 'client' (ForeignKey vers Client)
        # et que vous avez importé le modèle Commande
        return CommandeClient.objects.filter(client=client).order_by('-date_commande')




# Solution recommandée : Utiliser Django's DeleteView pour simplifier
from django.views.generic.edit import DeleteView

@method_decorator([
    login_required,
    permission_required('STOCK.delete_client', raise_exception=True)
], name='dispatch')


class ClientDeleteView(EntrepriseAccessMixin, DeleteView):
    model = Client
    template_name = 'clients/client_confirm_delete.html'
    context_object_name = 'client' # Le nom de la variable du contexte pour le template
    success_url = reverse_lazy('liste_client') # Redirige après succès

    # Si vous voulez ajouter des messages de succès/erreur personnalisés
    def form_valid(self, form):
        messages.success(self.request, f"Client {self.object.nom} supprimé avec succès.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, f"Erreur lors de la suppression du client.")
        return super().form_invalid(form)

    # Assurez-vous que l'entreprise correspond
    def get_queryset(self):
        return super().get_queryset().filter(entreprise=self.request.entreprise)

# Si vous voulez absolument rester avec votre vue 'SupClient' personnalisée,
# vous devez la modifier pour gérer GET (afficher le formulaire) et POST (effectuer la suppression)

# @method_decorator([
#     login_required,
#     permission_required('STOCK.delete_client', raise_exception=True)
# ], name='dispatch')
# class SupClient(EntrepriseAccessMixin, View):
#     template_name = 'clients/client_confirm_delete.html' # Nouveau template

#     def get(self, request, client_id): # Affiche le formulaire de confirmation
#         client = get_object_or_404(
#             Client,
#             pk=client_id,
#             entreprise=request.entreprise
#         )
#         return render(request, self.template_name, {'client': client})

#     def post(self, request, client_id): # Gère la suppression
#         client = get_object_or_404(
#             Client,
#             pk=client_id,
#             entreprise=request.entreprise
#         )
#         try:
#             nom_client = client.nom
#             client.delete()
#             messages.success(request, f"Client {nom_client} supprimé avec succès.")
#             return redirect('liste_clients') # Redirige vers la liste après suppression
#         except Exception as e:
#             messages.error(request, f"Erreur lors de la suppression: {str(e)}")
#             return redirect('details_client', client_id=client.pk) # Retourne aux détails en cas d'erreur



@method_decorator([
    login_required,
    permission_required('STOCK.change_client', raise_exception=True)
], name='dispatch')
class ModifClient(EntrepriseAccessMixin, View):
    template_name = "clients/modificatclient.html"
    
    def get(self, request, client_id):
        client = get_object_or_404(
            Client,
            pk=client_id,
            entreprise=request.entreprise
        )
        
        # Start with all active devises
        devises = Devise.objects.filter(active=True)

        # Order devises: if the client has a preferred currency, put it first
        if client.devise_preferee:
            devises = devises.order_by(
                Case(
                    When(id=client.devise_preferee_id, then=0), # Put preferred first
                    default=1 # Others after
                )
            )

        context = {
            'client': client,
            'devises': devises, # The ordered list of devises
            'current_entreprise': request.entreprise,
            'pays_liste': self._get_pays_liste()
        }
        return render(request, self.template_name, context)
        
    def post(self, request, client_id):
        # ... (Your existing post logic here, it remains the same)
        # Ensure you handle all fields and save the client object
        client = get_object_or_404(
            Client,
            pk=client_id,
            entreprise=request.entreprise
        )
        
        try:
            client.nom = request.POST.get('nom')
            client.email = request.POST.get('email')
            client.telephone = request.POST.get('telephone')
            client.adresse = request.POST.get('adresse')
            
            # Update other fields (ensure all fields are handled here if not using a form)
            client.ville = request.POST.get('ville')
            client.code_postal = request.POST.get('code_postal')
            client.pays = request.POST.get('pays')
            client.numero_tva = request.POST.get('numero_tva')
            client.notes = request.POST.get('notes')
            client.telephone_secondaire = request.POST.get('telephone_secondaire')
            client.website = request.POST.get('website')
            client.code_client = request.POST.get('code_client')
            client.numero_fiscal = request.POST.get('numero_fiscal')
            
            # Numeric fields (convert to Decimal/int)
            client.limite_credit = float(request.POST.get('limite_credit', '0.00'))
            client.delai_paiement = int(request.POST.get('delai_paiement', '0'))
            client.taux_remise = float(request.POST.get('taux_remise', '0.00'))

            # Select fields
            client.type_client = request.POST.get('type_client')
            client.statut = request.POST.get('statut')

            # Checkbox
            client.exonere_tva = 'exonere_tva' in request.POST
            
            # Devise preferee
            if 'devise_preferee' in request.POST and request.POST['devise_preferee']:
                devise_id = request.POST['devise_preferee']
                client.devise_preferee = Devise.objects.get(id=devise_id)
            else:
                client.devise_preferee = None
            
            # Photo management
            if 'photo' in request.FILES:
                client.photo = request.FILES['photo']
            elif 'clear_photo' in request.POST:
                client.photo = None
            
            client.save()
            messages.success(request, "Client mis à jour avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de la mise à jour: {str(e)}")
            return redirect('modif_client', client_id=client.pk)
        
        return redirect('details_client', my_id=client.pk)

    def _get_pays_liste(self):
        # Your existing _get_pays_liste logic
        try:
            from django_countries import countries
            return list(countries)
        except ImportError:
            return Client._meta.get_field('pays').choices

@method_decorator([login_required], name='dispatch')
class GenerateClientPDF(EntrepriseAccessMixin, View):
    def get(self, request, client_uuid):
        client = get_object_or_404(
            Client,
            uuid=client_uuid,
            entreprise=request.entreprise
        )
        
        template_path = 'clients/carte_client_pdf.html'
        context = {
            'client': client,
            'entreprise': request.entreprise
        }
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="carte_client_{client.code_client}.pdf"'
        
        template = get_template(template_path)
        html = template.render(context)
        
        pisa_status = pisa.CreatePDF(
            BytesIO(html.encode('UTF-8')),
            dest=response,
            encoding='UTF-8'
        )
        
        if pisa_status.err:
            return HttpResponse('Erreur lors de la génération du PDF')
        
        return response
    
    
# classe categorie pour affiche et voir
@method_decorator([
    login_required,
    permission_required('STOCK.view_categorie', raise_exception=True)
], name='dispatch')
class CatView(EntrepriseAccessMixin, View):
    template_name = 'categorie/liste_cat.html'
    
    def get(self, request):
        # Filtrage par entreprise courante
        categories = Categorie.objects.filter(entreprise=request.entreprise)
        
        # Vérification des permissions avancées
        can_add = request.user.has_perm('STOCK.add_categorie')
        can_change = request.user.has_perm('STOCK.change_categorie')
        can_delete = request.user.has_perm('STOCK.delete_categorie')
        
        # Pagination
        paginator = Paginator(categories, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'can_add': can_add,
            'can_change': can_change,
            'can_delete': can_delete,
            'current_entreprise': request.entreprise  # Pour l'affichage dans le template
        }
        return render(request, self.template_name, context)
      
    def post(self, request):
        pass
      
    #formulaire cat
@method_decorator([
    login_required,
    permission_required('STOCK.add_categorie', raise_exception=True)
], name='dispatch')

class FormulaireCat(EntrepriseAccessMixin, View):
    template_name = "categorie/form_cat.html"
    
    def get(self, request):
        # Vérification des permissions
        if not request.user.has_perm('STOCK.view_categorie'):
            messages.error(request, "Vous n'avez pas la permission de voir les catégories")
            return redirect('acces_refuse')
        
        # Filtrage par entreprise courante
        categories = Categorie.objects.filter(entreprise=request.entreprise)
        
        context = {
            'cat': categories,
            'can_change': request.user.has_perm('STOCK.change_categorie'),
            'can_delete': request.user.has_perm('STOCK.delete_categorie'),
            'current_entreprise': request.entreprise  # Pour le template
        }
        return render(request, self.template_name, context)
        
    def post(self, request):
        nom = request.POST.get('nom')
        photo = request.FILES.get('photo')

        if nom and photo:
            try:
                # Création avec l'entreprise courante
                categorie = Categorie.objects.create(
                    nom=nom, 
                    photo=photo,
                    entreprise=request.entreprise,
                    cree_par=request.user
                )
                messages.success(request, "Catégorie ajoutée avec succès !")
                return redirect('liste_cat')
            except Exception as e:
                messages.error(request, f"Erreur : {str(e)}")
        else:
            messages.error(request, "Nom et photo sont obligatoires")
        
        # Réaffichage du formulaire avec les catégories de l'entreprise
        return render(request, self.template_name, {
            'cat': Categorie.objects.filter(entreprise=request.entreprise)
        })
        
        
@method_decorator([
    login_required,
    permission_required('STOCK.change_categorie', raise_exception=True)
], name='dispatch')
class UpdateCatView(EntrepriseAccessMixin, View):
    template_name = "categorie/edit_cat.html"
    
    def get(self, request, pk):
        # Vérification que la catégorie appartient à l'entreprise
        categorie = get_object_or_404(
            Categorie, 
            pk=pk, 
            entreprise=request.entreprise
        )
        
        return render(request, self.template_name, {
            'categorie': categorie,
            'current_entreprise': request.entreprise
        })
        
    def post(self, request, pk):
        categorie = get_object_or_404(
            Categorie, 
            pk=pk, 
            entreprise=request.entreprise
        )
        
        nom = request.POST.get('nom')
        photo = request.FILES.get('photo')

        if nom:
            try:
                categorie.nom = nom
                if photo:
                    categorie.photo = photo
                categorie.save()
                
                messages.success(request, "Catégorie modifiée avec succès !")
                return redirect('liste_cat')
            except Exception as e:
                messages.error(request, f"Erreur : {str(e)}")
        else:
            messages.error(request, "Le nom est obligatoire")
        
        return render(request, self.template_name, {
            'categorie': categorie,
            'current_entreprise': request.entreprise
        })
        
@method_decorator([
    login_required,
    permission_required('STOCK.delete_categorie', raise_exception=True)
], name='dispatch')
class DeleteCatView(EntrepriseAccessMixin, View):
    def post(self, request, pk):
        # Vérification que la catégorie appartient à l'entreprise
        categorie = get_object_or_404(
            Categorie, 
            pk=pk, 
            entreprise=request.entreprise
        )
        
        try:
            # Vérifier qu'aucun produit n'est associé
            if categorie.produit_set.exists():
                messages.error(request, "Impossible de supprimer : des produits sont associés à cette catégorie")
            else:
                categorie.delete()
                messages.success(request, "Catégorie supprimée avec succès !")
        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
        
        return redirect('liste_cat')
    #ajouter produit 
# Mettez à jour vos imports
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.contrib import messages
from decimal import Decimal
from django.db import transaction

# Assurez-vous d'importer votre modèle ConfigurationSAAS
from parametres.models import ConfigurationSAAS
# Assurez-vous d'importer votre mixin
from parametres.mixins import EntrepriseAccessMixin

# Mettez à jour vos imports de modèles
from .models import Produit, Categorie


@method_decorator([
    login_required,
    permission_required('stock.add_produit', raise_exception=True)
], name='dispatch')
class AjouterProduitView(EntrepriseAccessMixin, View):
    template_name = 'produit/ajouter_produit.html'

    def get_devise_principale_symbole(self, request):
        """Récupère le symbole de la devise principale de l'entreprise."""
        try:
            config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
            return config_saas.devise_principale.symbole if config_saas.devise_principale else "€"
        except ConfigurationSAAS.DoesNotExist:
            return "€"

    def get(self, request, *args, **kwargs):
        """Affiche le formulaire d'ajout de produit."""
        context = {
            'categories': Categorie.objects.filter(entreprise=request.entreprise),
            'devise_principale_symbole': self.get_devise_principale_symbole(request),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """Traite les données du formulaire pour créer un produit."""
        try:
            with transaction.atomic():
                # Récupération et validation des données du formulaire
                nom = request.POST.get('nom')
                if not nom:
                    raise ValueError("Le nom du produit est requis.")

                description = request.POST.get('description', '')
                prix_achat = Decimal(request.POST.get('prix_achat', '0'))
                prix_vente = Decimal(request.POST.get('prix_vente', '0'))
                stock = int(request.POST.get('stock', '0'))
                seuil_alerte = int(request.POST.get('seuil_alerte', '0'))
                categorie_id = request.POST.get('categorie')
                photo = request.FILES.get('photo')

                categorie = get_object_or_404(Categorie, id=categorie_id, entreprise=request.entreprise)
                
                # Création du produit
                produit = Produit(
                    nom=nom,
                    description=description,
                    prix_achat=prix_achat,
                    prix_vente=prix_vente,
                    stock=stock,
                    seuil_alerte=seuil_alerte,
                    categorie=categorie,
                    photo=photo,
                    cree_par=request.user,
                    entreprise=request.entreprise
                )

                produit.save()
                
                # Génération du code-barre et sauvegarde
                produit.generate_barcode()
                produit.save()

            messages.success(request, f"Le produit '{produit.nom}' a été ajouté avec succès.")
            return redirect('produits_par_categorie')
            
        except ValueError as ve:
            messages.error(request, f"Erreur de données: {str(ve)}")
        except Exception as e:
            messages.error(request, f"Une erreur inattendue est survenue: {str(e)}")
            
        # Recharger le contexte pour afficher le formulaire avec les erreurs
        context = {
            'categories': Categorie.objects.filter(entreprise=request.entreprise),
            'form_data': request.POST,
            'devise_principale_symbole': self.get_devise_principale_symbole(request),
        }
        return render(request, self.template_name, context)
#etiquette

import base64
from django.shortcuts import get_object_or_404, render
from django.template.loader import get_template
from django.http import HttpResponse
from decimal import Decimal
from weasyprint import HTML
from django.core.files.storage import default_storage # Importez default_storage
from .models import Produit
from parametres.models import ConfigurationSAAS, TauxChange

def ticket_produit_pdf(request, pk):
    produit = get_object_or_404(Produit, pk=pk, entreprise=request.entreprise)
    
    devise_principale_symbole = "$"
    devise_principale_code = 'USD'
    logo_base64 = None
    code_barre_base64 = None
    
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        if config_saas.devise_principale:
            devise_principale_symbole = config_saas.devise_principale.symbole
            devise_principale_code = config_saas.devise_principale.code
    except ConfigurationSAAS.DoesNotExist:
        pass
    
    taux_usd = None
    prix_vente_usd = None
    if devise_principale_code != 'USD':
        taux_usd = TauxChange.objects.filter(
            devise_source__code=devise_principale_code,
            devise_cible__code='USD',
            est_actif=True
        ).order_by('-date_application').first()
        if taux_usd:
            try:
                prix_vente_usd = Decimal(produit.prix_vente) * Decimal(taux_usd.taux)
            except:
                prix_vente_usd = None

    # Conversion du logo en Base64 si le fichier existe
    if request.entreprise.logo:
        # Utilisez default_storage pour un accès compatible avec S3 ou le système de fichiers
        if default_storage.exists(request.entreprise.logo.name):
            with default_storage.open(request.entreprise.logo.name, "rb") as logo_file:
                logo_base64 = base64.b64encode(logo_file.read()).decode('utf-8')

    # Conversion du code-barres en Base64 si le fichier existe
    if produit.code_barre:
        if default_storage.exists(produit.code_barre.name):
            with default_storage.open(produit.code_barre.name, "rb") as barcode_file:
                code_barre_base64 = base64.b64encode(barcode_file.read()).decode('utf-8')

    template = get_template('produit/ticket_produit.html')
    html = template.render({
        'produit': produit,
        'entreprise': request.entreprise,
        'devise_principale_symbole': devise_principale_symbole,
        'taux_usd': taux_usd.taux if taux_usd else None,
        'prix_vente_usd': prix_vente_usd,
        'logo_base64': logo_base64,
        'code_barre_base64': code_barre_base64,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename=ticket_produit_{produit.id}.pdf'
    
    # WeasyPrint n'a plus besoin du base_url car les images sont encodées
    HTML(string=html).write_pdf(response)
    return response

from django.core.files.storage import default_storage

def imprimer_tickets_en_stock_pdf(request):
    """Génère un PDF contenant un ticket par produit en stock."""
    
    # 1. Récupérer les données essentielles
    produits = Produit.objects.filter(entreprise=request.entreprise, stock__gt=0)
    
    devise_principale_symbole = "$"
    taux_usd = None
    
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        if config_saas.devise_principale:
            devise_principale_symbole = config_saas.devise_principale.symbole
            devise_principale_code = config_saas.devise_principale.code
            if devise_principale_code != 'USD':
                taux_usd = TauxChange.objects.filter(
                    devise_source__code=devise_principale_code,
                    devise_cible__code='USD',
                    est_actif=True
                ).order_by('-date_application').first()
    except ConfigurationSAAS.DoesNotExist:
        pass
        
    # 2. Préparer le contexte pour chaque produit
    liste_tickets = []
    
    # Préparez le logo une seule fois pour tous les tickets
    logo_base64 = None
    if request.entreprise.logo:
        if default_storage.exists(request.entreprise.logo.name):
            with default_storage.open(request.entreprise.logo.name, "rb") as logo_file:
                logo_base64 = base64.b64encode(logo_file.read()).decode('utf-8')

    for produit in produits:
        prix_vente_usd = None
        if taux_usd:
            try:
                prix_vente_usd = Decimal(produit.prix_vente) * Decimal(taux_usd.taux)
            except:
                pass
        
        # Préparez le code-barres pour chaque produit
        code_barre_base64 = None
        if produit.code_barre:
            if default_storage.exists(produit.code_barre.name):
                with default_storage.open(produit.code_barre.name, "rb") as barcode_file:
                    code_barre_base64 = base64.b64encode(barcode_file.read()).decode('utf-8')
        
        liste_tickets.append({
            'produit': produit,
            'entreprise': request.entreprise,
            'devise_principale_symbole': devise_principale_symbole,
            'prix_vente_usd': prix_vente_usd,
            'taux_usd': taux_usd,
            'code_barre_base64': code_barre_base64,
            'logo_base64': logo_base64,
            'current_date': timezone.now().date(),
        })

    # 3. Générer le PDF
    template = get_template('produit/liste_tickets.html')
    html = template.render({'liste_tickets': liste_tickets})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=tickets_en_stock.pdf'
    
    HTML(string=html).write_pdf(response)
    return response


 #affiche produit
@login_required
@permission_required('STOCK.view_produit', raise_exception=True)
def produits_list(request):
    # Récupération de la configuration SAAS
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale.code if config_saas.devise_principale else 'USD'
    except ConfigurationSAAS.DoesNotExist:
        devise_principale = 'USD'

    # Vérification permission spéciale pour la conversion devise
    if not request.user.has_perm('STOCK.view_tauxchange'):
        devises_acceptees = [devise_principale]
        devise_demandee = devise_principale
    else:
        devises_acceptees = ['USD', 'EUR', 'CDF']  # Exemple de devises acceptées
        devise_demandee = request.GET.get('devise', devise_principale)

    search_query = request.GET.get('search', '')
    
    # Correction: Ajout d'un order_by() pour éviter le warning de pagination
    produits = Produit.objects.filter(entreprise=request.entreprise)\
                             .annotate(stock_reel=F('stock'))\
                             .order_by('nom')  # Tri par nom ou autre champ pertinent
    
    if search_query:
        produits = produits.filter(Q(nom__icontains=search_query) | Q(description__icontains=search_query))

    # Conversion des devises
    if (request.user.has_perm('STOCK.view_tauxchange') and 
        devise_demandee != devise_principale and 
        devise_demandee in devises_acceptees):
        
        taux = TauxChange.objects.filter(
            devise_source__code=devise_principale,
            devise_cible__code=devise_demandee,
            est_actif=True
        ).order_by('-date_application').first()
        
        if taux:
            taux_decimal = Decimal(taux.taux)
            for produit in produits:
                produit.prix_vente_converti = Decimal(produit.prix_vente) * taux_decimal
                produit.prix_achat_converti = Decimal(produit.prix_achat) * taux_decimal
                produit.devise_affichage = devise_demandee
    else:
        for produit in produits:
            produit.prix_vente_converti = produit.prix_vente
            produit.prix_achat_converti = produit.prix_achat
            produit.devise_affichage = devise_principale

    # Pagination avec queryset ordonné
    paginator = Paginator(produits, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'produit/liste_produit.html', {
        'page_obj': page_obj,
        'devise_principale': devise_principale,
        'devises_acceptees': devises_acceptees,
        'devise_active': devise_demandee,
        'devise_symbols': {'USD': '$', 'EUR': '€', 'CDF': 'FC'},
        'can_edit': request.user.has_perm('STOCK.change_produit'),
        'can_delete': request.user.has_perm('STOCK.delete_produit'),
        'current_entreprise': request.entreprise
    })

@login_required
@permission_required('STOCK.view_produit', raise_exception=True)
def produits_search(request):
    search_query = request.GET.get('search', '')
    
    # Correction: Ajout d'un order_by() pour la recherche
    produits = Produit.objects.filter(
        entreprise=request.entreprise,
        nom__icontains=search_query
    ).order_by('nom')[:5]  # Tri par nom

    # Récupération de la devise principale
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale.code if config_saas.devise_principale else 'USD'
    except ConfigurationSAAS.DoesNotExist:
        devise_principale = 'USD'

    devise_demandee = request.GET.get('devise', devise_principale)

    results = []
    for p in produits:
        prix_vente = p.prix_vente
        if (request.user.has_perm('STOCK.view_tauxchange') and 
            devise_demandee != devise_principale):
            
            taux = TauxChange.objects.filter(
                devise_source__code=devise_principale,
                devise_cible__code=devise_demandee,
                est_actif=True
            ).order_by('-date_application').first()
            
            if taux:
                try:
                    prix_vente = Decimal(p.prix_vente) * Decimal(taux.taux)
                except Exception as e:
                    print(f"Erreur conversion recherche: {e}")
        
        results.append({
            'id': p.id,
            'nom': p.nom,
            'prix': str(Decimal(str(prix_vente)).quantize(Decimal('0.01'))),
            'devise': devise_demandee,
            'can_view': request.user.has_perm('STOCK.view_produit')
        })

    return JsonResponse({'results': results})
from django.conf import settings
import os
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
from decimal import Decimal
from django.conf import settings
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.core.files.base import ContentFile
from django.db.models import F, Avg
from .models import Produit, Categorie # Assurez-vous d'importer vos modèles




@method_decorator([
    login_required,
    permission_required('STOCK.delete_produit', raise_exception=True)
], name='dispatch')
class DeleteProduitView(EntrepriseAccessMixin, View):
    def post(self, request, pk):
        produit = get_object_or_404(
            Produit, 
            pk=pk, 
            entreprise=request.entreprise
        )
        
        try:
            produit.delete()
            messages.success(request, f"Le produit '{produit.nom}' a été supprimé avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression du produit : {e}")
            
        return redirect('produits_par_categorie')







@login_required
@permission_required('STOCK.view_produit', raise_exception=True)
def produit_detail(request, id):
    """
    Vue détaillée d'un produit, avec gestion des permissions et conversion de devise.
    """
    # Récupération du produit avec filtre entreprise pour la sécurité
    produit = get_object_or_404(Produit, id=id, entreprise=request.entreprise)
    
    # DEBUG: Ajoutez ces prints pour voir ce qui se passe
    print(f"=== DEBUG PRODUIT {produit.id} ===")
    print(f"Photo: {produit.photo}")
    print(f"Photo name: {produit.photo.name if produit.photo else 'None'}")
    print(f"Photo path: {produit.photo.path if produit.photo else 'None'}")
    print(f"Photo URL: {produit.photo.url if produit.photo else 'None'}")
    print(f"Code-barres: {produit.code_barre}")
    print(f"Code-barres name: {produit.code_barre.name if produit.code_barre else 'None'}")
    
    # Vérifiez l'existence physique des fichiers
    import os
    if produit.photo:
        print(f"Photo exists: {os.path.exists(produit.photo.path)}")
    if produit.code_barre:
        print(f"Code-barres exists: {os.path.exists(produit.code_barre.path)}")
    
    
    # Initialisation des variables pour la devise et la configuration
    devise_principale = None
    taux_change = None
    devises_disponibles = []
    
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale
        devises_disponibles = config_saas.devises_acceptees.all()
        
        # Conversion de devise si l'utilisateur a la permission et une devise est demandée
        if request.user.has_perm('STOCK.view_tauxchange'):
            devise_demandee = request.GET.get('devise', devise_principale.code)
            if devise_demandee != devise_principale.code:
                taux = config_saas.tauxchanges.filter(
                    devise_cible__code=devise_demandee,
                    est_actif=True
                ).order_by('-date_application').first()
                
                if taux:
                    taux_change = {
                        'code': devise_demandee,
                        'taux': taux.taux,
                        'symbole': taux.devise_cible.symbole
                    }
                    produit.prix_vente_converti = Decimal(produit.prix_vente) * Decimal(taux.taux)
                    # Le prix d'achat n'est converti que si l'utilisateur peut le voir
                    if hasattr(produit, 'prix_achat') and request.user.has_perm('STOCK.view_prix_achat'):
                        produit.prix_achat_converti = Decimal(produit.prix_achat) * Decimal(taux.taux)
    except ConfigurationSAAS.DoesNotExist:
        # En cas d'absence de configuration, on peut définir des valeurs par défaut
        pass # Les variables sont déjà initialisées à None ou []

    # Vérification des permissions pour le contexte du template
    show_prix_achat = request.user.has_perm('STOCK.view_prix_achat')
    can_convert_devise = request.user.has_perm('STOCK.view_tauxchange')
    
    context = {
        'produit': produit,
        'can_edit': request.user.has_perm('STOCK.change_produit'),
        'can_delete': request.user.has_perm('STOCK.delete_produit'),
        'can_view_price': request.user.has_perm('STOCK.view_prix_produit'),
        'show_prix_achat': show_prix_achat,
        'current_entreprise': request.entreprise,
        'devise_principale': {
            'code': devise_principale.code if devise_principale else 'USD',
            'symbole': devise_principale.symbole if devise_principale else '$'
        },
        'taux_change': taux_change,
        'can_convert_devise': can_convert_devise,
        'devises_disponibles': devises_disponibles,
    }
    
    return render(request, 'produit/fiche_produit.html', context)
    


from django.db.models import Sum, Count
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
import logging
logger = logging.getLogger(__name__)

@login_required
@permission_required('STOCK.view_commande', raise_exception=True)
def ventes_du_jour(request):
    """
    Vue SAAS pour afficher les ventes du jour avec gestion multi-devises
    """
    # 1. Vérification de l'entreprise (doit être géré par middleware SAAS)
    if not hasattr(request, 'entreprise') or not request.entreprise:
        logger.error("Accès sans entreprise définie")
        return redirect('security:acces_refuse')

    entreprise = request.entreprise

    # 2. Récupération de la configuration SAAS
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=entreprise)
        devise_entreprise = config_saas.devise_principale
        if not devise_entreprise:
            raise ValueError("Devise principale non configurée")
    except (ObjectDoesNotExist, ValueError) as e:
        logger.error(f"Configuration SAAS manquante pour {entreprise}: {e}")
        messages.error(request, "Configuration commerciale incomplète")
        return redirect('security:acces_refuse')

    # 3. Filtrage SAAS des ventes du jour
    today = timezone.now().date()
    ventes = CommandeClient.objects.filter(
        entreprise=entreprise,  # Filtrage SAAS critique
        date_commande__date=today,
        vente_confirmee=True
    ).select_related('client', 'vendeur')

    # 4. Gestion multi-devises
    devise_code = request.GET.get('devise', devise_entreprise.code)
    conversion_data = handle_currency_conversion(
        ventes, 
        from_currency=devise_entreprise,  # On passe l'objet Devise directement
        to_currency_code=devise_code     # Nom du paramètre corrigé
    )

    # 5. Calcul des totaux
    totals = ventes.aggregate(
        total_original=Sum('montant_total'),
        count=Count('id')
    )

    context = {
        'ventes_converties': conversion_data['ventes'],
        'total_original': totals['total_original'] or 0,
        'total_converti': conversion_data['total_converti'],
        'devise_entreprise': devise_entreprise,
        'devise_affichee': conversion_data['devise_cible'].code if hasattr(conversion_data['devise_cible'], 'code') else conversion_data['devise_cible'],
        'taux_conversion': conversion_data['taux'],
        'sens_conversion': conversion_data['sens'],
        'devises_disponibles': Devise.objects.filter(active=True),
        'nombre_ventes': totals['count'] or 0,
        'date_jour': today,
        'config_saas': config_saas  # Pour accès aux autres paramètres
    }

    return render(request, 'vente/ventes_du_jour.html', context)

def handle_currency_conversion(ventes, from_currency, to_currency_code):
    """
    Helper SAAS pour gérer les conversions de devises
    from_currency: Objet Devise (la devise principale de l'entreprise)
    to_currency_code: Code de la devise cible (string)
    """
    if to_currency_code == from_currency.code:
        return {
            'ventes': [{
                'vente': v, 
                'montant_converti': v.montant_total,
                'devise_cible': from_currency
            } for v in ventes],
            'total_converti': sum(v.montant_total for v in ventes) if ventes else 0,
            'devise_cible': from_currency,
            'taux': 1.0,
            'sens': f"1 {from_currency.code} = 1 {from_currency.code}"
        }

    try:
        devise_cible = Devise.objects.get(code=to_currency_code, active=True)
        taux = devise_cible.taux_par_defaut / from_currency.taux_par_defaut
        
        ventes_converties = []
        total_converti = 0
        
        for vente in ventes:
            montant = vente.montant_total * taux
            ventes_converties.append({
                'vente': vente,
                'montant_converti': round(montant, devise_cible.decimales),
                'devise_cible': devise_cible
            })
            total_converti += montant

        return {
            'ventes': ventes_converties,
            'total_converti': round(total_converti, devise_cible.decimales),
            'devise_cible': devise_cible,
            'taux': taux,
            'sens': f"1 {from_currency.code} = {taux:.4f} {devise_cible.code}"
        }

    except Devise.DoesNotExist:
        logger.warning(f"Devise {to_currency_code} non trouvée, utilisation de {from_currency.code}")
        return handle_currency_conversion(ventes, from_currency, from_currency.code)
    
    
def ticket_pdf(request, commande_id):
    commande = get_object_or_404(Commande, id=commande_id)
    parametre = getattr(commande, 'parametre', None)
    if not parametre:
        parametre = Parametre.objects.first() # Assure-toi que ce champ existe sur le modèle Commande

    # Préparation du contexte
    context = {
        'commande': commande,
        'parametre': parametre
    }

    # Rendu HTML à partir du template
    template = get_template('vente/ticket_pdf.html')
    html_string = template.render(context)

    # Création du PDF via WeasyPrint
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="ticket_{commande.id}.pdf"'

    html.write_pdf(response)

    return response

def bon_commande_pdf(request, commande_id):
    commande = get_object_or_404(Commande, id=commande_id)
    lignes = LigneCommandeClient.objects.filter(commande=commande)
    context = {
        'commande': commande,
        'lignes': lignes,
    }
    return generer_pdf('pdf/bon_commande.html', context)


#code bare
def chercher_produit_par_code_barres(request):
    code = request.GET.get('code', '')
    try:
        produit = Produit.objects.get(code_barres=code)
        data = {
            'id': produit.id,
            'nom': produit.nom,
            'prix_vente': float(produit.prix_vente),
        }
        return JsonResponse(data)
    except Produit.DoesNotExist:


        return JsonResponse({'error': 'Produit non trouvé'}, status=404)
   
@login_required
@permission_required('security.historique_vente', raise_exception=True)
def historique_ventes(request):
    """
    Vue SAAS pour l'historique des ventes avec gestion multi-devises
    """
    # 1. Vérification de l'entreprise
    if not hasattr(request, 'entreprise') or not request.entreprise:
        logger.error("Accès sans entreprise définie")
        messages.error(request, "Accès non autorisé - entreprise non définie")
        return redirect('security:acces_refuse')

    entreprise = request.entreprise

    # 2. Récupération de la configuration SAAS
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=entreprise)
        devise_entreprise = config_saas.devise_principale
        if not devise_entreprise:
            raise ValueError("Devise principale non configurée")
    except (ObjectDoesNotExist, ValueError) as e:
        logger.error(f"Configuration SAAS manquante pour {entreprise}: {e}")
        messages.error(request, "Configuration commerciale incomplète")
        return redirect('config:completement_config')

    # 3. Filtrage SAAS des ventes
    ventes = CommandeClient.objects.filter(
        entreprise=entreprise,
        vente_confirmee=True
    ).select_related('client', 'vendeur')

    # Appliquer les filtres supplémentaires
    date = request.GET.get('date')
    vendeur = request.GET.get('vendeur')
    search_query = request.GET.get('search', '').strip()

    if date:
        ventes = ventes.filter(date_commande__date=date)
    if vendeur:
        ventes = ventes.filter(vendeur__id=vendeur)
    if search_query:
        ventes = ventes.filter(
            Q(id__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(vendeur__username__icontains=search_query)
        )

    # 4. Gestion multi-devises - Utilisation de la devise principale de la config SAAS
    devise_code = request.GET.get('devise', devise_entreprise.code)
    conversion_data = handle_currency_conversion(
        ventes, 
        from_currency=devise_entreprise,  # On passe l'objet Devise directement
        to_currency_code=devise_code
    )

    # 5. Calcul des totaux (optimisé)
    totals = ventes.aggregate(
        total_original=Sum('montant_total'),
        count=Count('id')
    )

    context = {
        'ventes_converties': conversion_data['ventes'],
        'total_original': totals['total_original'] or 0,
        'total_converti': conversion_data['total_converti'],
        'devise_entreprise': devise_entreprise,
        'devise_affichee': conversion_data['devise_cible'],
        'taux_conversion': conversion_data['taux'],
        'sens_conversion': conversion_data['sens'],
        'vendeurs': CommandeClient.objects.filter(entreprise=entreprise)
                          .values('vendeur__id', 'vendeur__username')
                          .distinct(),
        'devises_disponibles': Devise.objects.filter(active=True),
        'selected_date': date,
        'selected_vendeur': vendeur,
        'search_query': search_query,
        'config_saas': config_saas,
        'entreprise': entreprise,
        'can_export': request.user.has_perm('STOCK.export_commande')
    }

    return render(request, 'vente/historique_ventes.html', context)

def handle_currency_conversion(ventes, from_currency, to_currency_code):
    """
    Helper SAAS pour gérer les conversions de devises
    from_currency: Objet Devise (la devise principale de l'entreprise)
    to_currency_code: Code de la devise cible
    """
    if to_currency_code == from_currency.code:
        return {
            'ventes': [{
                'vente': v, 
                'montant_converti': v.montant_total,
                'devise_cible': from_currency
            } for v in ventes],
            'total_converti': sum(v.montant_total for v in ventes) if ventes else 0,
            'devise_cible': from_currency,
            'taux': 1.0,
            'sens': f"1 {from_currency.code} = 1 {from_currency.code}"
        }

    try:
        devise_cible = Devise.objects.get(code=to_currency_code, active=True)
        taux = devise_cible.taux_par_defaut / from_currency.taux_par_defaut
        
        ventes_converties = []
        total_converti = 0
        
        for vente in ventes:
            montant = vente.montant_total * taux
            ventes_converties.append({
                'vente': vente,
                'montant_converti': round(montant, devise_cible.decimales),
                'devise_cible': devise_cible
            })
            total_converti += montant

        return {
            'ventes': ventes_converties,
            'total_converti': round(total_converti, devise_cible.decimales),
            'devise_cible': devise_cible,
            'taux': taux,
            'sens': f"1 {from_currency.code} = {taux:.4f} {devise_cible.code}"
        }

    except Devise.DoesNotExist:
        logger.warning(f"Devise {to_currency_code} non trouvée, utilisation de {from_currency.code}")
        return handle_currency_conversion(ventes, from_currency, from_currency.code)
    
    #Statistiques
def format_devise(self, montant, devise=None):
    devise = devise or self.devise_principale
    symboles = {
        'USD': '$',
        'EUR': '€',
        'CDF': 'FC',
        'FC': 'FC'
    }
    symbole = symboles.get(devise, devise)
    return f"{symbole} {montant:,.2f}"

from decimal import Decimal
from django.db.models import Sum, Count, F
from django.db.models.functions import Coalesce
from django.utils.timezone import now
from django.contrib import messages
from django.shortcuts import redirect, render
from django.core.exceptions import ObjectDoesNotExist

@login_required
@permission_required('security.voir_statistiques', raise_exception=True)
def statistiques(request):
    # Vérification SAAS - entreprise requise
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Accès non autorisé - entreprise non définie")
        return redirect('security:acces_refuse')
    
    entreprise = request.entreprise
    
    # Récupération config SAAS
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=entreprise)
        devise_source = config_saas.devise_principale.code if config_saas.devise_principale else "USD"
    except ConfigurationSAAS.DoesNotExist:
        messages.error(request, "Configuration SAAS manquante")
        return redirect('security:acces_refuse')

    devise_cible = request.GET.get("devise", devise_source)
    aujourd_hui = now().date()

    # Taux de change avec conversion en Decimal
    try:
        if devise_source != devise_cible:
            taux_obj = TauxChange.objects.filter(
                devise_source=devise_source,
                devise_cible=devise_cible,
                entreprise=entreprise
            ).latest('date')
            taux_change = Decimal(str(taux_obj.taux))
        else:
            taux_change = Decimal('1.0')
    except (TauxChange.DoesNotExist, ObjectDoesNotExist):
        taux_change = Decimal('1.0')

    # Ventes du jour pour l'entreprise
    ventes_du_jour = CommandeClient.objects.filter(
        entreprise=entreprise,
        date_commande__date=aujourd_hui
    ).select_related('vendeur')
    
    # Calcul des agrégats avec Coalesce
    totals = ventes_du_jour.aggregate(
        total_journalier=Coalesce(Sum('montant_total'), Decimal('0.00')),
        total_transactions=Count('id')
    )
    
    # Conversion et calcul
    total_journalier_source = totals['total_journalier']
    total_journalier_converted = total_journalier_source * taux_change
    ventes_jour_converted = total_journalier_converted  # Alias pour le template

    # Produits populaires
    produits_stats = (
        LigneCommandeClient.objects
        .filter(
            commande__entreprise=entreprise,
            commande__date_commande__date=aujourd_hui
        )
        .annotate(
            ca_ligne=Coalesce(F('quantite') * F('prix_unitaire'), Decimal('0.00'))
        )
        .values(
            'produit__nom',
            'produit__code_barre_numero'
        )
        .annotate(
            qte_totale=Sum('quantite'),
            ca_produit=Sum('ca_ligne')
        )
        .order_by('-qte_totale')[:5]
    )

    # Performance vendeurs - version corrigée sans conflit d'annotation
    vendeurs_stats = (
        ventes_du_jour
        .values(
            'vendeur__id',  # Utilisation directe du champ relation
            'vendeur__username'
        )
        .annotate(
            total_vendeur=Coalesce(Sum('montant_total'), Decimal('0.00')),
            nb_ventes=Count('id')
        )
        .order_by('-total_vendeur')
    )

    # Renommage des champs pour le template
    vendeurs_stats = [
        {
            'id': v['vendeur__id'],
            'nom': v['vendeur__username'],
            'total_vendeur': v['total_vendeur'],
            'nb_ventes': v['nb_ventes']
        }
        for v in vendeurs_stats
    ]

    # Calcul des performances
    for vendeur in vendeurs_stats:
        vendeur['total_vendeur_converted'] = vendeur['total_vendeur'] * taux_change
        performance = (vendeur['total_vendeur_converted'] / total_journalier_converted * 100) if total_journalier_converted > 0 else 0
        vendeur['performance'] = performance
        
        if performance >= 50:
            vendeur['perf_color'] = 'success'
            vendeur['perf_icon'] = 'arrow-up'
        elif performance >= 20:
            vendeur['perf_color'] = 'warning'
            vendeur['perf_icon'] = 'arrow-right'
        else:
            vendeur['perf_color'] = 'danger'
            vendeur['perf_icon'] = 'arrow-down'

    # Autres statistiques
    nb_utilisateurs = UtilisateurPersonnalise.objects.filter(entreprise=entreprise).count()
    utilisateurs_actifs = UtilisateurPersonnalise.objects.filter(
        entreprise=entreprise,
        last_login__date=aujourd_hui
    ).count()
    
    # Calcul du total produits
    total_produits = sum(p['qte_totale'] for p in produits_stats) if produits_stats else 0
    for produit in produits_stats:
        produit['ca_produit_converted'] = produit['ca_produit'] * taux_change
        produit['part_market'] = (produit['qte_totale'] / total_produits * 100) if total_produits > 0 else 0
        produit['nom'] = produit['produit__nom']
        produit['reference'] = produit['produit__code_barre_numero']

    # Préparation du contexte
    context = {
        'now': now(),
        'devise': devise_cible,
        'taux_change': float(taux_change),
        'entreprise': entreprise,
        'config_saas': config_saas,
        
        # Statistiques principales
        'total_journalier_converted': total_journalier_converted,
        'ventes_jour_converted': ventes_jour_converted,
        'total_transactions': totals['total_transactions'],
        'total_produits': total_produits,
        
        # Utilisateurs
        'nb_utilisateurs': nb_utilisateurs,
        'utilisateurs_actifs': utilisateurs_actifs,
        
        # Tableaux
        'produits_populaires': produits_stats,
        'performance_vendeurs': vendeurs_stats,
        
        # Pour le sélecteur de devise
        'param': {
            'devises_acceptees': ['USD', 'CDF', 'EUR']  # À adapter selon votre modèle
        }
    }
    
    return render(request, 'stat/statistiquesVente.html', context)

def bilan_caisse_du_jour(vendeur):
    today = timezone.now().date()
    commandes = CommandeClient.objects.filter(
        vente_au_comptoir=True,
        vente_confirmee=True,
        vendeur=vendeur,
        date_commande__date=today
    )

    total = commandes.aggregate(Sum("montant_total"))["montant_total__sum"] or 0
    nb_commandes = commandes.count()

    return {
        "total": total,
        "nb_commandes": nb_commandes,
        "commandes": commandes
    }
from django.db.models import Sum, Count
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.shortcuts import render, redirect
from django.utils import timezone
from decimal import Decimal
from django.conf import settings
from django.core.mail import send_mail

@login_required
@permission_required('security.gerer_caisse', raise_exception=True)
def cloturer_caisse(request):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    today = timezone.now().date()
    user = request.user
    
    # Vérification si clôture déjà effectuée
    if ClotureCaisse.objects.filter(
        vendeur=user, 
        date_jour=today,
        entreprise=request.entreprise
    ).exists():
        messages.warning(request, "La caisse a déjà été clôturée aujourd'hui.")
        return redirect('dashboard')
    
    # Récupération des ventes du jour
    ventes_du_jour = CommandeClient.objects.filter(
        vendeur=user,
        vente_au_comptoir=True,
        date_commande__date=today,
        vente_confirmee=True,
        entreprise=request.entreprise
    )
    
    total_ventes = ventes_du_jour.aggregate(
        total=Sum('montant_total'),
        count=Count('id')
    )
    total_vendu = total_ventes['total'] or Decimal('0.00')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                montant_espece = Decimal(request.POST.get('montant_espece', '0'))
                montant_carte = Decimal(request.POST.get('montant_carte', '0'))
                commentaire = request.POST.get('commentaire', '')
                
                if montant_espece < 0 or montant_carte < 0:
                    raise ValueError("Les montants ne peuvent pas être négatifs")
                    
                total_declare = montant_espece + montant_carte
                tolerance = Decimal('0.01')
                
                # Création de la clôture
                cloture = ClotureCaisse.objects.create(
                    vendeur=user,
                    date_jour=today,
                    montant_total=total_vendu,
                    nombre_ventes=total_ventes['count'] or 0,
                    montant_espece=montant_espece,
                    montant_carte=montant_carte,
                    commentaire=commentaire,
                    validee=True,
                    entreprise=request.entreprise
                )
                
                # Gestion des écarts
                ecart_montant = abs(total_vendu - total_declare)
                if ecart_montant > tolerance:
                    type_ecart = 'manquant' if total_vendu > total_declare else 'excedent'
                    
                    # Enregistrement de l'écart
                    ecart = EcartCaisse.objects.create(
                        cloture=cloture,
                        montant=ecart_montant,
                        type_ecart=type_ecart,
                        commentaire=f"Écart détecté lors de la clôture. {commentaire}",
                        entreprise=request.entreprise
                    )
                    
                    # Notification admin
                    email_content = f"""
                    <h3>Nouvel écart de caisse enregistré</h3>
                    <p><strong>Entreprise:</strong> {request.entreprise.nom}</p>
                    <p><strong>Vendeur:</strong> {user.get_full_name()} ({user.username})</p>
                    <p><strong>Type:</strong> {ecart.get_type_ecart_display()}</p>
                    <p><strong>Montant:</strong> {ecart.montant} {request.entreprise.devise_principale.symbole}</p>
                    <p><strong>Clôture:</strong> #{cloture.id} du {cloture.date_jour}</p>
                    <p><strong>Détails:</strong></p>
                    <ul>
                        <li>Total système: {total_vendu} {request.entreprise.devise_principale.symbole}</li>
                        <li>Total déclaré: {total_declare} {request.entreprise.devise_principale.symbole}</li>
                        <li>Espèces: {montant_espece} {request.entreprise.devise_principale.symbole}</li>
                        <li>Carte: {montant_carte} {request.entreprise.devise_principale.symbole}</li>
                    </ul>
                    """
                    
                    send_mail(
                        subject=f"[{request.entreprise.nom}] ECART CAISSE - {ecart.get_type_ecart_display()} de {ecart.montant}{request.entreprise.devise_principale.symbole}",
                        message="",
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[admin[1] for admin in settings.ADMINS],
                        html_message=email_content,
                        fail_silently=False,
                    )
                    
                    ecart.notifie_admin = True
                    ecart.save()
                    
                    messages.warning(
                        request,
                        f"Clôture validée avec écart {type_ecart} de {ecart_montant}{request.entreprise.devise_principale.symbole} enregistré."
                    )
                else:
                    messages.success(request, "Clôture validée sans écart.")
                
                return redirect('rapport_cloture', pk=cloture.id)
                
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
            return redirect('cloturer_caisse')
    
    context = {
        'total_ventes': total_vendu,
        'nombre_ventes': total_ventes['count'] or 0,
        'today': today.strftime("%d/%m/%Y"),
        'entreprise': request.entreprise,
        'devise': request.entreprise.devise_principale.symbole if hasattr(request.entreprise, 'devise_principale') else 'FC'
    }
    
    return render(request, 'vente/cloturer_caisse.html', context)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def liste_ecarts(request):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    ecarts = EcartCaisse.objects.filter(
        entreprise=request.entreprise
    ).select_related('cloture', 'cloture__vendeur').order_by('-date_decouverte')
    
    return render(request, 'vente/liste_ecarts.html', {
        'ecarts': ecarts,
        'entreprise': request.entreprise
    })
    
@login_required
def rapport_cloture(request, pk):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    cloture = get_object_or_404(
        ClotureCaisse.objects.select_related('vendeur'),
        pk=pk,
        entreprise=request.entreprise
    )
    
    return render(request, 'vente/rapport_cloture.html', {
        'cloture': cloture,
        'entreprise': request.entreprise,
        'devise': request.entreprise.devise_principale.symbole if hasattr(request.entreprise, 'devise_principale') else 'FC'
    })

#bila
@login_required
def bilan_du_jour(request):
    vendeur = request.user
    today = timezone.now().date()

    commandes = CommandeClient.objects.filter(
        vente_au_comptoir=True,
        vente_confirmee=True,
        vendeur=vendeur,
        date_commande__date=today
    )

    total = commandes.aggregate(Sum("montant_total"))["montant_total__sum"] or 0
    nb_commandes = commandes.count()

    moyen_commande = total / nb_commandes if nb_commandes > 0 else 0

    derniere_commande = commandes.order_by("-date_commande").first()

    lignes = LigneCommandeClient.objects.filter(commande__in=commandes)

    produits_vendus = {}
    for ligne in lignes:
        produit = ligne.produit
        if produit.id not in produits_vendus:
            produits_vendus[produit.id] = {
                "nom": produit.nom,
                "quantite": 0,
                "montant": 0,
            }
        produits_vendus[produit.id]["quantite"] += ligne.quantite
        produits_vendus[produit.id]["montant"] += ligne.sous_total()

    produits_vendus_list = sorted([
        {
            "nom": infos["nom"],
            "quantite": infos["quantite"],
            "montant": infos["montant"]
        }
        for infos in produits_vendus.values()
    ], key=lambda x: x["montant"], reverse=True)

    context = {
        "vendeur": vendeur,
        "date": today,
        "total": total,
        "nb_commandes": nb_commandes,
        "moyen_commande": moyen_commande,
        "derniere_commande": derniere_commande.date_commande if derniere_commande else None,
        "commandes": commandes,
        "produits_vendus": produits_vendus_list
    }

    return render(request, "vente/bilan_du_jour.html", context)
    
def afficher_bilan(request):
    bilan = bilan_caisse_du_jour(request.user)
    nb = bilan["nb_commandes"]
    total = bilan["total"]
    bilan["moyen_commande"] = total / nb if nb > 0 else 0
    bilan["vendeur"] = request.user
    bilan["date"] = timezone.now().date()
    return render(request, "vente/bilan_du_jour.html", bilan)

    
    
def ticket_caisse(request, commande_id):
    commande = get_object_or_404(Commande, pk=commande_id)
    
    # Récupérer les paramètres de l'entreprise
    try:
        parametre = Parametre.objects.first()
    except Parametre.DoesNotExist:
        # Créer un objet par défaut si aucun paramètre n'existe
        parametre = Parametre(
            nom_societe="Ma Société",
            adresse="Adresse non définie",
            telephone="0000000000",
            email="contact@example.com"
        )
    
    context = {
        'commande': commande,
        'parametre': parametre
    }
    return render(request, 'ticket_caisse.html', context)   
from .models import Parametre

    
def ma_vue(request, commande_id):
    commande = get_object_or_404(Commande, pk=commande_id)
    parametre = Parametre.objects.first()  # ou la méthode appropriée pour récupérer les paramètres
    return render(request, 'template.html', {'commande': commande, 'parametre': parametre})
# def afficher_P(request):
#     parametres = Parametre.objects.first()
#     return render(request, "parametres/parametres.html", {
#         "parA": parametres
#     })

def principal(request):
    parametres = Parametre.objects.first()
    return render(request, "stat/principale.html")



#tableau de bord
def dashboard(request):
    today = timezone.now().date()

    ventes_jour = CommandeClient.objects.filter(date__date=today)
    total_journalier = ventes_jour.aggregate(total=Sum('montant'))['total'] or 0
    total_transactions = ventes_jour.count()

    # Top vendeur
    meilleurs = ventes_jour.values('vendeur__username') \
        .annotate(total=Sum('montant')) \
        .order_by('-total')
    meilleur_vendeur = {'nom': meilleurs[0]['vendeur__username']} if meilleurs else {'nom': 'Aucun'}

    # Produits populaires
    produits_populaires = ventes_jour.values('produit__nom') \
        .annotate(total_vendu=Sum('quantite')) \
        .order_by('-total_vendu')[:5]

    context = {
        'total_journalier': total_journalier,
        'total_transactions': total_transactions,
        'meilleur_vendeur': meilleur_vendeur,
        'produits_populaires': [{'nom': p['produit__nom'], 'total_vendu': p['total_vendu']} for p in produits_populaires],
        'now': timezone.now(),
    }
    return render(request, 'stat/dashboard.html', context)



def dashboard(request):
    return render(request, 'stat/statistiquesVente.html')

# def logout_view(request):
#     logout(request)
#     messages.info(request, "Vous êtes maintenant connécté ")
#     return redirect('login')  # Redirection vers la page de login


# #role

# Fonctions de vérification de rôle
def est_admin(user):
    return user.groups.filter(name='Admin').exists()

def est_caissier(user):
    return user.groups.filter(name='Caissier').exists()

def est_stock(user):
    return user.groups.filter(name='Stock').exists()

# Vue pour les caissiers
@login_required
@user_passes_test(est_caissier)
def vue_pour_caissier(request):
    return render(request, "vente/vente_comptoire.html")

# Vue pour le responsable stock
@login_required
@user_passes_test(est_stock)
def vue_stock(request):
    return render(request, "stock/liste_produits.html")

#vue utilisateurs
def is_admin(user):
    return user.groups.filter(name="Admin").exists()

@user_passes_test(is_admin)
def liste_utilisateurs(request):
    utilisateurs = User.objects.all()
    return render(request, 'utilisateurs/liste.html', {'utilisateurs': utilisateurs})

#modifier article
@user_passes_test(is_admin)
def modifier_utilisateur(request, user_id):
    user = User.objects.get(pk=user_id)
    groupes = Group.objects.all()

    if request.method == "POST":
        selected_group_id = request.POST.get("groupe")
        group = Group.objects.get(id=selected_group_id)
        user.groups.clear()
        user.groups.add(group)
        user.save()
        return redirect('liste_utilisateurs')

    return render(request, 'utilisateurs/modifier.html', {'user': user, 'groupes': groupes})

@user_passes_test(is_admin)
def desactiver_utilisateur(request, user_id):
    user = User.objects.get(pk=user_id)
    user.is_active = False
    user.save()
    return redirect('liste_utilisateurs')

@user_passes_test(is_admin)
def activer_utilisateur(request, user_id):
    user = User.objects.get(pk=user_id)
    user.is_active = True
    user.save()
    return redirect('liste_utilisateurs')


def reset_password(request, user_id):
    user = User.objects.get(pk=user_id)
    nouveau_mot_de_passe = get_random_string(10)

    user.password = make_password(nouveau_mot_de_passe)
    user.save()

    sujet = "Réinitialisation de votre mot de passe"
    message = f"""
Bonjour {user.get_full_name() or user.username},

Votre mot de passe a été réinitialisé.

Nom d'utilisateur : {user.username}
Mot de passe : {nouveau_mot_de_passe}

Merci de changer votre mot de passe après connexion.

Cordialement,
L'équipe support
    """.strip()

    send_mail(sujet, message, settings.DEFAULT_FROM_EMAIL, [user.email])

    messages.success(request, f"Mot de passe réinitialisé et envoyé à {user.email}")
    return redirect('liste_utilisateurs')

#ajouter user
def ajouter_utilisateur(request):
    groupes = Group.objects.all()

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        groupe_id = request.POST.get('groupe')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Ce nom d'utilisateur existe déjà.")
            return redirect('ajouter_utilisateur')

        user = User.objects.create(
            username=username,
            email=email,
            password=make_password(password),
        )

        if groupe_id:
            group = Group.objects.get(id=groupe_id)
            user.groups.add(group)

        messages.success(request, "Utilisateur créé avec succès.")
        return redirect('liste_utilisateurs')

    return render(request, 'utilisateurs/ajouter_utilisateur.html', {'groupes': groupes})

#categorie par commande
def nouvelle_commande(request):
    # Récupérer tous les produits et catégories
    produits = Produit.objects.select_related('categorie').all()
    categories = Categorie.objects.all()
    clients = Client.objects.all()
    
    context = {
        'produits': produits,
        'categories': categories,
        'clients': clients,
    }
    
    return render(request, 'commande/creer_commande.html', context)

@login_required
def tableau_de_bord(request):
    # 1. Vérification et récupération de l'entreprise et configuration SAAS
    entreprise = request.entreprise
    
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=entreprise)
        devise_principale = config_saas.devise_principale
        if not devise_principale:
            messages.error(request, "Aucune devise principale configurée")
            return redirect('tableau_de_bord_erreur')
    except ConfigurationSAAS.DoesNotExist:
        messages.error(request, "Configuration SAAS manquante")
        return redirect('tableau_de_bord_erreur')

    # 2. Période du mois en cours
    aujourdhui = timezone.now()  # Utilisation de timezone pour la compatibilité Django
    debut_mois = aujourdhui.replace(day=1)
    fin_mois = (debut_mois + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # 3. Récupération des données filtrées par entreprise
    produits = Produit.objects.filter(entreprise=entreprise).annotate(
        quantite_vendue=Sum('lignecommande__quantite',
            filter=Q(lignecommande__commande__date_commande__range=[debut_mois, fin_mois],
                    lignecommande__commande__entreprise=entreprise)),  # Filtre SAAS ajouté
        montant_vendu=Sum(F('lignecommande__quantite') * F('lignecommande__prix_unitaire'),
            filter=Q(lignecommande__commande__date_commande__range=[debut_mois, fin_mois],
                    lignecommande__commande__entreprise=entreprise))  # Filtre SAAS ajouté
    ).select_related('categorie').order_by('-montant_vendu')

    # 4. Calcul des totaux avec conversion devise si nécessaire
    total_ventes = sum(p.quantite_vendue or 0 for p in produits)
    total_ca = sum(p.montant_vendu or 0 for p in produits)
    
    # Formatage des montants selon la devise principale
    total_ca_formatted = devise_principale.formater_montant(Decimal(total_ca)) if total_ca else "0.00"

    # 5. Stats par catégorie (filtrées par entreprise)
    categories = Categorie.objects.filter(entreprise=entreprise).annotate(
        total_ventes=Sum('produit__lignecommande__quantite',
            filter=Q(produit__lignecommande__commande__date_commande__range=[debut_mois, fin_mois],
                    produit__lignecommande__commande__entreprise=entreprise)),
        total_ca=Sum(F('produit__lignecommande__quantite') * F('produit__lignecommande__prix_unitaire'),
            filter=Q(produit__lignecommande__commande__date_commande__range=[debut_mois, fin_mois],
                    produit__lignecommande__commande__entreprise=entreprise))
    ).filter(total_ventes__gt=0)

    # 6. Préparation du contexte avec les données formatées
    context = {
        'stats': produits,
        'categories': categories,
        'total_produits': Produit.objects.filter(entreprise=entreprise).count(),
        'total_ventes': total_ventes,
        'total_ca': total_ca,
        'total_ca_formatted': total_ca_formatted,
        'produits_alerte': Produit.objects.filter(entreprise=entreprise, stock__lte=F('seuil_alerte')).count(),
        'mois_courant': debut_mois.strftime("%B %Y"),
        'devise_principale': devise_principale,
        'config_saas': config_saas,
    }
    
    return render(request, 'stat/produits.html', context)
#stat commandes
# views.py - Version complète

from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, Q
from django.db.models.functions import TruncMonth
from django.shortcuts import render, redirect
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.contrib import messages
from decimal import Decimal
from datetime import timedelta

from django.http import Http404
import logging

logger = logging.getLogger(__name__)

@login_required
@permission_required('STOCK.view_commande', raise_exception=True)
def commande_stats(request):
    # 1. Vérification de l'entreprise dans la requête
    if not hasattr(request, 'entreprise') or not request.entreprise:
        raise Http404("Informations d'entreprise manquantes pour l'utilisateur. Accès refusé.")

    entreprise = request.entreprise

    # 2. Récupération de la configuration SAAS
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=entreprise)
        devise_principale = config_saas.devise_principale
        if not devise_principale:
            messages.error(request, "Aucune devise principale configurée pour cette entreprise")
            return redirect('security:acces_refuse')
            
        # Récupération des devises acceptées
        devises_acceptees = [devise_principale.code]  # Par défaut, seule la devise principale
        if hasattr(config_saas, 'devises_acceptees'):  # Si la relation existe
            devises_acceptees += [d.code for d in config_saas.devises_acceptees.all()]
            
    except ConfigurationSAAS.DoesNotExist:
        logger.error(f"Configuration SAAS manquante pour l'entreprise {entreprise.id}")
        messages.error(request, "Erreur de configuration SAAS. Contactez l'administrateur.")
        return redirect('security:acces_refuse')
    except Exception as e:
        logger.error(f"Erreur configuration SAAS entreprise {entreprise.id}: {str(e)}")
        messages.error(request, "Erreur de configuration. Contactez l'administrateur.")
        return redirect('security:acces_refuse')

    # 3. Gestion des filtres et de la devise
    period = request.GET.get('period', 'month')
    status = request.GET.get('status')
    sale_type = request.GET.get('sale_type')
    devise_cible = request.GET.get("devise", request.session.get('devise_affichee', devise_principale.code))

    # Vérification devise autorisée
    if devise_cible not in devises_acceptees:
        devise_cible = devise_principale.code

    # Sauvegarde en session si nouvelle devise valide
    if 'devise' in request.GET and devise_cible in devises_acceptees:
        request.session['devise_affichee'] = devise_cible

    # 4. Récupération du taux de change
    try:
        if devise_principale.code == devise_cible:
            taux_change = Decimal('1.0')
        else:
            taux_change = TauxChange.get_taux(devise_principale.code, devise_cible)
            taux_change = Decimal(str(taux_change)) if taux_change else Decimal('1.0')
    except ValidationError as e:
        logger.error(f"Erreur taux de change {devise_principale.code}->{devise_cible}: {str(e)}")
        messages.warning(request, f"Taux de change non disponible. Affichage en {devise_principale.code}")
        devise_cible = devise_principale.code
        taux_change = Decimal('1.0')

    # 5. Filtrage des commandes par entreprise
    queryset = CommandeClient.objects.filter(entreprise=entreprise)
    
    # Filtres supplémentaires
    if status:
        queryset = queryset.filter(statut=status)
    if sale_type == 'comptoir':
        queryset = queryset.filter(vente_au_comptoir=True)
    elif sale_type == 'normal':
        queryset = queryset.filter(vente_au_comptoir=False)

    # Filtre par période
    today = timezone.now().date()
    if period == 'today':
        queryset = queryset.filter(date_commande__date=today)
    elif period == 'week':
        queryset = queryset.filter(date_commande__date__gte=today - timedelta(days=7))
    elif period == 'month':
        queryset = queryset.filter(date_commande__date__gte=today - timedelta(days=30))
    elif period == 'year':
        queryset = queryset.filter(date_commande__date__gte=today - timedelta(days=365))

    # 6. Calculs statistiques
    total_commandes = queryset.count()
    total_montant_source = queryset.aggregate(total=Sum('montant_total'))['total'] or Decimal('0')
    total_montant = total_montant_source * taux_change
    commandes_annulees = queryset.filter(statut='annulee').count()
    
    # Formatage des montants
    total_montant_formatted = devise_principale.formater_montant(total_montant)
    total_montant_source_formatted = devise_principale.formater_montant(total_montant_source)

    # 7. Données mensuelles pour graphiques
    monthly_data = (
        queryset.annotate(month=TruncMonth('date_commande'))
               .values('month')
               .annotate(
                   total=Count('id'), 
                   amount=Sum('montant_total')
               )
               .order_by('month')
    )

    months = [entry['month'].strftime("%b") for entry in monthly_data]
    monthly_totals = [entry['total'] for entry in monthly_data]
    monthly_amounts = [
        float((Decimal(str(entry['amount'] or 0)) * taux_change) / Decimal('1000'))
        for entry in monthly_data
    ]

    # 8. Top produits (avec conversion devise)
    top_produits = LigneCommandeClient.objects.filter(
        commande__in=queryset,
        commande__entreprise=entreprise
    ).annotate(
        montant_ligne=ExpressionWrapper(
            F('quantite') * F('prix_unitaire'),
            output_field=DecimalField()
        )
    ).values(
        'produit__nom', 'produit__code_barre_numero'
    ).annotate(
        quantite=Sum('quantite'),
        montant=Sum('montant_ligne')
    ).order_by('-quantite')[:5]

    for produit in top_produits:
        produit['montant_converted'] = Decimal(str(produit['montant'] or 0)) * taux_change
        produit['montant_formatted'] = devise_principale.formater_montant(produit['montant_converted'])

    # 9. Top clients (avec conversion devise)
    top_clients = queryset.values('client__nom', 'client__id').annotate(
        commandes=Count('id'),
        montant=Sum('montant_total')
    ).order_by('-montant')[:5]

    for client in top_clients:
        client['montant_converted'] = Decimal(str(client['montant'] or 0)) * taux_change
        client['montant_formatted'] = devise_principale.formater_montant(client['montant_converted'])

    # 10. Commandes récentes
    recent_orders = queryset.select_related('client').order_by('-date_commande')[:10]
    for commande in recent_orders:
        commande.montant_formatted = devise_principale.formater_montant(commande.montant_total * taux_change)

    # 11. Préparation du contexte
    context = {
        'entreprise': entreprise,
        'config_saas': config_saas,
        'devise_principale': devise_principale,
        'devise_cible': devise_cible,
        'devises_acceptees': devises_acceptees,
        'today': today,
        'total_commandes': total_commandes,
        'total_montant': total_montant,
        'total_montant_formatted': total_montant_formatted,
        'total_montant_source': total_montant_source,
        'total_montant_source_formatted': total_montant_source_formatted,
        'commandes_annulees': commandes_annulees,
        'statuts': queryset.values('statut').annotate(total=Count('id')).order_by('-total'),
        'monthly_labels': months,
        'monthly_totals': monthly_totals,
        'monthly_amounts': monthly_amounts,
        'taux_change': taux_change,
        'top_clients': top_clients,
        'top_produits': top_produits,
        'recent_orders': recent_orders,
        'active_filters': {
            'period': period,
            'status': status,
            'sale_type': sale_type,
            'devise': devise_cible
        }
    }

    return render(request, 'stat/commande_stats.html', context)

def inventaire(request):
    return render(request,"inventaire/inventaire.html")

# Dans votre fichier views.py (corrigé)
from django.db.models import F
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Produit, Categorie
from parametres.models import ConfigurationSAAS

@login_required
def etat_stock(request):
    # Vérification de l'entreprise
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie.")
        return redirect('security:acces_refuse')

    # Récupération de la devise principale
    devise_principale = 'USD'  # Devise par défaut
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale.symbole
    except ConfigurationSAAS.DoesNotExist:
        pass

    # Requête de base pour les produits
    produits = Produit.objects.filter(entreprise=request.entreprise)
    categories = Categorie.objects.filter(entreprise=request.entreprise)

    # Filtres
    cat = request.GET.get('categorie')
    seuil = request.GET.get('alerte')

    # Filtrage par catégorie
    if cat:
        produits = produits.filter(categorie_id=cat)
        
    # Filtrage par seuil d'alerte
    if seuil == "1":
        produits = produits.filter(stock__lte=F('seuil_alerte'))

    context = {
        'produits': produits,
        'categories': categories,
        'cat_selected': int(cat) if cat else None,
        'seuil_selected': seuil,
        'entreprise': request.entreprise,
        'devise_principale': devise_principale
    }
    return render(request, 'inventaire/etat_stock.html', context)



# In STOCK/views.py
# Dans votre fichier STOCK/views.py

import csv
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import F
from .models import Produit, Categorie
from parametres.models import ConfigurationSAAS

def get_devise_principale(request):
    """Récupère la devise principale de l'entreprise."""
    devise_principale = 'USD'  # Devise par défaut
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale.symbole
    except (ConfigurationSAAS.DoesNotExist, AttributeError):
        pass
    return devise_principale

@login_required
def etat_stock(request):
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie.")
        return redirect('security:acces_refuse')

    produits = Produit.objects.filter(entreprise=request.entreprise)
    categories = Categorie.objects.filter(entreprise=request.entreprise)
    
    cat = request.GET.get('categorie')
    seuil = request.GET.get('alerte')

    if cat:
        produits = produits.filter(categorie_id=cat)
    if seuil == "1":
        produits = produits.filter(stock__lte=F('seuil_alerte'))

    context = {
        'produits': produits,
        'categories': categories,
        'cat_selected': int(cat) if cat else None,
        'seuil_selected': seuil,
        'entreprise': request.entreprise,
        'devise_principale': get_devise_principale(request)
    }
    return render(request, 'inventaire/etat_stock.html', context)
# In STOCK/views.py

import csv
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import F
from .models import Produit, Categorie
from parametres.models import Entreprise # Make sure this import path is correct
from django.conf import settings
import os

def get_devise_principale(request):
    """Retrieves the main currency of the company."""
    # (Existing function, no changes needed)
    devise_principale = 'USD'
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale.symbole
    except (ConfigurationSAAS.DoesNotExist, AttributeError):
        pass
    return devise_principale

@login_required
def exporter_stock(request):
    if not hasattr(request, 'entreprise'):
        return redirect('security:acces_refuse')

    entreprise = request.entreprise
    format_type = request.GET.get('format', 'csv')
    produits = Produit.objects.filter(entreprise=request.entreprise)

    headers = ['Nom', 'Catégorie', 'Stock', 'Seuil d\'alerte', 'Prix de vente']

    # --- CSV Export (No Design Changes) ---
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="etat_stock.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for produit in produits:
            writer.writerow([
                produit.nom,
                produit.categorie.nom if produit.categorie else 'N/A',
                produit.stock,
                produit.seuil_alerte,
                produit.prix_vente
            ])
        return response

    # --- XLSX Export (Improved Design) ---
    elif format_type == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="etat_stock.xlsx"'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Company Header
        sheet['B2'] = 'Rapport d\'État du Stock'
        sheet['B3'] = f"Entreprise: {entreprise.nom}"
        sheet['B4'] = f"Adresse: {entreprise.adresse}"
        
        # Add a blank row for spacing
        sheet.append([])
        
        # Headers with bold styling
        header_row = sheet.append(headers)
        for cell in sheet[sheet.max_row]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Data rows with light alternating color
        for i, produit in enumerate(produits):
            row_data = [
                produit.nom,
                produit.categorie.nom if produit.categorie else 'N/A',
                produit.stock,
                produit.seuil_alerte,
                produit.prix_vente
            ]
            sheet.append(row_data)
            if i % 2 == 0:
                for cell in sheet[sheet.max_row]:
                    cell.fill = openpyxl.styles.PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

        workbook.save(response)
        return response

    # --- PDF Export (Improved Design with ReportLab) ---
    elif format_type == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="etat_stock.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Company Header
        company_logo_path = entreprise.logo.path if entreprise.logo else None
        if company_logo_path and os.path.exists(company_logo_path):
            story.append(Image(company_logo_path, width=1.5*inch, height=0.5*inch))
        
        story.append(Paragraph(f"<font size=14><b>{entreprise.nom}</b></font>", styles['Normal']))
        story.append(Paragraph(f"<font size=10>{entreprise.adresse}</font>", styles['Normal']))
        story.append(Spacer(1, 0.25*inch))
        
        # Report Title
        report_title_style = ParagraphStyle('ReportTitle', parent=styles['Normal'], fontSize=16, alignment=1, spaceAfter=0.25*inch, fontName='Helvetica-Bold')
        story.append(Paragraph("RAPPORT D'ÉTAT DU STOCK", report_title_style))
        story.append(Spacer(1, 0.25*inch))

        # Table Data
        data = [headers]
        for produit in produits:
            data.append([
                produit.nom,
                produit.categorie.nom if produit.categorie else 'N/A',
                produit.stock,
                produit.seuil_alerte,
                produit.prix_vente
            ])
        
        # Table Styling
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#bdbdbd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ])
        
        # Apply alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fafafa'))
        
        table = Table(data)
        table.setStyle(table_style)
        
        story.append(table)
        
        doc.build(story)
        return response

    return HttpResponse("Format non supporté.", status=400)

from django.db import transaction
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.shortcuts import render, redirect
import logging
from comptabilite.models import *

logger = logging.getLogger(__name__)

from django.db import transaction
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.shortcuts import render, redirect
import logging

logger = logging.getLogger(__name__)

@login_required
@permission_required('STOCK.add_mouvementstock', raise_exception=True)
def ajouter_mouvement_stock(request):
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    produits = Produit.objects.filter(entreprise=request.entreprise)

    if request.method == "POST":
        produit_id = request.POST.get("produit")
        type_mouvement = request.POST.get("type")
        quantite = int(request.POST.get("quantite", 0))
        commentaire = request.POST.get("commentaire", "")

        try:
            with transaction.atomic():
                produit = Produit.objects.get(id=produit_id, entreprise=request.entreprise)

                if type_mouvement == "sortie" and quantite > produit.stock:
                    messages.error(request, "Stock insuffisant pour cette sortie.")
                    return redirect("ajouter_mouvement_stock")

                # CORRECTION: Utiliser prix_achat pour la valorisation comptable
                valeur_mouvement = quantite * produit.prix_achat
                
                mouvement = MouvementStock.objects.create(
                    produit=produit,
                    type_mouvement=type_mouvement,
                    quantite=quantite,
                    utilisateur=request.user,
                    commentaire=commentaire,
                    entreprise=request.entreprise,
                    prix_unitaire_moment=produit.prix_achat  # CORRECTION: prix_achat
                )

                # Mise à jour du stock
                if type_mouvement == "entree":
                    produit.stock += quantite
                elif type_mouvement == "sortie":
                    produit.stock -= quantite
                produit.save()

                # Créer l'écriture comptable
                create_ecriture_comptable_mouvement(mouvement, valeur_mouvement, request)

                messages.success(request, "Mouvement enregistré avec succès.")
                return redirect("ajouter_mouvement_stock")

        except Exception as e:
            logger.error(f"Erreur lors de l'ajout du mouvement de stock: {e}")
            messages.error(request, f"Erreur: {str(e)}")

    return render(request, "inventaire/ajouter_mouvement.html", {
        "produits": produits,
        "entreprise": request.entreprise
    })

def create_ecriture_comptable_mouvement(mouvement, valeur_mouvement, request):
    """
    Crée une écriture comptable pour un mouvement de stock
    """
    try:
        from comptabilite.models import JournalComptable, PlanComptableOHADA, EcritureComptable, LigneEcriture
        
        entreprise = request.entreprise
        produit = mouvement.produit
        
        # Vérifier et initialiser le plan comptable si nécessaire
        if not PlanComptableOHADA.objects.filter(entreprise=entreprise).exists():
            PlanComptableOHADA.initialiser_plan_comptable(entreprise)
        
        if not JournalComptable.objects.filter(entreprise=entreprise).exists():
            JournalComptable.initialiser_journaux(entreprise)
        
        # Déterminer le journal en fonction du type de mouvement
        if mouvement.type_mouvement == "entree":
            journal_code = 'STK'
            journal_libelle = 'Journal des Stocks'
            libelle_ecriture = f"Entrée stock - {produit.nom}"
        else:
            journal_code = 'STK'
            journal_libelle = 'Journal des Stocks'
            libelle_ecriture = f"Sortie stock - {produit.nom}"
        
        # Journal des stocks
        journal, created = JournalComptable.objects.get_or_create(
            code=journal_code,
            entreprise=entreprise,
            defaults={
                'intitule': journal_libelle,
                'type_journal': 'divers'
            }
        )
        
        # Récupérer les comptes comptables
        compte_stock, created = PlanComptableOHADA.objects.get_or_create(
            numero='31',  # Stocks
            entreprise=entreprise,
            defaults={
                'classe': '3',
                'intitule': 'Stocks',
                'type_compte': 'actif',
                'description': 'Stocks de matières et marchandises'
            }
        )
        
        # Compte de variation de stock (classe 60)
        compte_variation_stock, created = PlanComptableOHADA.objects.get_or_create(
            numero='603',  # Variation des stocks
            entreprise=entreprise,
            defaults={
                'classe': '6',
                'intitule': 'Variation des stocks',
                'type_compte': 'charge',
                'description': 'Variation des stocks de marchandises'
            }
        )
        
        # Compte de production (pour les sorties de stock)
        compte_production, created = PlanComptableOHADA.objects.get_or_create(
            numero='71',  # Production stockée
            entreprise=entreprise,
            defaults={
                'classe': '7',
                'intitule': 'Production stockée',
                'type_compte': 'produit',
                'description': 'Production stockée et déstockée'
            }
        )
        
        # Générer un numéro d'écriture
        numero_ecriture = generate_ecriture_number(journal)
        
        # Créer l'écriture comptable
        ecriture = EcritureComptable.objects.create(
            journal=journal,
            numero=numero_ecriture,
            date_ecriture=mouvement.date_mouvement,
            date_comptable=mouvement.date_mouvement,
            libelle=libelle_ecriture,
            piece_justificative=f"MVT-{mouvement.id}",
            montant_devise=valeur_mouvement,
            entreprise=entreprise,
            created_by=request.user,
            mouvement_stock_lie=mouvement
        )
        
        if mouvement.type_mouvement == "entree":
            # ENTREE DE STOCK
            # Débit: Stock (augmentation de l'actif)
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_stock,
                libelle=f"Entrée stock - {produit.nom}",
                debit=valeur_mouvement,
                credit=0,
                entreprise=entreprise
            )
            
            # Crédit: Variation de stock (charge)
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_variation_stock,
                libelle=f"Variation stock - {produit.nom}",
                debit=0,
                credit=valeur_mouvement,
                entreprise=entreprise
            )
            
        else:
            # SORTIE DE STOCK
            # Débit: Variation de stock (charge)
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_variation_stock,
                libelle=f"Variation stock - {produit.nom}",
                debit=valeur_mouvement,
                credit=0,
                entreprise=entreprise
            )
            
            # Crédit: Stock (diminution de l'actif)
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_stock,
                libelle=f"Sortie stock - {produit.nom}",
                debit=0,
                credit=valeur_mouvement,
                entreprise=entreprise
            )
        
        logger.info(f"Écriture comptable {numero_ecriture} créée pour le mouvement de stock {mouvement.id}")
        
        return ecriture
        
    except ImportError:
        logger.warning("Module comptabilité non installé - écriture non créée")
        return None
    except Exception as e:
        logger.error(f"Erreur création écriture comptable mouvement: {e}", exc_info=True)
        return None

def generate_ecriture_number(journal):
    """Génère un numéro d'écriture séquentiel"""
    from django.utils import timezone
    
    today = timezone.now().date()
    last_ecriture = EcritureComptable.objects.filter(
        journal=journal,
        date_ecriture=today
    ).order_by('-numero').first()
    
    sequence = 1
    if last_ecriture:
        try:
            parts = last_ecriture.numero.split('-')
            if len(parts) >= 3:
                sequence = int(parts[-1]) + 1
        except (ValueError, IndexError):
            pass
    
    return f"{journal.code}-{today.strftime('%Y%m%d')}-{sequence:04d}"
from django.contrib.auth import get_user_model
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator

@login_required
@permission_required('STOCK.view_mouvementstock', raise_exception=True)
def liste_mouvements_stock(request):
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    # Récupérer tous les mouvements de l'entreprise
    mouvements = MouvementStock.objects.filter(
        entreprise=request.entreprise
    ).select_related('produit', 'utilisateur').order_by('-date_mouvement')

    # Filtres
    produit_id = request.GET.get("produit")
    type_mouvement = request.GET.get("type")
    utilisateur_id = request.GET.get("utilisateur")
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    # Appliquer les filtres
    if produit_id and produit_id != '':
        mouvements = mouvements.filter(produit_id=produit_id)
    if type_mouvement and type_mouvement != '':
        mouvements = mouvements.filter(type_mouvement=type_mouvement)
    if utilisateur_id and utilisateur_id != '':
        mouvements = mouvements.filter(utilisateur_id=utilisateur_id)
    if date_debut:
        try:
            date_debut_obj = parse_date(date_debut)
            if date_debut_obj:
                mouvements = mouvements.filter(date_mouvement__date__gte=date_debut_obj)
        except (ValueError, TypeError):
            pass
    if date_fin:
        try:
            date_fin_obj = parse_date(date_fin)
            if date_fin_obj:
                mouvements = mouvements.filter(date_mouvement__date__lte=date_fin_obj)
        except (ValueError, TypeError):
            pass

    # Pagination
    paginator = Paginator(mouvements, 25)  # 25 éléments par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Données pour les filtres
    produits = Produit.objects.filter(entreprise=request.entreprise)
    utilisateurs = get_user_model().objects.filter(
        mouvementstock__entreprise=request.entreprise
    ).distinct()

 # Récupérer la devise
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_symbole = config_saas.devise_principale.symbole if config_saas.devise_principale else "€"
    except ConfigurationSAAS.DoesNotExist:
        devise_symbole = "€"
        
    context = {
        'mouvements': page_obj,  # Utiliser l'objet paginé
        'produits': produits,
        'utilisateurs': utilisateurs,
        'filtres': request.GET.dict(),  # Conserver tous les paramètres de filtre
        'total_mouvements': mouvements.count(),
        'entreprise': request.entreprise,
        'devise_symbole': devise_symbole
    }
    return render(request, "inventaire/liste_mouvements.html", context)

# 🔢 Saisie de l'inventaire physique
@login_required
@permission_required('STOCK.add_inventairephysique', raise_exception=True)
def saisie_inventaire(request):
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    produits = Produit.objects.filter(entreprise=request.entreprise)

    if request.method == "POST":
        try:
            with transaction.atomic():
                for produit in produits:
                    field_name = f"physique_{produit.id}"
                    stock_physique = int(request.POST.get(field_name, 0))
                    stock_theorique = produit.stock
                    ecart = stock_physique - stock_theorique

                    InventairePhysique.objects.create(
                        produit=produit,
                        stock_theorique=stock_theorique,
                        stock_physique=stock_physique,
                        ecart=ecart,
                        utilisateur=request.user,
                        entreprise=request.entreprise
                    )

                messages.success(request, "Inventaire enregistré avec succès.")
                return redirect("liste_inventaires")

        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")

    return render(request, "inventaire/saisir_inventaire.html", {
        "produits": produits,
        "entreprise": request.entreprise
    })

@login_required
@permission_required('STOCK.change_inventairephysique', raise_exception=True)
def liste_inventaires(request):
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    inventaires = InventairePhysique.objects.filter(
        entreprise=request.entreprise,
        valide=False
    ).select_related("produit")

    if request.method == "POST":
        try:
            with transaction.atomic():
                for inv in inventaires:
                    produit = inv.produit
                    produit.stock = inv.stock_physique
                    produit.save()

                    inv.valide = True
                    inv.save()

                    MouvementStock.objects.create(
                        produit=produit,
                        type_mouvement="ajustement",
                        quantite=abs(inv.ecart),
                        commentaire=f"Ajustement inventaire (écart: {inv.ecart:+})",
                        utilisateur=request.user,
                        entreprise=request.entreprise
                    )

                messages.success(request, "Stocks ajustés avec succès.")
                return redirect("liste_inventaires")

        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")

    return render(request, "inventaire/liste_inventaires.html", {
        "inventaires": inventaires,
        "entreprise": request.entreprise
    })
    
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib.auth import get_user_model
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from STOCK.models import Produit, MouvementStock, InventairePhysique
import logging

logger = logging.getLogger(__name__)
@require_POST
@login_required
def valider_inventaire(request):
    """
    Valide l'inventaire physique et crée les écritures comptables
    """
    try:
        logger.info(f"=== DÉBUT VALIDATION INVENTAIRE ===")
        logger.info(f"Utilisateur: {request.user.username}")
        logger.info(f"Entreprise: {request.entreprise.nom if hasattr(request, 'entreprise') else 'Non définie'}")
        logger.info(f"Données POST: {dict(request.POST)}")
        
        if not hasattr(request, 'entreprise'):
            logger.error("ERREUR: Aucune entreprise définie dans la request")
            messages.error(request, "Entreprise non définie")
            return redirect("liste_inventaires")
        
        with transaction.atomic():
            inventaires_crees = []
            erreurs_produits = []
            produits_traites = 0
            
            for key, value in request.POST.items():
                if key.startswith("quantite_physique_"):
                    produits_traites += 1
                    produit_id = key.split("_")[-1]
                    logger.info(f"Traitement produit ID: {produit_id}, valeur: {value}")
                    
                    try:
                        produit = Produit.objects.get(id=produit_id, entreprise=request.entreprise)
                        quantite_physique = int(value)
                        ecart = quantite_physique - produit.stock

                        logger.info(f"Produit {produit.nom}: Stock théorique={produit.stock}, Physique={quantite_physique}, Écart={ecart}")

                        if ecart != 0:
                            ancien_stock = produit.stock
                            
                            # Créer d'abord le mouvement de stock
                            mouvement = MouvementStock.objects.create(
                                produit=produit,
                                type_mouvement="ajustement",
                                quantite=abs(ecart),
                                commentaire=f"Ajustement inventaire: {ancien_stock} → {quantite_physique} (écart: {ecart:+})",
                                utilisateur=request.user,
                                entreprise=request.entreprise,
                                prix_unitaire_moment=produit.prix_achat
                            )
                            logger.info(f"Mouvement de stock créé: ID {mouvement.id}")

                            # Mettre à jour le produit
                            produit.stock = quantite_physique
                            produit.save()
                            logger.info(f"Stock du produit {produit.nom} mis à jour: {ancien_stock} → {quantite_physique}")

                            # Créer l'écriture comptable
                            logger.info(f"Appel create_ecriture_comptable_inventaire...")
                            ecriture = create_ecriture_comptable_inventaire(mouvement, abs(ecart), ecart, request)
                            
                            if ecriture:
                                inventaires_crees.append({
                                    'produit': produit.nom,
                                    'ecart': ecart,
                                    'ecriture': ecriture.numero
                                })
                                logger.info(f"SUCCÈS: Écriture comptable créée: {ecriture.numero}")
                            else:
                                logger.warning(f"ÉCHEC: Aucune écriture comptable créée pour le produit {produit.nom}")

                    except Produit.DoesNotExist:
                        erreur_msg = f"Produit avec ID {produit_id} non trouvé"
                        logger.warning(erreur_msg)
                        erreurs_produits.append(erreur_msg)
                    except ValueError as e:
                        erreur_msg = f"Valeur invalide pour le produit {produit_id}: {value} - {e}"
                        logger.warning(erreur_msg)
                        erreurs_produits.append(erreur_msg)
                    except Exception as e:
                        erreur_msg = f"Erreur avec le produit {produit_id}: {str(e)}"
                        logger.error(erreur_msg, exc_info=True)
                        erreurs_produits.append(erreur_msg)

            logger.info(f"Produits traités: {produits_traites}")
            logger.info(f"Ajustements créés: {len(inventaires_crees)}")
            logger.info(f"Erreurs: {len(erreurs_produits)}")

            # Gestion des messages de résultat
            if inventaires_crees:
                success_msg = f"Inventaire validé: {len(inventaires_crees)} ajustements comptabilisés."
                logger.info(success_msg)
                messages.success(request, success_msg)
                
                for ajustement in inventaires_crees:
                    logger.info(f"Ajustement - Produit: {ajustement['produit']}, Écart: {ajustement['ecart']}, Écriture: {ajustement['ecriture']}")
                    
            else:
                info_msg = "Inventaire validé - aucun écart significatif."
                logger.info(info_msg)
                messages.info(request, info_msg)
                
            if erreurs_produits:
                for erreur in erreurs_produits:
                    logger.warning(f"Erreur produit: {erreur}")
                    messages.warning(request, f"Erreur: {erreur}")
                
            logger.info("=== FIN VALIDATION INVENTAIRE ===")
            return redirect("liste_inventaires")

    except Exception as e:
        error_msg = f"Erreur critique lors de la validation de l'inventaire: {str(e)}"
        logger.error(error_msg, exc_info=True)
        messages.error(request, f"Erreur lors de la validation: {str(e)}")
        return redirect("liste_inventaires")
def create_ecriture_comptable_inventaire(mouvement, quantite_ajustement, ecart, request):
    """
    Crée une écriture comptable pour un ajustement d'inventaire
    """
    try:
        logger.info(f"=== DÉBUT CRÉATION ÉCRITURE COMPTABLE ===")
        logger.info(f"Mouvement ID: {mouvement.id}, Produit: {mouvement.produit.nom}")
        
        # Import des modèles comptables avec vérification
        try:
            from comptabilite.models import JournalComptable, PlanComptableOHADA, EcritureComptable, LigneEcriture
            logger.info("✓ Import des modèles comptables réussi")
        except ImportError as e:
            logger.error(f"✗ Erreur import modèles comptables: {e}")
            return None
        
        # Vérifier que l'utilisateur a les permissions
        if not hasattr(request.user, 'has_perm') or not request.user.has_perm('comptabilite.add_ecriturecomptable'):
            logger.warning("✗ Utilisateur sans permission pour créer des écritures")
            return None
        
        entreprise = request.entreprise
        produit = mouvement.produit
        
        # Calcul de la valeur
        valeur_ajustement = quantite_ajustement * mouvement.prix_unitaire_moment
        logger.info(f"Valeur ajustement: {valeur_ajustement}")
        
        # Déterminer le type d'ajustement
        libelle_ecriture = f"{'Excédent' if ecart > 0 else 'Déficit'} inventaire - {produit.nom}"
        
        # Journal des stocks - création si nécessaire
        journal, created = JournalComptable.objects.get_or_create(
            code='STK',
            entreprise=entreprise,
            defaults={
                'intitule': 'Journal des Stocks',
                'type_journal': 'divers'
            }
        )
        
        # Comptes comptables - création si nécessaire
        compte_stock, created = PlanComptableOHADA.objects.get_or_create(
            numero='31',
            entreprise=entreprise,
            defaults={
                'classe': '3',
                'intitule': 'Stocks',
                'type_compte': 'actif',
                'description': 'Stocks de matières et marchandises'
            }
        )
        
        compte_variation, created = PlanComptableOHADA.objects.get_or_create(
            numero='6037',
            entreprise=entreprise,
            defaults={
                'classe': '6',
                'intitule': 'Variation des stocks (inventaires)',
                'type_compte': 'charge',
                'description': 'Ajustements de stocks suite aux inventaires'
            }
        )
        
        # Générer un numéro d'écriture simple
        from django.utils import timezone
        numero_ecriture = f"STK-{timezone.now().strftime('%Y%m%d')}-{EcritureComptable.objects.filter(journal=journal, date_ecriture__date=timezone.now().date()).count() + 1:04d}"
        
        # Créer l'écriture comptable
        ecriture = EcritureComptable.objects.create(
            journal=journal,
            numero=numero_ecriture,
            date_ecriture=timezone.now(),
            date_comptable=timezone.now().date(),
            libelle=libelle_ecriture,
            piece_justificative=f"INV-{mouvement.id}",
            montant_devise=valeur_ajustement,
            entreprise=entreprise,
            created_by=request.user,
            mouvement_stock_lie=mouvement
        )
        
        # Créer les lignes d'écriture
        if ecart > 0:  # Excédent
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_stock,
                libelle=f"Excédent inventaire - {produit.nom}",
                debit=valeur_ajustement,
                credit=0,
                entreprise=entreprise
            )
            
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_variation,
                libelle=f"Excédent inventaire - {produit.nom}",
                debit=0,
                credit=valeur_ajustement,
                entreprise=entreprise
            )
        else:  # Déficit
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_variation,
                libelle=f"Déficit inventaire - {produit.nom}",
                debit=valeur_ajustement,
                credit=0,
                entreprise=entreprise
            )
            
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_stock,
                libelle=f"Déficit inventaire - {produit.nom}",
                debit=0,
                credit=valeur_ajustement,
                entreprise=entreprise
            )
        
        logger.info(f"✓ Écriture créée avec succès: {ecriture.numero}")
        return ecriture
        
    except Exception as e:
        logger.error(f"✗ Erreur création écriture: {e}", exc_info=True)
        return None
        
        # Générer un numéro d'écriture
        try:
            numero_ecriture = generate_ecriture_number(journal)
            logger.info(f"Numéro d'écriture généré: {numero_ecriture}")
        except Exception as e:
            logger.error(f"✗ Erreur génération numéro: {e}", exc_info=True)
            numero_ecriture = f"STK-{mouvement.id}-ERR"
        
        # Créer l'écriture comptable
        try:
            ecriture = EcritureComptable.objects.create(
                journal=journal,
                numero=numero_ecriture,
                date_ecriture=mouvement.date_mouvement,
                date_comptable=mouvement.date_mouvement.date(),
                libelle=libelle_ecriture,
                piece_justificative=f"INV-{mouvement.id}",
                montant_devise=valeur_ajustement,
                entreprise=entreprise,
                created_by=request.user,
                mouvement_stock_lie=mouvement
            )
            logger.info(f"✓ Écriture comptable créée: {ecriture.numero}")
        except Exception as e:
            logger.error(f"✗ Erreur création écriture: {e}", exc_info=True)
            return None
        
        # Créer les lignes d'écriture
        try:
            if type_ajustement == "excédent":
                # EXCÉDENT D'INVENTAIRE
                ligne1 = LigneEcriture.objects.create(
                    ecriture=ecriture,
                    compte=compte_stock,
                    libelle=f"Excédent inventaire - {produit.nom}",
                    debit=valeur_ajustement,
                    credit=0,
                    entreprise=entreprise
                )
                
                ligne2 = LigneEcriture.objects.create(
                    ecriture=ecriture,
                    compte=compte_variation,
                    libelle=f"Excédent inventaire - {produit.nom}",
                    debit=0,
                    credit=valeur_ajustement,
                    entreprise=entreprise
                )
                logger.info(f"✓ Lignes créées: Débit {ligne1.debit} sur {compte_stock.numero}, Crédit {ligne2.credit} sur {compte_variation.numero}")
                
            else:
                # DÉFICIT D'INVENTAIRE
                ligne1 = LigneEcriture.objects.create(
                    ecriture=ecriture,
                    compte=compte_variation,
                    libelle=f"Déficit inventaire - {produit.nom}",
                    debit=valeur_ajustement,
                    credit=0,
                    entreprise=entreprise
                )
                
                ligne2 = LigneEcriture.objects.create(
                    ecriture=ecriture,
                    compte=compte_stock,
                    libelle=f"Déficit inventaire - {produit.nom}",
                    debit=0,
                    credit=valeur_ajustement,
                    entreprise=entreprise
                )
                logger.info(f"✓ Lignes créées: Débit {ligne1.debit} sur {compte_variation.numero}, Crédit {ligne2.credit} sur {compte_stock.numero}")
                
        except Exception as e:
            logger.error(f"✗ Erreur création lignes: {e}", exc_info=True)
            # Supprimer l'écriture incomplète
            try:
                ecriture.delete()
                logger.info("Écriture supprimée à cause d'erreur de lignes")
            except:
                pass
            return None
        
        # Mise à jour des documents comptables
        try:
            update_grand_livre(ecriture, entreprise)
            logger.debug("✓ Grand Livre mis à jour")
            
            update_balance(entreprise, mouvement.date_mouvement.date())
            logger.debug("✓ Balance mise à jour")
            
            update_bilan(entreprise, mouvement.date_mouvement.date())
            logger.debug("✓ Bilan mis à jour")
            
        except Exception as e:
            logger.warning(f"⚠ Erreur mise à jour documents comptables: {e}")
        
        logger.info(f"=== SUCCÈS: Écriture {ecriture.numero} créée ===")
        return ecriture
        
    except Exception as e:
        logger.error(f"✗ ERREUR CRITIQUE création écriture: {e}", exc_info=True)
        return None

@login_required
def liste_mouvements_stock(request):
    """
    Liste des mouvements de stock avec filtres
    """
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    mouvements = MouvementStock.objects.filter(
        entreprise=request.entreprise
    ).select_related('produit', 'utilisateur').order_by('-date_mouvement')

    # Filtres
    produit_id = request.GET.get("produit")
    type_mouvement = request.GET.get("type")
    utilisateur_id = request.GET.get("utilisateur")
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    # Appliquer les filtres
    if produit_id and produit_id != '':
        mouvements = mouvements.filter(produit_id=produit_id)
    if type_mouvement and type_mouvement != '':
        mouvements = mouvements.filter(type_mouvement=type_mouvement)
    if utilisateur_id and utilisateur_id != '':
        mouvements = mouvements.filter(utilisateur_id=utilisateur_id)
    if date_debut:
        try:
            date_debut_obj = parse_date(date_debut)
            if date_debut_obj:
                mouvements = mouvements.filter(date_mouvement__date__gte=date_debut_obj)
        except (ValueError, TypeError):
            pass
    if date_fin:
        try:
            date_fin_obj = parse_date(date_fin)
            if date_fin_obj:
                mouvements = mouvements.filter(date_mouvement__date__lte=date_fin_obj)
        except (ValueError, TypeError):
            pass

    # Pagination
    paginator = Paginator(mouvements, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Données pour les filtres
    produits = Produit.objects.filter(entreprise=request.entreprise)
    utilisateurs = get_user_model().objects.filter(
        mouvementstock__entreprise=request.entreprise
    ).distinct()

    context = {
        'mouvements': page_obj,
        'produits': produits,
        'utilisateurs': utilisateurs,
        'filtres': request.GET.dict(),
        'total_mouvements': mouvements.count(),
        'entreprise': request.entreprise
    }
    
    logger.debug(f"Liste mouvements affichée: {len(page_obj)} résultats")
    return render(request, "inventaire/liste_mouvements.html", context)
    

#rapport mouvement
@login_required
@permission_required('STOCK.view_report', raise_exception=True)
def rapport_mouvements(request):
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    mouvements = MouvementStock.objects.filter(
        entreprise=request.entreprise
    ).select_related('produit', 'utilisateur').order_by('-date_mouvement')
    # Filtres
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")
    type_mouvement = request.GET.get("type")

    if date_debut:
        mouvements = mouvements.filter(date_mouvement__date__gte=parse_date(date_debut))
    if date_fin:
        mouvements = mouvements.filter(date_mouvement__date__lte=parse_date(date_fin))
    if type_mouvement:
        mouvements = mouvements.filter(type_mouvement=type_mouvement)

    # Export CSV
    if request.GET.get("export") == "csv":
        return export_mouvements_csv(mouvements)
    
    # Impression spécifique
    if request.GET.get("print") == "1":
        context = {
            "mouvements": mouvements,
            "date_debut": date_debut,
            "date_fin": date_fin,
            "type_mouvement": type_mouvement,
            "print_mode": True  # Nouveau contexte pour le mode impression
        }
        return render(request, "rapports/rapport_mouvements_print.html", context)

    context = {
        "mouvements": mouvements,
        "date_debut": date_debut,
        "date_fin": date_fin,
        "type_mouvement": type_mouvement,
    }
    return render(request, "rapports/rapport_mouvements.html", context)

def export_mouvements_csv(mouvements):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=mouvements_stock.csv"

    writer = csv.writer(response)
    writer.writerow(["Date", "Produit", "Type", "Quantité", "Commentaire", "Utilisateur"])

    for m in mouvements:
        writer.writerow([
            m.date_mouvement.strftime("%Y-%m-%d %H:%M"),
            m.produit.nom,
            m.get_type_mouvement_display(),
            m.quantite,
            m.commentaire,
            m.utilisateur.username,
        ])

    return response

#rapport d'alert

@login_required
def rapport_alertes_pdf(request):
    if not hasattr(request, 'entreprise'):
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    produits_alertes = Produit.objects.filter(
        entreprise=request.entreprise,
        stock__lte=F('seuil_alerte')
    )

    template = get_template("rapports/rapport_alertes_pdf.html")
    html = template.render({
        "produits_alertes": produits_alertes,
        "entreprise": request.entreprise,
        "user": request.user
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="produits_en_alerte.pdf"'
    pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=response)
    return response
#ecart rapport

def generate_ecriture_number(journal):
    """Génère un numéro d'écriture séquentiel"""
    from django.utils import timezone
    
    today = timezone.now().date()
    last_ecriture = EcritureComptable.objects.filter(
        journal=journal,
        date_ecriture=today
    ).order_by('-numero').first()
    
    sequence = 1
    if last_ecriture:
        try:
            parts = last_ecriture.numero.split('-')
            if len(parts) >= 3:
                sequence = int(parts[-1]) + 1
        except (ValueError, IndexError):
            pass
    
    return f"{journal.code}-{today.strftime('%Y%m%d')}-{sequence:04d}"

@login_required
def rapport_ecarts_inventaire_pdf(request):
    ecarts = InventairePhysique.objects.filter(ecart__isnull=False).order_by('-date')
    template = get_template("rapports/rapport_ecarts_pdf.html")
    context = {
        "ecarts": ecarts,
        "user": request.user,
        "now": now(),
    }
    html_content = template.render(context)

    buffer = BytesIO()
    HTML(string=html_content).write_pdf(buffer)
    pdf_content = buffer.getvalue()

    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="rapport_ecarts_inventaire.pdf"'
    response['Content-Length'] = len(pdf_content)
    return response

#liste cloture
@login_required
def clotures_du_jour(request):
    today = timezone.now().date()
    clotures = ClotureCaisse.objects.filter(date_jour=today)

    return render(request, "cloture/clotures_du_jour.html", {
        "clotures": clotures,
        "aujourd_hui": today
    })

#pdf cloture
def telecharger_rapport_cloture_pdf(request):
    cloture = ClotureCaisse.objects.filter(vendeur=request.user, date_jour=timezone.now().date()).first()
    if not cloture:
        return HttpResponse("Aucune clôture trouvée pour aujourd’hui.")

    html_string = render_to_string("cloture/rapport_cloture_pdf.html", {"cloture": cloture})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rapport_cloture.pdf"'

    HTML(string=html_string).write_pdf(response)
    return response



@method_decorator([
    login_required,
    permission_required('stock.add_produit', raise_exception=True)
], name='dispatch')
class AjoutMultipleProduitsView(EntrepriseAccessMixin, View):

    def get_devise_principale_symbole(self, request):
        """Récupère le symbole de la devise principale de l'entreprise."""
        try:
            config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
            return config_saas.devise_principale.symbole if config_saas.devise_principale else "€"
        except ConfigurationSAAS.DoesNotExist:
            return "€"

    def get(self, request, *args, **kwargs):
        """Affiche le formulaire pour l'ajout multiple de produits."""
        categories = Categorie.objects.filter(entreprise=request.entreprise)
        
        context = {
            'categories': categories,
            'devise_principale_symbole': self.get_devise_principale_symbole(request)
        }
        return render(request, 'produit/ajout_multiple.html', context)

    def post(self, request, *args, **kwargs):
        """Traite les données JSON pour la création multiple de produits."""
        try:
            if not request.POST.get('produits_data'):
                return JsonResponse({
                    'success': False, 
                    'error': 'Aucune donnée reçue'
                }, status=400)

            produits_data = json.loads(request.POST['produits_data'])
            categorie_id = request.POST.get('categorie')
            produits_crees = []
            errors = []

            with transaction.atomic():
                for idx, produit_data in enumerate(produits_data):
                    try:
                        # Validation des champs obligatoires
                        nom = produit_data.get('nom')
                        prix_achat = produit_data.get('prix_achat')
                        prix_vente = produit_data.get('prix_vente')

                        if not all([nom, prix_achat, prix_vente]):
                            errors.append(f"Ligne {idx + 1}: Champs obligatoires manquants (nom, prix_achat, prix_vente)")
                            continue

                        # Convertir en Decimal et int
                        prix_achat_dec = Decimal(str(prix_achat))
                        prix_vente_dec = Decimal(str(prix_vente))
                        stock_int = int(produit_data.get('stock', 0))
                        seuil_alerte_int = int(produit_data.get('seuil_alerte', 10))

                        # Création de l'objet Produit
                        produit = Produit(
                            nom=nom,
                            description=produit_data.get('description', ''),
                            prix_achat=prix_achat_dec,
                            prix_vente=prix_vente_dec,
                            stock=stock_int,
                            seuil_alerte=seuil_alerte_int,
                            entreprise=request.entreprise,
                            cree_par=request.user
                        )

                        # Gestion de la catégorie
                        if categorie_id:
                            categorie = Categorie.objects.get(id=categorie_id, entreprise=request.entreprise)
                            produit.categorie = categorie
                        
                        # Traitement de la photo
                        if produit_data.get('photo_data'):
                            format, imgstr = produit_data['photo_data'].split(';base64,')
                            ext = format.split('/')[-1]
                            photo_data = base64.b64decode(imgstr)
                            produit.photo.save(f"{nom}_{request.entreprise.id}_{idx}.{ext}", ContentFile(photo_data), save=False)
                        
                        produit.save()
                        produit.generate_barcode()
                        produit.save() # Pour sauvegarder le code-barre

                        produits_crees.append({
                            'id': produit.id,
                            'nom': produit.nom
                        })
                    
                    except (ValueError, TypeError, Categorie.DoesNotExist) as e:
                        errors.append(f"Ligne {idx + 1}: Erreur de données ou de catégorie - {str(e)}")
                    except Exception as e:
                        errors.append(f"Ligne {idx + 1}: Erreur inattendue - {str(e)}")

            response_data = {
                'success': True if not errors else False,
                'message': f'{len(produits_crees)}/{len(produits_data)} produits créés. '
                           f'{len(errors)} erreurs trouvées.',
                'produits_crees': produits_crees,
                'errors': errors
            }
            return JsonResponse(response_data, status=200 if not errors else 400)

        except json.JSONDecodeError as e:
            return JsonResponse({
                'success': False,
                'error': f'Données JSON invalides: {str(e)}'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur serveur inattendue: {str(e)}'
            }, status=500)
@method_decorator([
    login_required,
    permission_required('stock.change_produit', raise_exception=True)
], name='dispatch')
class ModifierProduitView(EntrepriseAccessMixin, View):
    template_name = 'produit/modifier_produit.html'

    def get_devise_principale_symbole(self, request):
        """Récupère le symbole de la devise principale de l'entreprise."""
        try:
            config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
            return config_saas.devise_principale.symbole if config_saas.devise_principale else "€"
        except ConfigurationSAAS.DoesNotExist:
            return "€"

    def get(self, request, pk, *args, **kwargs):
        """Affiche le formulaire de modification du produit."""
        produit = get_object_or_404(Produit, pk=pk, entreprise=request.entreprise)
        
        context = {
            'produit': produit,
            'categories': Categorie.objects.filter(entreprise=request.entreprise),
            'devise_principale_symbole': self.get_devise_principale_symbole(request),
        }
        return render(request, self.template_name, context)

    def post(self, request, pk, *args, **kwargs):
        """Traite les données du formulaire pour modifier un produit."""
        produit = get_object_or_404(Produit, pk=pk, entreprise=request.entreprise)
        
        try:
            with transaction.atomic():
                produit.nom = request.POST.get('nom')
                produit.reference = request.POST.get('reference')
                produit.categorie_id = request.POST.get('categorie') or None
                produit.prix_achat = Decimal(request.POST.get('prix_achat') or 0)
                produit.prix_vente = Decimal(request.POST.get('prix_vente') or 0)
                produit.stock = int(request.POST.get('stock') or 0)
                produit.tva = int(request.POST.get('tva') or 0)
                produit.description = request.POST.get('description')

                # Gère le téléchargement d'une nouvelle photo
                if 'photo' in request.FILES:
                    produit.photo = request.FILES['photo']
                
                produit.save()
            
            messages.success(request, f"Le produit '{produit.nom}' a été modifié avec succès.")
            return redirect('produits_par_categorie')

        except Exception as e:
            messages.error(request, f"Une erreur s'est produite lors de la modification : {str(e)}")
            
            context = {
                'produit': produit,
                'categories': Categorie.objects.filter(entreprise=request.entreprise),
                'devise_principale_symbole': self.get_devise_principale_symbole(request),
            }
            return render(request, self.template_name, context)


def imprimer_produit(request, pk):
    """Génère la version PDF de la fiche produit"""
    produit = get_object_or_404(Produit, pk=pk)
    parametres = Parametre.objects.first()
    
    # Génération du code-barres (identique à la vue HTML)
    codebarres_img = None
    if produit.code:
        try:
            code = barcode.get('code128', produit.code, writer=ImageWriter())
            buffer = BytesIO()
            code.write(buffer)
            codebarres_img = base64.b64encode(buffer.getvalue()).decode('utf-8')
        except:
            pass
    
    context = {
        'produit': produit,
        'parametres': parametres,
        'date_impression': timezone.now().strftime("%d/%m/%Y %H:%M"),
        'codebarres_img': codebarres_img
    }
    
    template = get_template('produit/fiche_produit_pdf.html')
    html_content = template.render(context)
    
    buffer_pdf = BytesIO()
    HTML(string=html_content).write_pdf(buffer_pdf)
    buffer_pdf.seek(0)
    
    identifiant = getattr(produit, 'reference', None) or produit.code or produit.id
    nom_fichier = f"Fiche_Produit_{identifiant}.pdf"
    
    return FileResponse(buffer_pdf, content_type='application/pdf', filename=nom_fichier)

from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.http import FileResponse
from django.template.loader import get_template
from weasyprint import HTML
from io import BytesIO
import base64


from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.db.models import Avg, F
from .models import Produit, Parametre, TauxChange, Categorie
from decimal import Decimal, InvalidOperation
import base64
from io import BytesIO

def voir_produit(request, pk):
    """Affiche la fiche produit avec gestion robuste des données"""
    try:
        # Récupération sécurisée du produit
        produit = get_object_or_404(Produit, pk=pk)
        
        # Préparation des données du produit avec valeurs par défaut
        produit_data = {
            'id': produit.id,
            'nom': produit.nom,
            'description': produit.description or "Aucune description disponible",
            'prix_achat': float(produit.prix_achat),
            'prix_vente': float(produit.prix_vente),
            'prix_vente_suggere': float(produit.prix_vente_suggere()),
            'stock': produit.stock,
            'seuil_alerte': produit.seuil_alerte,
            'taux_tva': float(produit.taux_tva),
            'code_barre_numero': produit.code_barre_numero or "",
            'categorie': produit.categorie,
            'actif': produit.actif,
            'has_photo': bool(produit.photo),
            'has_code_barre': bool(produit.code_barre),
        }

        # Gestion du created_by si le champ existe
        if hasattr(produit, 'created_by') and produit.created_by:
            created_by_info = produit.created_by.get_full_name() or produit.created_by.username
        else:
            created_by_info = "Système"

        # Paramètres et taux de change
        parametres = Parametre.objects.first()
        taux_usd = taux_eur = taux_produit = None
        prix_achat_usd = prix_vente_usd = None
        date_taux = timezone.now().date().strftime("%d/%m/%Y")

        if parametres:
            try:
                # Taux de change USD
                if parametres.devise_principale != 'USD':
                    taux_usd = TauxChange.get_taux('USD', parametres.devise_principale)
                    if taux_usd:
                        prix_achat_usd = produit_data['prix_achat'] / float(taux_usd)
                        prix_vente_usd = produit_data['prix_vente'] / float(taux_usd)
                
                # Taux de change EUR
                if parametres.devise_principale not in ['USD', 'EUR']:
                    taux_eur = TauxChange.get_taux('EUR', parametres.devise_principale)
            except (TypeError, InvalidOperation, AttributeError) as e:
                print(f"Erreur conversion devise: {e}")

        # Gestion des images
        photo_base64 = None
        code_barre_base64 = None
        
        if produit.photo:
            try:
                with produit.photo.open('rb') as photo_file:
                    photo_base64 = base64.b64encode(photo_file.read()).decode('utf-8')
            except Exception as e:
                print(f"Erreur lecture photo: {e}")

        if produit.code_barre:
            try:
                with produit.code_barre.open('rb') as barcode_file:
                    code_barre_base64 = base64.b64encode(barcode_file.read()).decode('utf-8')
            except Exception as e:
                print(f"Erreur lecture code-barres: {e}")

        # Calcul de la marge
        marge_brute = produit_data['prix_vente'] - produit_data['prix_achat']
        try:
            pourcentage_marge = (marge_brute / produit_data['prix_achat']) * 100
        except ZeroDivisionError:
            pourcentage_marge = 0

        # Préparation du contexte final
        context = {
            'produit': produit_data,
            'parametres': parametres,
            'photo_base64': photo_base64,
            'code_barre_base64': code_barre_base64,
            'taux_usd': float(taux_usd) if taux_usd else None,
            'taux_eur': float(taux_eur) if taux_eur else None,
            'prix_achat_usd': round(prix_achat_usd, 2) if prix_achat_usd else None,
            'prix_vente_usd': round(prix_vente_usd, 2) if prix_vente_usd else None,
            'marge_brute': round(marge_brute, 2),
            'pourcentage_marge': round(pourcentage_marge, 2),
            'created_by_info': created_by_info,
            'date_impression': timezone.now().strftime("%d/%m/%Y à %H:%M"),
            'date_taux': date_taux,
        }

        return render(request, 'produit/fiche_produit.html', context)

    except Exception as e:
        # Contexte minimal en cas d'erreur grave
        context = {
            'error': str(e),
            'produit_id': pk,
            'produit': {
                'nom': 'Produit non trouvé',
                'description': 'Impossible de charger les détails du produit',
                'prix_achat': 0,
                'prix_vente': 0,
                'stock': 0,
                'seuil_alerte': 0,
            },
            'created_by_info': "Système",
            'photo_base64': None,
            'code_barre_base64': None,
        }

def imprimer_produit(request, pk):
    """Génère la version PDF"""
    try:
        produit = get_object_or_404(Produit, pk=pk)
        
        # Récupération de la devise principale de l'entreprise
        devise_principale_symbole = "€"  # Valeur par défaut
        try:
            config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
            if config_saas.devise_principale:
                devise_principale_symbole = config_saas.devise_principale.symbole
        except ConfigurationSAAS.DoesNotExist:
            # Gérer le cas où la configuration n'existe pas
            pass

        context = {
            'produit': produit,
            # Le modèle Parametre est-il toujours pertinent ? Si non, retirez cette ligne.
            'parametres': Parametre.objects.first() if Parametre.objects.exists() else None,
            'devise_principale_symbole': devise_principale_symbole,
            'date_impression': timezone.now().strftime("%d/%m/%Y %H:%M"),
            'code_barre_img': None
        }
        
        if produit.code_barre:
            try:
                with open(produit.code_barre.path, "rb") as image_file:
                    context['code_barre_img'] = base64.b64encode(image_file.read()).decode('utf-8')
            except Exception as e:
                print(f"Erreur lecture code-barres PDF: {e}")
        
        template = get_template('produit/fiche_produit.html')
        html_content = template.render(context)
        
        buffer = BytesIO()
        HTML(string=html_content).write_pdf(buffer)
        buffer.seek(0)
        
        nom_fichier = f"Fiche_Produit_{produit.nom.replace(' ', '_')}_{produit.id}.pdf"
        return FileResponse(buffer, content_type='application/pdf', filename=nom_fichier)
        
    except Exception as e:
        return render(request, 'produit/fiche_produit.html', {'error': str(e)}, status=500)
    

from openpyxl.styles import Font


import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import Produit

def exporter_produits_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits en stock"

    # En-têtes
    headers = ['ID', 'Nom', 'Catégorie', 'Prix Achat', 'Prix Vente', 'Stock','Seuil_alerte']
    ws.append(headers)

    # Style pour l'entête
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Données
    produits = Produit.objects.filter(stock__gt=0)

    for p in produits:
        ws.append([
            p.id,
            p.nom,
            p.categorie.nom if p.categorie else 'Non classé',
            p.prix_achat,
            p.prix_vente,
            p.stock,
            p.seuil_alerte,
        ])

    # Génération du fichier
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=produits_en_stock.xlsx'
    wb.save(response)
    return response
import io
def exporter_produits_pdf(request):
    categorie_id = request.GET.get('categorie')

    produits = Produit.objects.all()
    if categorie_id:
        produits = produits.filter(categorie_id=categorie_id)

    context = {
        'produits': produits,
        'parametres': Parametre.objects.first(),
        'user': request.user,
        'now': timezone.now(),
    }

    html = render_to_string("produit/produits_pdf.html", context)
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="liste_produits.pdf"'  # Changé de 'inline' à 'attachment'
    HTML(string=html).write_pdf(response)
    return response




#export bd
from django.http import HttpResponse
from django.core.management import call_command
import io
from datetime import datetime
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def telecharger_sauvegarde(request):
    now = datetime.now().strftime('%Y-%m-%d_%H-%M')
    buffer = io.StringIO()
    call_command('dumpdata', indent=2, stdout=buffer)
    response = HttpResponse(buffer.getvalue(), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename=backup_{now}.json'
    return response
#facture

def creer_facture(request, commande_id):
    commande = get_object_or_404(Commande, pk=commande_id)
    
    if request.method == 'POST':
        numero = request.POST.get('numero')
        if not numero:
            numero = f"FACT-{commande.id}-{commande.date_commande.year}"
        
        facture = Facture.objects.create(
            commande=commande,
            numero=numero,
            montant_total=commande.montant_total
        )
        messages.success(request, 'Facture créée avec succès!')
        return redirect('detail_facture', pk=facture.id)
    
    return render(request, 'factures/creer_facture.html', {'commande': commande})

from django.views.decorators.http import require_http_methods
from django.contrib import messages

@require_http_methods(["POST"])
def enregistrer_paiement(request, facture_id):
    facture = get_object_or_404(Facture, id=facture_id)
    
    try:
        montant = Decimal(request.POST.get('montant', 0))
        methode = request.POST.get('methode', 'especes')
        
        if montant <= 0:
            messages.error(request, "Le montant doit être positif")
            return redirect('detail_facture', pk=facture.id)
            
        # Crée le paiement (qui mettra à jour la facture automatiquement)
        Paiement.objects.create(
            facture=facture,
            montant=montant,
            methode=methode
        )
        
        messages.success(request, "Paiement enregistré avec succès !")
        return redirect('detail_facture', pk=facture.id)
        
    except (ValueError, InvalidOperation):
        messages.error(request, "Montant invalide")
        return redirect('detail_facture', pk=facture.id)
    
    
def detail_facture(request, pk):
    facture = get_object_or_404(Facture, pk=pk)
    parametres = Parametre.objects.first()
    context = {
        'facture': facture,
        'paiements': facture.paiement_set.all(),
        'reste': facture.reste_a_payer(),
        'parametres': parametres
    }
    return render(request, 'factures/detail_facture.html', context)

from reportlab.pdfgen import canvas
def facture_pdf(request, pk):
    facture = get_object_or_404(Facture, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    p = canvas.Canvas(response)
    p.drawString(100, 750, f"Facture {facture.numero}")
    p.showPage()
    p.save()
    return response

# Dans views.py
from django.views.generic import CreateView

class AjoutPaiement(CreateView):
    model = Paiement
    fields = ['montant', 'methode']
    template_name = 'factures/ajout_paiement.html'

    def form_valid(self, form):
        form.instance.facture = get_object_or_404(Facture, pk=self.kwargs['pk'])
        return super().form_valid(form)

def liste_paiements(request):
    # Récupération des paramètres de l'entreprise
    parametres = Parametre.objects.first()
    
    # Récupération de tous les paiements avec les relations optimisées
    paiements = Paiement.objects.select_related(
        'facture', 
        'facture__commande', 
        'facture__commande__client'
    ).order_by('-date')

    # Calcul des statistiques globales
    stats = paiements.aggregate(
        total=Sum('montant'),
        count=Count('id'),
        moyenne=Avg('montant')
    )

    # Préparation du contexte
    context = {
        'parametres': parametres,
        'paiements': paiements,
        'methodes': Paiement.methodes_disponibles(),
        'total_paiements': stats['total'] or 0,
        'moyenne_paiements': stats['moyenne'] or 0,
    }
    
    return render(request, 'factures/liste_paiements.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages


def liste_factures(request):
    factures = Facture.objects.all().order_by('-date')
    context = {'factures': factures}
    return render(request, 'factures/liste_factures.html', context)

def liste_paiements(request):
    paiements = Paiement.objects.select_related('facture').order_by('-date')
    context = {'paiements': paiements}
    return render(request, 'factures/liste_paiements.html', context)

def actions_facture(request, facture_id):
    facture = get_object_or_404(Facture, pk=facture_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'marquer_payee' and facture.reste_a_payer() <= 0:
            facture.statut = 'payee'
            facture.save()
            messages.success(request, "Facture marquée comme payée")
        
        elif action == 'annuler':
            facture.statut = 'annulee'
            facture.save()
            messages.warning(request, "Facture annulée")
            
        return redirect('liste_factures')
    
    return render(request, 'factures/actions_facture.html', {'facture': facture})







# from django.db import IntegrityError  # Ajoutez cette ligne
# from django.db import models  # Si pas déjà présent
# from django.db import transaction  # Si pas déjà présent
# from django.views.generic import ListView, CreateView, UpdateView, DetailView
# from django.urls import reverse_lazy
# from django.contrib.messages.views import SuccessMessageMixin

# from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
# from django.db.models import Q

# class FournisseurListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
#     model = Fournisseur
#     template_name = 'achats/fournisseur_list.html'
#     context_object_name = 'fournisseurs'
#     paginate_by = 25
#     permission_required = 'STOCK.view_fournisseur'

#     def get_queryset(self):
#         if not hasattr(self.request, 'entreprise'):
#             raise PermissionDenied("Entreprise non définie")
        
#         queryset = super().get_queryset().filter(entreprise=self.request.entreprise)
        
#         # Filtres
#         search_query = self.request.GET.get('search')
#         type_filter = self.request.GET.get('type')
        
#         if search_query:
#             queryset = queryset.filter(
#                 Q(nom__icontains=search_query) | 
#                 Q(telephone__icontains=search_query) |
#                 Q(email__icontains=search_query)
#             )
        
#         if type_filter:
#             queryset = queryset.filter(type=type_filter)
            
#         return queryset.order_by('-created_at')

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context.update({
#             'search_query': self.request.GET.get('search', ''),
#             'selected_type': self.request.GET.get('type', ''),
#             'can_add': self.request.user.has_perm('STOCK.add_fournisseur'),
#             'can_export': self.request.user.has_perm('STOCK.export_fournisseurs'),
#             'type_choices': Fournisseur.TYPE_FOURNISSEUR,
#             'entreprise': self.request.entreprise
#         })
#         return context

#     def handle_no_permission(self):
#         messages.error(self.request, "Accès refusé")
#         return redirect('security:acces_refuse')

# class FournisseurCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
#     model = Fournisseur
#     template_name = 'achats/fournisseur_form.html'
#     fields = [
#         'nom', 'type', 'telephone', 'email', 
#         'adresse', 'ville', 'pays', 'code_postal',
#         'taux_tva', 'delai_livraison', 'notes'
#     ]
#     success_url = reverse_lazy('fournisseur_list')
#     success_message = "Fournisseur créé avec succès"
#     permission_required = 'STOCK.add_fournisseur'
    
#     def form_valid(self, form):
#         form.instance.entreprise = self.request.entreprise
#         form.instance.created_by = self.request.user
#         try:
#             return super().form_valid(form)
#         except IntegrityError:
#             form.add_error('telephone', 'Un fournisseur avec ce numéro existe déjà')
#             return self.form_invalid(form)

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['title'] = "Ajouter un nouveau fournisseur"
#         context['entreprise'] = self.request.entreprise
#         return context

# class FournisseurDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
#     model = Fournisseur
#     template_name = 'achats/fournisseur_detail.html'
#     context_object_name = 'fournisseur'
#     permission_required = 'STOCK.view_fournisseur'
    
#     def get_queryset(self):
#         return super().get_queryset().filter(entreprise=self.request.entreprise)

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['can_edit'] = self.request.user.has_perm('STOCK.change_fournisseur')
#         context['can_delete'] = self.request.user.has_perm('STOCK.delete_fournisseur')
#         context['entreprise'] = self.request.entreprise
#         return context

# class FournisseurUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
#     model = Fournisseur
#     template_name = 'achats/fournisseur_form.html'
#     fields = [
#         'nom', 'type', 'telephone', 'email', 
#         'adresse', 'ville', 'pays', 'code_postal',
#         'taux_tva', 'delai_livraison', 'notes'
#     ]
#     success_url = reverse_lazy('fournisseur_list')
#     success_message = "Fournisseur modifié avec succès"
#     permission_required = 'STOCK.change_fournisseur'
    
#     def get_queryset(self):
#         return super().get_queryset().filter(entreprise=self.request.entreprise)

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['title'] = f"Modifier {self.object.nom}"
#         context['entreprise'] = self.request.entreprise
#         return context

# class FournisseurDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
#     model = Fournisseur
#     template_name = 'achats/fournisseur_confirm_delete.html'
#     success_url = reverse_lazy('fournisseur_list')
#     permission_required = 'STOCK.delete_fournisseur'
    
#     def get_queryset(self):
#         return super().get_queryset().filter(entreprise=self.request.entreprise)

#     def post(self, request, *args, **kwargs):
#         try:
#             response = super().post(request, *args, **kwargs)
#             messages.success(request, "Fournisseur supprimé avec succès")
#             return response
#         except models.ProtectedError:
#             messages.error(
#                 request,
#                 "Impossible de supprimer ce fournisseur car il est associé à des commandes"
#             )
#             return redirect('fournisseur_detail', pk=kwargs['pk'])

# from django.shortcuts import render

# def acces_refuse(request):
#     return render(request, 'acces_refuse.html', status=403)



# class AchatCreateView(EntrepriseAccessMixin, SuccessMessageMixin, CreateView):
#     model = Achat
#     template_name = 'achats/achat_form.html'
#     fields = ['fournisseur', 'produit', 'quantite', 'prix_unitaire', 'date_achat', 'numero_facture', 'notes']
#     success_message = "Achat enregistré avec succès"

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         try:
#             context['devise_principale'] = ConfigurationSAAS.objects.get(
#                 entreprise=self.request.entreprise
#             ).devise_principale
#         except ConfigurationSAAS.DoesNotExist:
#             messages.warning(self.request, "Configuration devise manquante")
#         return context

#     @transaction.atomic
#     def form_valid(self, form):
#         form.instance.entreprise = self.request.entreprise
#         form.instance.created_by = self.request.user
#         form.instance.date_achat = form.instance.date_achat or timezone.now().date()
        
#         # Définition du montant original (identique au prix unitaire dans ce cas simple)
#         form.instance.montant_original = form.cleaned_data['prix_unitaire']
        
#         # Mise à jour atomique du stock
#         Produit.objects.filter(pk=form.instance.produit.pk).update(
#             stock=F('stock') + form.instance.quantite
#         )

#         return super().form_valid(form)

#     def get_success_url(self):
#         return reverse_lazy('achat_detail', kwargs={'pk': self.object.pk})

#     def get_form(self, form_class=None):
#         form = super().get_form(form_class)
        
#         # Filtrage strict par entreprise
#         form.fields['fournisseur'].queryset = Fournisseur.objects.filter(
#             entreprise=self.request.entreprise
#         ).order_by('nom')
        
#         form.fields['produit'].queryset = Produit.objects.filter(
#             entreprise=self.request.entreprise,
#             actif=True
#         ).order_by('nom')

#         # Configuration des champs
#         decimal_fields = ['prix_unitaire', 'quantite']
#         for field in decimal_fields:
#             form.fields[field].widget.attrs.update({
#                 'class': 'form-control',
#                 'min': '0.01',
#                 'step': '0.01' if field == 'prix_unitaire' else '1'
#             })
        
#         form.fields['date_achat'].widget.attrs.update({
#             'class': 'form-control datepicker'
#         })

#         return form
# class AchatDetailView(LoginRequiredMixin, PermissionRequiredMixin, EntrepriseAccessMixin, DetailView):
#     model = Achat
#     template_name = 'achats/achat_detail.html'
#     permission_required = 'STOCK.view_achat'

#     def get_object(self, queryset=None):
#         obj = super().get_object(queryset)
#         if obj.entreprise != self.request.entreprise:
#             raise Http404("Cet achat n'existe pas ou vous n'y avez pas accès")
#         return obj

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         achat = self.object
#         entreprise = self.request.entreprise  # Récupération de l'entreprise

#         try:
#             config = ConfigurationSAAS.objects.get(entreprise=entreprise)
#             devise_principale = config.devise_principale
#             symbole_devise = devise_principale.symbole if devise_principale else '$'
#             taux_tva = config.taux_tva
#         except (ConfigurationSAAS.DoesNotExist, AttributeError):
#             symbole_devise = '$'
#             taux_tva = Decimal('16.00')

#         # Calcul des montants
#         montant_ht = achat.total_achat
#         montant_tva = montant_ht * taux_tva / 100
#         montant_ttc = montant_ht + montant_tva

#         context.update({
#             'entreprise': entreprise,  # Ajout de l'entreprise au contexte
#             'stock_avant': achat.produit.stock - achat.quantite if achat.quantite else achat.produit.stock,
#             'auteur_achat': achat.created_by,
#             'symbole_devise': symbole_devise,
#             'taux_tva': taux_tva,
#             'montant_tva': montant_tva,
#             'montant_ttc': montant_ttc,
#             'can_edit': self.request.user.has_perm('STOCK.change_achat'),
#             'can_delete': self.request.user.has_perm('STOCK.delete_achat'),
#             'can_create_transaction': self.request.user.has_perm('COMPTA.add_transaction'),
#         })
#         return context
#     def handle_no_permission(self):
#         messages.error(self.request, "Accès refusé")
#         return redirect('security:acces_refuse')
    
 
# class AchatDeleteView(LoginRequiredMixin, PermissionRequiredMixin, EntrepriseAccessMixin, DeleteView):
#     model = Achat
#     template_name = 'achats/achat_confirm_delete.html'
#     permission_required = 'STOCK.delete_achat'
#     success_url = reverse_lazy('achat_list')

#     def get_object(self, queryset=None):
#         obj = super().get_object(queryset)
#         if obj.entreprise != self.request.entreprise:
#             raise Http404("Cet achat n'existe pas ou vous n'y avez pas accès")
#         return obj

#     def delete(self, request, *args, **kwargs):
#         messages.success(request, "L'achat a été supprimé avec succès.")
#         return super().delete(request, *args, **kwargs)

#     def handle_no_permission(self):
#         messages.error(self.request, "Vous n'avez pas la permission de supprimer cet achat.")
#         return redirect('security:acces_refuse')   
# from django.views.generic.edit import UpdateView
# from .forms import AchatForm

# class AchatUpdateView(LoginRequiredMixin, PermissionRequiredMixin, EntrepriseAccessMixin, UpdateView):
#     model = Achat
#     form_class = AchatForm
#     template_name = 'achats/achat_form.html'
#     permission_required = 'STOCK.change_achat'

#     def get_object(self, queryset=None):
#         obj = super().get_object(queryset)
#         if obj.entreprise != self.request.entreprise:
#             raise Http404("Cet achat n'existe pas ou vous n'y avez pas accès")
#         return obj

#     def get_success_url(self):
#         messages.success(self.request, "L'achat a été modifié avec succès.")
#         return reverse('achat_detail', kwargs={'pk': self.object.pk})

#     def get_form_kwargs(self):
#         kwargs = super().get_form_kwargs()
#         kwargs['entreprise'] = self.request.entreprise
#         return kwargs

#     def handle_no_permission(self):
#         messages.error(self.request, "Vous n'avez pas la permission de modifier cet achat.")
#         return redirect('security:acces_refuse')   
# from weasyprint import HTML, CSS  
#   # STOCK/views/achat_pdf_view.py



# logger = logging.getLogger(__name__)

# def achat_pdf_view(request, pk):
#     # Récupération de l'achat avec vérification entreprise
#     achat = get_object_or_404(Achat, pk=pk, entreprise=request.entreprise)

#     try:
#         config = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
#         taux_tva = config.taux_tva
#         devise_principale = config.devise_principale
#         symbole_devise = devise_principale.symbole if devise_principale else '$'
#     except ConfigurationSAAS.DoesNotExist:
#         taux_tva = Decimal('0.00')
#         symbole_devise = '$'

#     # Calcul des montants
#     total_ht = Decimal(achat.quantite) * Decimal(achat.prix_unitaire)
#     montant_tva = total_ht * taux_tva / 100
#     total_ttc = total_ht + montant_tva

#     context = {
#         'achat': achat,
#         'entreprise': request.entreprise,
#         'total_ht': total_ht,
#         'montant_tva': montant_tva,
#         'total_ttc': total_ttc,
#         'taux_tva': taux_tva,
#         'symbole_devise': symbole_devise,
#         'stock_avant': achat.produit.stock - achat.quantite,
#         'auteur': achat.created_by,
#         'date_generation': timezone.now(),
#     }

#     if hasattr(request.entreprise, 'logo') and request.entreprise.logo:
#         context['logo_path'] = request.entreprise.logo.path

#     # Charger le template sans les CSS @page problématiques
#     template = get_template("achats/achat_pdf.html")
#     html = template.render(context)

#     # Définir le type de réponse (inline ou download)
#     response_type = HttpResponse(content_type='application/pdf')
#     if 'download' in request.GET:
#         response_type['Content-Disposition'] = f'attachment; filename="achat_{achat.id}.pdf"'
#     else:
#         response_type['Content-Disposition'] = f'inline; filename="achat_{achat.id}.pdf"'

#     # Génération du PDF
#     pisa_status = pisa.CreatePDF(html, dest=response_type, encoding='UTF-8')

#     if pisa_status.err:
#         logger.error(f"Erreur lors de la génération du PDF: {pisa_status.err}")
#         return HttpResponse("Une erreur est survenue lors de la génération du PDF.")
    
#     return response_type

# class AchatListView(LoginRequiredMixin, PermissionRequiredMixin, EntrepriseAccessMixin, ListView):
#     model = Achat
#     template_name = 'achats/achat_list.html'
#     context_object_name = 'achats'
#     paginate_by = 15
#     permission_required = 'STOCK.view_achat'

#     def get_queryset(self):
#         queryset = super().get_queryset().filter(entreprise=self.request.entreprise)
#         queryset = queryset.select_related('fournisseur', 'produit')

#         # Filtres
#         date_debut = self.request.GET.get('date_debut')
#         date_fin = self.request.GET.get('date_fin')
#         fournisseur_id = self.request.GET.get('fournisseur')

#         if date_debut:
#             queryset = queryset.filter(date_achat__gte=date_debut)
#         if date_fin:
#             queryset = queryset.filter(date_achat__lte=date_fin)
#         if fournisseur_id:
#             queryset = queryset.filter(fournisseur_id=fournisseur_id)

#         return queryset.order_by('-date_achat')

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
        
#         # Récupération de la devise principale depuis ConfigurationSAAS
#         try:
#             config = ConfigurationSAAS.objects.get(entreprise=self.request.entreprise)
#             devise_principale = config.devise_principale
#             # Accès au symbole via la relation ForeignKey
#             symbole_devise = devise_principale.symbole if devise_principale else '$'
#             code_devise = devise_principale.code if devise_principale else 'USD'
#         except (ConfigurationSAAS.DoesNotExist, AttributeError):
#             code_devise = 'USD'
#             symbole_devise = '$'

#         # Récupération des totaux
#         achats = context['achats']
#         total_achats = sum(a.total_achat for a in achats)
#         total_quantite = sum(a.quantite for a in achats)

#         context.update({
#             'total_achats': total_achats,
#             'total_quantite': total_quantite,
#             'achats_count': achats.count(),
#             'fournisseurs': Fournisseur.objects.filter(entreprise=self.request.entreprise).order_by('nom'),
#             'selected_fournisseur': self.request.GET.get('fournisseur', ''),
#             'date_debut': self.request.GET.get('date_debut', ''),
#             'date_fin': self.request.GET.get('date_fin', ''),
#             'can_add': self.request.user.has_perm('STOCK.add_achat'),
#             'can_export': self.request.user.has_perm('STOCK.export_achats'),
#             'devise_principale': code_devise,
#             'symbole_devise': symbole_devise,
#         })
#         return context

#     def handle_no_permission(self):
#         messages.error(self.request, "Accès refusé")
#         return redirect('security:acces_refuse')
   

@login_required
def liste_notifications(request):
    notifications = request.user.notifications.all().order_by('-date_creation')
    return render(request, 'notifications/liste.html', {'notifications': notifications})


def notifier_facture_impayee(facture):
    if not facture.est_payee:
        for utilisateur in User.objects.filter(is_staff=True):
            Notification.objects.create(
                utilisateur=utilisateur,
                message=f"Facture #{facture.id} non payée pour {facture.client.nom}.",
                url=f"/factures/{facture.id}/"
            )

def rapport(request):
    return render(request,"rapports/rapport.html")
#facture pdf 
from django.http import HttpResponse, Http404
from django.template.loader import get_template
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from xhtml2pdf import pisa
from io import BytesIO
import logging
import os
logger = logging.getLogger(__name__)

class FacturePDFView(LoginRequiredMixin, View):
    def get(self, request, pk):
        try:
            facture = Facture.objects.select_related(
                'commande__client',
                'commande__vendeur'
            ).prefetch_related(
                'commande__lignes__produit'
            ).get(pk=pk)
        except Facture.DoesNotExist:
            raise Http404("Facture non trouvée.")
        
        parametres = Parametre.objects.first()
        
        context = {
            "facture": facture,
            "parametres": parametres,
            "user": request.user,
        }

        try:
            template = get_template("factures/facture_pdf.html")
            html = template.render(context)
            
            result = BytesIO()
            
            # Configuration supplémentaire pour xhtml2pdf
            pdf = pisa.pisaDocument(
                BytesIO(html.encode("UTF-8")),
                result,
                encoding='UTF-8',
                link_callback=self.link_callback
            )
            
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                filename = f"facture_{facture.numero or facture.id}.pdf"
                response['Content-Disposition'] = f'inline; filename="{filename}"'
                return response
            else:
                logger.error(f"Erreur PDF: {pdf.err}")
                return HttpResponse("Erreur lors de la génération du PDF", status=500)
                
        except Exception as e:
            logger.error(f"Erreur génération facture: {str(e)}")
            return HttpResponse(f"Erreur serveur: {str(e)}", status=500)

    def link_callback(self, uri, rel):
        """
        Callback pour gérer les ressources (images, CSS)
        """
        # Chemin absolu pour les médias
        if uri.startswith(settings.MEDIA_URL):
            path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
        # Chemin absolu pour les fichiers statiques
        elif uri.startswith(settings.STATIC_URL):
            path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        else:
            path = None
            
        if path and os.path.isfile(path):
            return path
        return None
    
    
 
  #retour mses 
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db import transaction

from django.db import transaction
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)

@login_required
def liste_retours(request):
    """Liste des retours filtrés par entreprise SAAS"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    retours = RetourProduit.objects.filter(
        entreprise=request.entreprise
    ).select_related('commande', 'responsable', 'commande__client').order_by('-date_creation')

    context = {
        'retours': retours,
        'statuts': dict(RetourProduit.STATUT_RETOUR),
        'entreprise': request.entreprise
    }
    return render(request, 'retours/liste.html', context)

@login_required
def creer_retour(request, commande_id):
    """Création d'un retour avec gestion SAAS"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    commande = get_object_or_404(Commande, id=commande_id, entreprise=request.entreprise)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Création du retour
                retour = RetourProduit.objects.create(
                    commande=commande,
                    responsable=request.user,
                    motif=request.POST.get('motif'),
                    notes=request.POST.get('notes', ''),
                    entreprise=request.entreprise
                )
                
                # Création des lignes de retour
                for ligne in commande.lignes.all():
                    quantite = int(request.POST.get(f'quantite_{ligne.id}', 0))
                    if quantite > 0:
                        LigneRetour.objects.create(
                            retour=retour,
                            produit=ligne.produit,
                            quantite=quantite,
                            prix_unitaire=ligne.prix_unitaire,
                            entreprise=request.entreprise  # Maintenant supporté
                        )
                
                messages.success(request, "Retour créé avec succès.")
                return redirect('detail_retour', retour_id=retour.id)
                
        except Exception as e:
            messages.error(request, f"Erreur lors de la création : {str(e)}")
            logger.error(f"Erreur création retour: {str(e)}", exc_info=True)
    
    context = {
        'commande': commande,
        'motifs': dict(RetourProduit.MOTIF_RETOUR),
        'entreprise': request.entreprise
    }
    return render(request, 'retours/creer.html', context)


@login_required
def detail_retour(request, retour_id):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    retour = get_object_or_404(
        RetourProduit.objects.select_related('commande__client', 'responsable')
                           .prefetch_related('lignes__produit'),
        id=retour_id,
        entreprise=request.entreprise  # Filtre SAAS
    )

    context = {
        'retour': retour,
        'user': request.user,
        'is_responsable': retour.responsable == request.user,
        'is_admin': request.user.is_superuser,
        'entreprise': request.entreprise,
        'can_traiter': retour.peut_etre_traite() and request.user.has_perm('STOCK.change_retourproduit')
    }

    return render(request, 'retours/detail.html', context)

@login_required
def traiter_retour(request, retour_id):
    """Traitement d'un retour avec gestion SAAS"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    retour = get_object_or_404(
        RetourProduit.objects.select_related('commande'),
        id=retour_id,
        entreprise=request.entreprise
    )
    
    if not retour.peut_etre_traite():
        messages.warning(request, "Ce retour ne peut pas être traité.")
        return redirect('detail_retour', retour.id)
    
    if not request.user.has_perm('STOCK.change_retourproduit'):
        messages.error(request, "Vous n'avez pas la permission de traiter ce retour.")
        return redirect('detail_retour', retour.id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                if retour.traiter(request.user):
                    messages.success(request, "Retour traité avec succès. Stock mis à jour.")
                else:
                    messages.error(request, "Erreur lors du traitement.")
        except Exception as e:
            messages.error(request, f"Erreur lors du traitement : {str(e)}")
            logger.error(f"Erreur traitement retour: {str(e)}", exc_info=True)
        
        return redirect('detail_retour', retour.id)
    
    return render(request, 'retours/traiter.html', {
        'retour': retour,
        'entreprise': request.entreprise
    })

@login_required
def get_lignes_commande(request, commande_id):
    """API JSON pour les lignes de commande (SAAS)"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        return JsonResponse({'error': 'Entreprise non définie'}, status=403)

    commande = get_object_or_404(Commande, id=commande_id, entreprise=request.entreprise)
    lignes = LigneCommandeClient.objects.filter(commande=commande).select_related('produit')
    
    data = [{
        'id': ligne.id,
        'produit_id': ligne.produit.id,
        'nom': ligne.produit.nom,
        'reference': ligne.produit.code_barre_numero or '-',
        'quantite_max': ligne.quantite,
        'prix': str(ligne.prix_unitaire),
    } for ligne in lignes]
    
    return JsonResponse({'data': data}, safe=False)

def parametre(request):
    return render(request,'paramettre/paramettre.html')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import JsonResponse
from django.db import transaction
from django.utils import timezone
from decimal import Decimal
from django.db.models import F, Sum, Q
from django.db.models.functions import Coalesce
from django.core.exceptions import ValidationError
import logging

logger = logging.getLogger(__name__)

@login_required
@permission_required('STOCK.view_promotion', raise_exception=True)
def liste_promotions(request):
    """Liste toutes les promotions avec filtres"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    now = timezone.now()
    promotions = Promotion.objects.select_related('produit', 'categorie', 'created_by') \
                                 .filter(entreprise=request.entreprise) \
                                 .order_by('-date_creation')
    
    # Filtres
    statut = request.GET.get('statut')
    if statut == 'actives':
        promotions = promotions.filter(actif=True, date_debut__lte=now, date_fin__gte=now)
    elif statut == 'inactives':
        promotions = promotions.filter(actif=False)
    elif statut == 'futures':
        promotions = promotions.filter(date_debut__gt=now)
    elif statut == 'expirees':
        promotions = promotions.filter(date_fin__lt=now)
    
    type_remise = request.GET.get('type_remise')
    if type_remise in ['pourcentage', 'montant_fixe']:
        promotions = promotions.filter(type_remise=type_remise)
    
    return render(request, 'promotions/liste.html', {
        'promotions': promotions,
        'now': now,
        'filtres': {
            'statut': statut,
            'type_remise': type_remise
        },
        'entreprise': request.entreprise
    })

@login_required
@permission_required('STOCK.add_promotion', raise_exception=True)
def creer_promotion(request):
    """Crée une nouvelle promotion"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    if request.method == 'POST':
        try:
            with transaction.atomic():
                actif = 'actif' not in request.POST or request.POST.get('actif') == 'on'
                
                type_remise = request.POST.get('type_remise')
                if type_remise not in dict(Promotion.TypeRemise.choices):
                    raise ValidationError("Type de remise invalide")
                
                promotion = Promotion.objects.create(
                    nom=request.POST.get('nom'),
                    description=request.POST.get('description', ''),
                    type_remise=type_remise,
                    valeur=Decimal(request.POST.get('valeur')),
                    appliquer_a=request.POST.get('appliquer_a'),
                    produit_id=request.POST.get('produit') or None,
                    categorie_id=request.POST.get('categorie') or None,
                    date_debut=request.POST.get('date_debut'),
                    date_fin=request.POST.get('date_fin'),
                    actif=actif,
                    code_promo=request.POST.get('code_promo') or None,
                    usage_unique=request.POST.get('usage_unique') == 'on',
                    utilisations_max=int(request.POST.get('utilisations_max', 1)),
                    created_by=request.user,
                    entreprise=request.entreprise
                )
                
                messages.success(request, "Promotion créée avec succès")
                return redirect('liste_promotions')
        
        except ValidationError as e:
            messages.error(request, f"Erreur de validation: {str(e)}")
        except ValueError as e:
            messages.error(request, f"Valeur incorrecte: {str(e)}")
        except Exception as e:
            messages.error(request, f"Erreur lors de la création: {str(e)}")
            logger.error(f"Erreur création promotion: {str(e)}", exc_info=True)
    
    produits = Produit.objects.filter(entreprise=request.entreprise, actif=True).order_by('nom')
    categories = Categorie.objects.filter(entreprise=request.entreprise).order_by('nom')
    
    return render(request, 'promotions/creer.html', {
        'produits': produits,
        'categories': categories,
        'type_remise_choices': Promotion.TypeRemise.choices,
        'appliquer_a_choices': Promotion.TypeApplication.choices,
        'defaults': {
            'actif': True,
            'utilisations_max': 1,
            'type_remise': Promotion.TypeRemise.POURCENTAGE,
        },
        'entreprise': request.entreprise
    })
@login_required
@permission_required('STOCK.apply_promotion', raise_exception=True)
def appliquer_promotion_commande(request, commande_id):
    """Applique une promotion à une commande et met à jour les montants"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    commande = get_object_or_404(Commande, id=commande_id, entreprise=request.entreprise)
    
    if request.method == 'POST':
        code_promo = request.POST.get('code_promo', '').strip()

        try:
            with transaction.atomic():
                # 1. Vérification de la promotion
                promotion = Promotion.objects.get(
                    code_promo=code_promo,
                    entreprise=request.entreprise
                )

                if not promotion.est_valide():
                    messages.error(request, "Cette promotion n'est plus valide.")
                    return redirect('detail_commande', commande_id=commande.id)

                if RemiseCommandeClient.objects.filter(commande=commande, promotion=promotion).exists():
                    messages.error(request, "Cette promotion a déjà été appliquée à cette commande.")
                    return redirect('detail_commande', commande_id=commande.id)

                # 2. Calcul des montants initiaux
                commande.calculer_totaux()  # S'assurer que les montants sont à jour
                montant_initial = commande.montant_ht + commande.montant_tva

                # 3. Application de la remise
                montant_remise = promotion.appliquer_remise(montant_initial)
                
                # 4. Enregistrement de la remise
                RemiseCommandeClient.objects.create(
                    commande=commande,
                    promotion=promotion,
                    montant_remise=montant_remise,
                    entreprise=request.entreprise
                )

                # 5. Mise à jour de la commande
                commande.montant_remise += montant_remise
                commande.montant_total = max(Decimal('0'), (commande.montant_ht + commande.montant_tva) - commande.montant_remise)
                commande.save()

                # 6. Mise à jour de la promotion
                promotion.incrementer_utilisation()

                # 7. Message de succès avec détails
                devise = commande.client.devise_preferee.symbole if hasattr(commande.client, 'devise_preferee') else ''
                messages.success(
                    request, 
                    f"Promotion '{promotion.nom}' appliquée. "
                    f"Remise: {montant_remise:.2f} {devise} - "
                    f"Nouveau total: {commande.montant_total:.2f} {devise}"
                )

        except Promotion.DoesNotExist:
            messages.error(request, "Code promo invalide.")
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Une erreur est survenue: {str(e)}")
            logger.error(f"Erreur application promotion: {str(e)}", exc_info=True)

    return redirect('detail_commande', commande_id=commande.id)

@login_required
@permission_required('STOCK.view_remisecommande', raise_exception=True)
def detail_remise(request, remise_id):
    """Affiche le détail d'une remise appliquée"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    remise = get_object_or_404(
        RemiseCommandeClient.objects.select_related('commande', 'promotion', 'commande__client'),
        id=remise_id,
        entreprise=request.entreprise
    )

    # Calcul des impacts financiers
    impact = {
        'montant_initial': remise.commande.montant_ht + remise.commande.montant_tva,
        'pourcentage_remise': (remise.montant_remise / (remise.commande.montant_ht + remise.commande.montant_tva)) * 100 
            if (remise.commande.montant_ht + remise.commande.montant_tva) > 0 else 0,
        'nouveau_total': remise.commande.montant_total
    }

    return render(request, 'promotions/detail.html', {
        'remise': remise,
        'impact': impact,
        'entreprise': request.entreprise
    })

@login_required
@permission_required('STOCK.change_promotion', raise_exception=True)
def toggle_promotion(request, promotion_id):
    """Active/désactive une promotion"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    promotion = get_object_or_404(Promotion, id=promotion_id, entreprise=request.entreprise)
    promotion.actif = not promotion.actif
    promotion.save()
    
    action = "activée" if promotion.actif else "désactivée"
    messages.success(request, f"Promotion {action} avec succès")
    return redirect('liste_promotions')

@login_required
@permission_required('STOCK.delete_promotion', raise_exception=True)
def supprimer_promotion(request, promotion_id):
    """Supprime une promotion"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    promotion = get_object_or_404(Promotion, id=promotion_id, entreprise=request.entreprise)
    
    if request.method == 'POST':
        promotion.delete()
        messages.success(request, "Promotion supprimée avec succès")
        return redirect('liste_promotions')
    
    return render(request, 'promotions/supprimer.html', {
        'promotion': promotion,
        'entreprise': request.entreprise
    })
    
    


@login_required
def modifier_parametres(request):
    # Récupère ou crée les paramètres en associant l'utilisateur
    param, created = Parametre.objects.get_or_create(
        user=request.user,
        defaults={
            'nom_societe': 'Nom par défaut',
            'adresse': 'Adresse par défaut',
            'telephone': '+0000000000',
            'email': 'contact@example.com',
            'taux_tva': 20.0,
            'devise_principale': 'FC',  # Changé de 'devise' à 'devise_principale'
            'devises_acceptees': ['USD', 'EUR', 'CDF'],  # Ajouté
            'taux_change_auto': False  # Ajouté
        }
    )

    if request.method == 'POST':
        form_data = request.POST
        param.nom_societe = form_data.get('nom_societe', param.nom_societe)
        param.adresse = form_data.get('adresse', param.adresse)
        param.telephone = form_data.get('telephone', param.telephone)
        param.email = form_data.get('email', param.email)
        param.taux_tva = Decimal(form_data.get('taux_tva', param.taux_tva))  # Changé en Decimal
        
        # Gestion des devises
        param.devise_principale = form_data.get('devise_principale', param.devise_principale)  # Changé ici
        
        # Gestion des devises acceptées (convertir la chaîne en liste)
        devises_acceptees = form_data.get('devises_acceptees', '')
        if devises_acceptees:
            param.devises_acceptees = [d.strip() for d in devises_acceptees.split(',')]
        
        param.taux_change_auto = form_data.get('taux_change_auto') == 'on'  # Pour les checkbox
        
        if 'logo' in request.FILES:
            param.logo = request.FILES['logo']

        param.save()
        
        messages.success(request, "Paramètres mis à jour avec succès")
        return redirect('afficher_parametres')

    return render(request, 'parametres/edit_parametres.html', {'param': param})


def afficher_parametres(request):
    parametre, created = Parametre.objects.get_or_create(
        user=request.user,
        defaults={
            'nom_societe': 'Ma Société',
            # ... valeurs par défaut ...
        }
    )
    return render(request, 'parametres/parametres.html', {'parametre': parametre}) 




#IA PREDICATION

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from STOCK.ml.chatbot_engine import *


logger = logging.getLogger(__name__)

@csrf_exempt
@login_required
@require_http_methods(["POST"])
def chatbot_query(request):
    """
    Endpoint SaaS pour les requêtes chatbot avec isolation par entreprise
    """
    try:
        # Vérification de l'accès au chatbot
        if not hasattr(request.user, 'entreprise') and not request.user.is_saas_admin:
            raise PermissionDenied("Accès au chatbot non autorisé")

        query = request.POST.get('query', '').strip()
        if not query:
            return JsonResponse({'error': 'Requête vide'}, status=400)
        
        # Initialisation du moteur avec contexte SaaS
        engine = ChatbotEngine(
            user=request.user,
            enterprise=getattr(request.user, 'entreprise', None),
            is_saas_admin=request.user.is_saas_admin
        )
        
        # Traitement de la requête avec filtrage SaaS
        response = engine.process_query(query)
        
        # Enregistrement avec contexte multi-tenant
        ChatbotConversation.objects.create(
            user=request.user,
            entreprise=getattr(request.user, 'entreprise', None),
            query=query,
            response=response,
            metadata={
                'ip_address': request.META.get('REMOTE_ADDR'),
                'user_agent': request.META.get('HTTP_USER_AGENT'),
                'enterprise_id': getattr(request.user.entreprise, 'id', None)
            }
        )
        
        return JsonResponse({
            'response': response,
            'enterprise': getattr(request.user.entreprise, 'name', None)
        })
    
    except Exception as e:
        logger.error(f"Erreur chatbot SaaS - User: {request.user.id} - Error: {str(e)}")
        return JsonResponse({
            'error': 'Erreur de traitement',
            'details': str(e)
        }, status=500)

@login_required
def chatbot_view(request):
    """
    Vue SaaS pour l'interface chatbot avec filtrage par entreprise
    """
    # Vérification des permissions
    if not request.user.has_perm('ia.access_chatbot'):
        raise PermissionDenied("Vous n'avez pas accès au chatbot")
    
    # Filtrage des conversations selon le contexte SaaS
    conversations = ChatbotConversation.objects.filter(
        user=request.user
    ).select_related('entreprise').order_by('-timestamp')[:20]
    
    context = {
        'conversations': conversations,
        'chatbot_enabled': True,
        'is_saas_admin': request.user.is_saas_admin,
        'enterprise_name': getattr(request.user.entreprise, 'name', 'Compte SaaS')
    }
    
    return render(request, 'IA/chatbot_saas.html', context)

import logging
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.conf import settings
from django.core.cache import cache
from django.db.models import Avg, Sum
from decimal import Decimal
import json
import pandas as pd
from prophet import Prophet

def forecast_view(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    
    config = {
        'min_data_points': max(7, getattr(settings, 'FORECAST_MIN_DATA', 7)),
        'forecast_days': getattr(settings, 'FORECAST_DAYS', 30),
        'confidence_interval': getattr(settings, 'FORECAST_CONFIDENCE', 0.8),
    }
    
    try:
        # 1. Vérification des données historiques
        commandes = LigneCommandeClient.objects.filter(
            produit=produit,
            commande__vente_confirmee=True,
            quantite__gt=0
        ).values('commande__date_commande').annotate(
            total_quantite=Sum('quantite')
        ).order_by('commande__date_commande')
        
        data_points = len(commandes)
        
        # 2. Gestion du fallback si données insuffisantes
        if data_points < config['min_data_points']:
            fallback_reason = (
                f"Données insuffisantes (nécessaire: {config['min_data_points']}, "
                f"disponible: {data_points})"
            )
            return handle_forecast_fallback(
                request, 
                produit, 
                config,
                fallback_reason,
                data_points
            )
        
        # 3. Préparation des données pour Prophet
        try:
            df = pd.DataFrame([{
                'ds': cmd['commande__date_commande'].strftime('%Y-%m-%d'),
                'y': float(cmd['total_quantite'])
            } for cmd in commandes])
            
            df['ds'] = pd.to_datetime(df['ds'])
            
            # Vérification des dates uniques
            if df['ds'].nunique() < config['min_data_points']:
                raise ValueError("Pas assez de dates uniques pour l'analyse")
                
        except Exception as e:
            logger.error(f"Erreur préparation données: {str(e)}", exc_info=True)
            return handle_forecast_fallback(
                request,
                produit,
                config,
                f"Erreur préparation données: {str(e)}",
                data_points
            )
        
        # 4. Génération de la prévision avec Prophet
        try:
            model = Prophet(
                interval_width=config['confidence_interval'],
                daily_seasonality=False,
                weekly_seasonality=True,
                yearly_seasonality=False,
                changepoint_prior_scale=0.05  # Réduit la sensibilité aux changements
            )
            
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                model.fit(df)
            
            future = model.make_future_dataframe(
                periods=config['forecast_days'],
                freq='D'
            )
            forecast = model.predict(future)
            
            # Filtre seulement les futures prévisions
            last_historical_date = df['ds'].max()
            forecast = forecast[forecast['ds'] > last_historical_date]
            
            # 5. Calcul des métriques de performance
            try:
                from prophet.diagnostics import cross_validation, performance_metrics
                initial_days = max(30, len(df)-7)  # Correction ici
                df_cv = cross_validation(
                    model,
                    initial=f'{initial_days} days',  # Syntaxe corrigée
                    period='7 days',
                    horizon='7 days'
                )
                df_p = performance_metrics(df_cv)
                performance = {
                    'mdape': float(df_p['mdape'].mean()),
                    'mape': float(df_p['mape'].mean()),
                    'rmse': float(df_p['rmse'].mean()),
                    'interval': f"{int(config['confidence_interval']*100)}%"
                }
            except Exception as e:
                logger.warning(f"Erreur calcul performance: {str(e)}")
                performance = {'interval': f"{int(config['confidence_interval']*100)}%"}
            
            # 6. Formatage et enregistrement des résultats
            forecast_data = []
            for _, row in forecast.iterrows():
                forecast_data.append({
                    'date': row['ds'].strftime('%Y-%m-%d'),  # Format cohérent
                    'pred': round(float(row['yhat']), 2),
                    'min': round(float(row['yhat_lower']), 2),
                    'max': round(float(row['yhat_upper']), 2)
                })
                
              
                
                # Enregistrement en base si la date est dans le futur
                if row['ds'].date() > timezone.now().date():
                    HistoriquePrevision.objects.update_or_create(
                        produit=produit,
                        date_prevision=row['ds'].date(),
                        defaults={
                            'quantite_predite': pred,
                            'modele_utilise': 'prophet',
                            'parametres': {
                                'confidence': config['confidence_interval'],
                                'data_points': data_points,
                                'performance': performance
                            },
                            'precision': performance.get('mdape')
                        }
                    )
            
            # 7. Récupération de l'historique actualisé
            historique_previsions = HistoriquePrevision.objects.filter(
                produit=produit
            ).order_by('-date_prevision')[:30]
            
            return render(request, 'IA/forecast.html', {
                'produit': produit,
                'forecast_data': forecast_data,
                'performance': performance,
                'data_points': data_points,
                'historique_previsions': historique_previsions,
                'is_fallback': False
            })
            
        except Exception as e:
            logger.error(f"Erreur Prophet: {str(e)}", exc_info=True)
            return handle_forecast_fallback(
                request,
                produit,
                config,
                f"Erreur modélisation: {str(e)}",
                data_points
            )
            
    except Exception as e:
        logger.error(f"Erreur générale: {str(e)}", exc_info=True)
        return handle_forecast_fallback(
            request,
            produit,
            config,
            f"Erreur système: {str(e)}",
            0
        )
        
        
def handle_forecast_fallback(request, produit, config, reason, data_points):
    try:
        # 1. Calcul des statistiques de base
        stats = LigneCommandeClient.objects.filter(
            produit=produit,
            commande__vente_confirmee=True
        ).aggregate(
            avg=Avg('quantite'),
            total=Sum('quantite'),
            count=Count('id')
        )
        
        # 2. Calcul des ventes quotidiennes moyennes
        if stats['count'] > 3 and stats['avg']:
            daily_sales = float(stats['avg'])
        else:
            # Fallback basé sur le stock si pas assez de données
            daily_sales = max(
                0.1,  # Minimum de 0.1 unité/jour
                float(produit.stock) / 30.0 if produit.stock else 1.0
            )
        
        # 3. Génération des prévisions basiques
        forecast_data = []
        for i in range(1, config['forecast_days'] + 1):
            date = (timezone.now() + timedelta(days=i)).strftime('%Y-%m-%d')
            pred = round(daily_sales, 2)
            min_val = round(max(0.1, daily_sales * 0.7), 2)  # -30% avec minimum de 0.1
            max_val = round(daily_sales * 1.3, 2)            # +30%
            
            forecast_data.append((date, pred, min_val, max_val))
            
            # Enregistrement en base
            if i <= 30:  # Limite à 30 jours pour éviter la surcharge
                HistoriquePrevision.objects.update_or_create(
                    produit=produit,
                    date_prevision=timezone.now().date() + timedelta(days=i),
                    defaults={
                        'quantite_predite': pred,
                        'modele_utilise': 'moyenne',
                        'parametres': {
                            'fallback': True,
                            'avg_sales': daily_sales,
                            'data_points': data_points
                        }
                    }
                )
        
        # 4. Récupération de l'historique
        historique_previsions = HistoriquePrevision.objects.filter(
            produit=produit
        ).order_by('-date_prevision')[:30]
        
        return render(request, 'IA/forecast.html', {
            'produit': produit,
            'forecast_data': forecast_data,
            'performance': {'interval': '70-130%'},
            'data_points': data_points,
            'historique_previsions': historique_previsions,
            'is_fallback': True,
            'fallback_reason': reason,
            'daily_sales': daily_sales
        })
        
    except Exception as e:
        logger.critical(f"Erreur critique dans fallback: {str(e)}", exc_info=True)
        return render(request, 'IA/forecast.html', {
            'produit': produit,
            'error': "Erreur critique dans le système de prévision",
            'show_system_error': True
        })
        

# STOCK/views.py
def product_detail(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    return render(request, 'STOCK/product_detail.html', {'produit': produit})







# @login_required
# def retour_fournisseur(request, achat_id):
#     """
#     Gère les retours fournisseurs avec intégration SAAS
#     Version fonctionnelle avec vérification des permissions et filtrage par entreprise
#     """
    
#     # Vérification de l'entreprise de l'utilisateur
#     if not hasattr(request.user, 'entreprise') or not request.user.entreprise:
#         messages.error(request, "Votre compte n'est associé à aucune entreprise")
#         return redirect('dashboard')
    
#     user_entreprise = request.user.entreprise
    
#     # Récupération de l'achat avec vérification entreprise
#     achat = get_object_or_404(Achat, pk=achat_id, entreprise=user_entreprise)
    
#     # Récupération des paramètres SAAS
#     try:
#         config = ConfigurationSAAS.objects.get(entreprise=user_entreprise)
#         devise_principale = config.devise_principale
#         symbole_devise = devise_principale.symbole if devise_principale else '$'
#     except ConfigurationSAAS.DoesNotExist:
#         symbole_devise = '$'

#     # Gestion des requêtes POST (création de retour)
#     if request.method == 'POST':
#         return handle_retour_post(request, achat, user_entreprise)
    
#     # GET - Affichage du formulaire
#     retours_existants = RetourFournisseur.objects.filter(
#         achat=achat,
#         entreprise=user_entreprise
#     ).order_by('-date_retour')
    
#     context = {
#         'achat': achat,
#         'quantite_retournee_totale': achat.quantite_retournee,
#         'quantite_disponible': achat.quantite,
#         'retours_existants': retours_existants,
#         'quantite_effective': achat.quantite_effective,
#         'montant_effectif': achat.montant_effectif,
#         'can_delete': request.user.has_perm('STOCK.delete_retourfournisseur'),
#         'symbole_devise': symbole_devise,
#         'parametres': Parametre.objects.filter(entreprise=user_entreprise).first()
#     }
    
#     return render(request, 'achats/retour_fournisseur.html', context)

# def handle_retour_post(request, achat, user_entreprise):
#     """Gère la soumission du formulaire de retour"""
#     if not request.user.has_perm('STOCK.add_retourfournisseur'):
#         messages.error(request, "Vous n'avez pas la permission d'effectuer cette action")
#         return redirect('retour_fournisseur', achat_id=achat.id)
    
#     quantite_retournee = request.POST.get('quantite_retournee')
#     motif = request.POST.get('motif', '').strip()
    
#     # Validation des données
#     try:
#         quantite = int(quantite_retournee)
#         if quantite <= 0:
#             raise ValueError("La quantité doit être positive")
#         if quantite > achat.quantite_effective:
#             raise ValueError("La quantité dépasse le stock disponible")
#     except (ValueError, TypeError) as e:
#         messages.error(request, f"Erreur: {str(e)}")
#         return redirect('retour_fournisseur', achat_id=achat.id)
    
#     # Création du retour
#     try:
#         with transaction.atomic():
#             RetourFournisseur.objects.create(
#                 achat=achat,
#                 quantite_retournee=quantite,
#                 motif=motif,
#                 created_by=request.user,
#                 entreprise=user_entreprise
#             )
#             messages.success(request, f"Retour de {quantite} unités enregistré avec succès")
#     except Exception as e:
#         messages.error(request, f"Erreur lors de l'enregistrement: {str(e)}")
    
#     return redirect('retour_fournisseur', achat_id=achat.id)

# @login_required
# @permission_required('STOCK.delete_retourfournisseur', raise_exception=True)
# def supprimer_retour(request, retour_id):
#     """Supprime un retour avec vérification de l'entreprise"""
#     if not hasattr(request.user, 'entreprise'):
#         messages.error(request, "Problème de configuration du compte")
#         return redirect('dashboard')
    
#     retour = get_object_or_404(
#         RetourFournisseur, 
#         id=retour_id,
#         entreprise=request.user.entreprise
#     )
#     achat_id = retour.achat.id
    
#     if request.method == 'POST':
#         try:
#             with transaction.atomic():
#                 retour.delete()
#                 messages.success(request, "Retour supprimé avec succès")
#         except Exception as e:
#             messages.error(request, f"Erreur lors de la suppression: {str(e)}")
    
#     return redirect('retour_fournisseur', achat_id=achat_id)

# class RetourFournisseurDetailPDF(DetailView):
#     """Génération du PDF avec vérification SAAS"""
#     model = RetourFournisseur
#     template_name = 'achats/retour_fournisseur_pdf.html'
    
#     def get_object(self, queryset=None):
#         obj = super().get_object(queryset)
#         if not hasattr(self.request.user, 'entreprise') or obj.entreprise != self.request.user.entreprise:
#             raise Http404("Accès refusé")
#         return obj
    
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         retour = context['retour']
        
#         try:
#             config = ConfigurationSAAS.objects.get(entreprise=retour.entreprise)
#             context['symbole_devise'] = config.devise_principale.symbole if config.devise_principale else '$'
#         except ConfigurationSAAS.DoesNotExist:
#             context['symbole_devise'] = '$'
        
#         return context



from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import get_template
from weasyprint import HTML
@login_required
@permission_required('STOCK.add_bonlivraison', raise_exception=True)
def generer_bon_livraison(request, commande_id):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    commande = get_object_or_404(Commande, id=commande_id, entreprise=request.entreprise)

    if hasattr(commande, 'bon_livraison'):
        messages.warning(request, "Un Bon de Livraison existe déjà pour cette commande.")
        return redirect('liste_bons_livraison')

    bon_livraison = BonLivraison.objects.create(
        commande=commande,
        client=commande.client,
        cree_par=request.user,
        entreprise=request.entreprise,
    )

    messages.success(request, "Bon de Livraison créé avec succès.")
    return redirect('detail_bon_livraison', bl_id=bon_livraison.id)

@login_required
@permission_required('STOCK.view_bonlivraison', raise_exception=True)
def liste_bons_livraison(request):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    # Correction ici : suppression de config_saas du select_related
    bons = BonLivraison.objects.filter(entreprise=request.entreprise)\
                              .select_related('commande', 'client', 'cree_par', 'entreprise')\
                              .order_by('-date_creation')
    
    # Récupération séparée de la configuration SAAS
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale
    except ConfigurationSAAS.DoesNotExist:
        devise_principale = None

    return render(request, 'bons_livraison/liste.html', {
        'bons': bons,
        'entreprise': request.entreprise,
        'devise_principale': devise_principale
    })

@login_required
@permission_required('STOCK.view_bonlivraison', raise_exception=True)
def detail_bon_livraison(request, bl_id):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    bon = get_object_or_404(BonLivraison, id=bl_id, entreprise=request.entreprise)

    # Récupération de la configuration SAAS
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale
    except ConfigurationSAAS.DoesNotExist:
        devise_principale = None

    if request.method == 'POST' and request.user.has_perm('STOCK.change_bonlivraison'):
        bon.livre = request.POST.get('livre') == 'true'
        bon.date_livraison = request.POST.get('date_livraison') or None
        bon.save()
        messages.success(request, "Statut mis à jour")
        return redirect('liste_bons_livraison')

    return render(request, 'bons_livraison/detail_bon_livraison.html', {
        'bon': bon,
        'devise_principale': devise_principale,
        'can_change': request.user.has_perm('STOCK.change_bonlivraison')
    })

@login_required
@permission_required('STOCK.view_bonlivraison', raise_exception=True)
def bon_livraison_pdf(request, bl_id):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    bon = get_object_or_404(BonLivraison, id=bl_id, entreprise=request.entreprise)

    # Récupération de la configuration SAAS
    try:
        config_saas = ConfigurationSAAS.objects.get(entreprise=request.entreprise)
        devise_principale = config_saas.devise_principale
    except ConfigurationSAAS.DoesNotExist:
        devise_principale = None

    # Calcul des totaux
    lignes = bon.commande.lignes.all()
    quantite_totale = sum(ligne.quantite for ligne in lignes)
    montant_total = sum(ligne.quantite * ligne.prix_unitaire for ligne in lignes)

    # Formatage des montants
    if devise_principale and hasattr(devise_principale, 'formater_montant'):
        montant_formatted = devise_principale.formater_montant(montant_total)
    else:
        montant_formatted = f"{montant_total:.2f}"

    context = {
        'bon': bon,
        'lignes': lignes,
        'quantite_totale': quantite_totale,
        'montant_total': montant_total,
        'montant_formatted': montant_formatted,
        'devise_principale': devise_principale,
        'entreprise': request.entreprise
    }

    template = get_template('bons_livraison/pdf.html')
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename=BL_{bon.id}_{request.entreprise.nom}.pdf'
    
    HTML(string=html).write_pdf(response)
    return response



@login_required
@permission_required('STOCK.add_bonlivraison', raise_exception=True)
def choisir_commande_pour_bl(request):
    if not hasattr(request, 'entreprise') or not request.entreprise:
        messages.error(request, "Entreprise non définie")
        return redirect('security:acces_refuse')

    # Commandes sans bon de livraison
    commandes = CommandeClient.objects.filter(
        entreprise=request.entreprise
    ).exclude(
        bon_livraison__isnull=False
    ).select_related('client').order_by('-date_commande')

    # Pagination
    paginator = Paginator(commandes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'bons_livraison/choisir_commande.html', {
        'page_obj': page_obj,
        'entreprise': request.entreprise
    })
#comptabilites

@login_required
def liste_comptes(request):
    try:
        # Essayez d'obtenir les paramètres de l'entreprise
        entreprise = getattr(request.user, 'parametre', None)
        
        if entreprise is None:
            # Si aucun paramètre n'est associé, utilisez une entreprise par défaut
            entreprise = Parametre.objects.first()  # ou une autre logique
            
        comptes = Compte.objects.filter(entreprise=entreprise)
        
        for compte in comptes:
            compte.solde = compte.solde_actuel()
            
        return render(request, 'tresorerie/comptes/liste.html', {
            'comptes': comptes,
            'entreprise': entreprise
        })
        
    except Exception as e:
        # Journalisez l'erreur pour le débogage
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erreur dans liste_comptes: {str(e)}")
        
        # Retournez une réponse vide ou une page d'erreur
        return render(request, 'tresorerie/comptes/liste.html', {
            'comptes': [],
            'error': "Une erreur s'est produite lors du chargement des comptes."
        })

@login_required
def ajouter_compte(request):
    if request.method == 'POST':
        nom = request.POST.get('nom')
        type_compte = request.POST.get('type_compte')
        solde_initial = request.POST.get('solde_initial', 0)
        
        compte = Compte(
            nom=nom,
            type_compte=type_compte,
            solde_initial=solde_initial,
            entreprise=request.user.parametre
        )
        compte.save()
        return redirect('liste_comptes')
    
    return render(request, 'tresorerie/comptes/ajouter.html')

from django.contrib.auth.decorators import login_required
from django.db.models import Q, F, Sum, Case, When, DecimalField
from django.core.paginator import Paginator
from django.shortcuts import render
from django.utils import timezone


@login_required
def liste_transactions(request):
    entreprise = request.user.parametre
    transactions_list = Transaction.objects.filter(compte__entreprise=entreprise).order_by('-date_transaction')

    devise_affichee = request.session.get('devise_affichee', entreprise.devise_principale)

    devises_disponibles = set()
    devises_taux = TauxChange.objects.values_list('devise_source', 'devise_cible')
    for source, cible in devises_taux:
        devises_disponibles.add(source)
        devises_disponibles.add(cible)

    devises_disponibles.add(entreprise.devise_principale)
    if entreprise.devises_acceptees:
        devises_disponibles.update(entreprise.devises_acceptees)

    if devise_affichee not in devises_disponibles:
        devise_affichee = entreprise.devise_principale
        request.session['devise_affichee'] = devise_affichee

    search_query = request.GET.get('q', '')
    type_filter = request.GET.get('type', '')
    date_filter = request.GET.get('date', '')

    if search_query:
        transactions_list = transactions_list.filter(
            Q(description__icontains=search_query) |
            Q(compte__nom__icontains=search_query) |
            Q(categorie__nom__icontains=search_query)
        )

    if type_filter:
        transactions_list = transactions_list.filter(type_transaction=type_filter)

    if date_filter:
        transactions_list = transactions_list.filter(date_transaction__date=date_filter)

    paginator = Paginator(transactions_list, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    solde_total_principal = transactions_list.aggregate(
        total=Sum(
            Case(
                When(type_transaction='ENTREE', then=F('montant')),
                When(type_transaction='SORTIE', then=-F('montant')),
                default=0,
                output_field=DecimalField()
            )
        )
    )['total'] or 0

    transactions_formatted = []
    for transaction in page_obj.object_list:
        try:
            if devise_affichee == entreprise.devise_principale:
                montant_affiche = transaction.montant
            else:
                montant_affiche = entreprise.convertir_vers_devise_affichee(transaction.montant, devise_affichee)

            transactions_formatted.append({
                'obj': transaction,
                'montant_affiche': entreprise.format_devise(montant_affiche, devise_affichee),
                'montant_original': entreprise.format_devise(transaction.montant, entreprise.devise_principale)
                    if devise_affichee != entreprise.devise_principale else None
            })
        except Exception as e:
            print(f"Erreur conversion transaction {transaction.id}: {e}")
            transactions_formatted.append({
                'obj': transaction,
                'montant_affiche': entreprise.format_devise(transaction.montant, entreprise.devise_principale),
                'montant_original': None
            })

    try:
        if devise_affichee == entreprise.devise_principale:
            solde_total = solde_total_principal
        else:
            solde_total = entreprise.convertir_vers_devise_affichee(solde_total_principal, devise_affichee)
    except Exception as e:
        print(f"Erreur conversion solde: {e}")
        solde_total = solde_total_principal
        devise_affichee = entreprise.devise_principale

    solde_total_formatted = entreprise.format_devise(solde_total, devise_affichee)
    solde_total_principal_formatted = (
        entreprise.format_devise(solde_total_principal, entreprise.devise_principale)
        if devise_affichee != entreprise.devise_principale else None
    )

    return render(request, 'tresorerie/transactions/liste.html', {
        'page_obj': page_obj,
        'transactions': transactions_formatted,
        'solde_total_formatted': solde_total_formatted,
        'solde_total_principal_formatted': solde_total_principal_formatted,
        'devise_affichee': devise_affichee,
        'devises_disponibles': sorted(devises_disponibles),
        'devise_principale': entreprise.devise_principale,
        'search_query': search_query,
        'type_filter': type_filter,
        'date_filter': date_filter,
    })



@login_required
def ajouter_transaction(request):
    entreprise = request.user.parametre
    if request.method == 'POST':
        compte_id = request.POST.get('compte')
        type_transaction = request.POST.get('type_transaction')
        categorie_id = request.POST.get('categorie')
        montant = request.POST.get('montant')
        mode_paiement = request.POST.get('mode_paiement')
        description = request.POST.get('description', '')
        
        transaction = Transaction(
            compte_id=compte_id,
            type_transaction=type_transaction,
            categorie_id=categorie_id,
            montant=montant,
            mode_paiement=mode_paiement,
            description=description,
            utilisateur=request.user
        )
        
        if 'piece_jointe' in request.FILES:
            transaction.piece_jointe = request.FILES['piece_jointe']
        
        transaction.save()
        return redirect('liste_transactions')
    
    comptes = Compte.objects.filter(entreprise=entreprise)
    categories = CategorieTransaction.objects.filter(entreprise=entreprise)
    return render(request, 'tresorerie/transactions/ajouter.html', {
        'comptes': comptes,
        'categories': categories
    })


@login_required
def journal_caisse(request):
    entreprise = request.user.parametre
    transactions_list = Transaction.objects.filter(
        compte__entreprise=entreprise
    ).order_by('-date_transaction')
    
    # Filtres
    date_debut = request.GET.get('date_debut', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_fin = request.GET.get('date_fin', datetime.now().strftime('%Y-%m-%d'))
    compte_id = request.GET.get('compte')
    devise_affichee = request.GET.get('devise', entreprise.devise_principale)
    
    transactions_list = transactions_list.filter(
        date_transaction__date__gte=date_debut,
        date_transaction__date__lte=date_fin
    )
    
    if compte_id:
        transactions_list = transactions_list.filter(compte_id=compte_id)
    
    # Préparer les transactions avec les montants convertis
    transactions_converted = []
    for transaction in transactions_list:
        montant_converti = entreprise.convertir_vers_devise_affichee(
            transaction.montant, 
            devise_affichee
        )
        transactions_converted.append({
            'original': transaction,
            'montant_converti': montant_converti,
            'devise_affichee': devise_affichee
        })
    
    # Pagination
    paginator = Paginator(transactions_converted, 25)  # Utiliser la liste convertie
    page_number = request.GET.get('page')
    transactions_page = paginator.get_page(page_number)
    
    # Conversion des totaux (reste inchangé)
    total_entrees = transactions_list.filter(type_transaction='ENTREE').aggregate(
        total=Sum('montant')
    )['total'] or 0
    total_entrees = entreprise.convertir_vers_devise_affichee(total_entrees, devise_affichee)
    
    total_sorties = transactions_list.filter(type_transaction='SORTIE').aggregate(
        total=Sum('montant')
    )['total'] or 0
    total_sorties = entreprise.convertir_vers_devise_affichee(total_sorties, devise_affichee)
    
    return render(request, 'tresorerie/rapports/journal_caisse.html', {
        'transactions_page': transactions_page,  # Utiliser la page convertie
        'total_entrees': total_entrees,
        'total_sorties': total_sorties,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'comptes': Compte.objects.filter(entreprise=entreprise),
        'devises_disponibles': entreprise.devises_acceptees if entreprise.devises_acceptees else [entreprise.devise_principale],
        'devise_affichee': devise_affichee,
        'devise_principale': entreprise.devise_principale
    })
  

@login_required
def detail_compte(request, pk):
    compte = get_object_or_404(Compte, pk=pk, entreprise=request.user.parametre)
    transactions = Transaction.objects.filter(compte=compte).order_by('-date_transaction')[:10]
    return render(request, 'tresorerie/comptes/detail.html', {
        'compte': compte,
        'transactions': transactions,
        'solde_actuel': compte.solde_actuel()
    })






@login_required
def detail_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, compte__entreprise=request.user.parametre)
    return render(request, 'tresorerie/transactions/detail.html', {'transaction': transaction})

@login_required
def liste_categories(request):
    categories = CategorieTransaction.objects.filter(entreprise=request.user.parametre)
    return render(request, 'tresorerie/categories/liste.html', {'categories': categories})

@login_required
def ajouter_categorie(request):
    if request.method == 'POST':
        nom = request.POST.get('nom')
        type_categorie = request.POST.get('type_categorie')
        description = request.POST.get('description', '')
        
        categorie = CategorieTransaction(
            nom=nom,
            type_categorie=type_categorie,
            description=description,
            entreprise=request.user.parametre
        )
        categorie.save()
        return redirect('liste_categories')
    
    return render(request, 'tresorerie/categories/ajouter.html')

@login_required
def balance_comptes(request):
    entreprise = request.user.parametre
    date_debut = request.GET.get('date_debut', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_fin = request.GET.get('date_fin', datetime.now().strftime('%Y-%m-%d'))
    
    comptes = Compte.objects.filter(entreprise=entreprise)
    balance = []
    
    for compte in comptes:
        transactions = Transaction.objects.filter(
            compte=compte,
            date_transaction__date__range=[date_debut, date_fin]
        )
        
        entree = transactions.filter(type_transaction='ENTREE').aggregate(total=Sum('montant'))['total'] or 0
        sortie = transactions.filter(type_transaction='SORTIE').aggregate(total=Sum('montant'))['total'] or 0
        
        balance.append({
            'compte': compte,
            'entree': entree,
            'sortie': sortie,
            'solde': entree - sortie
        })
    
    return render(request, 'tresorerie/rapports/balance.html', {
        'balance': balance,
        'date_debut': date_debut,
        'date_fin': date_fin
    })

@login_required
def bilan_simplifie(request):
    entreprise = request.user.parametre
    date_fin = request.GET.get('date_fin', datetime.now().strftime('%Y-%m-%d'))
    
    # Actifs (comptes avec solde positif)
    actifs = Compte.objects.filter(
        entreprise=entreprise
    ).annotate(
        solde=Sum(
            Case(
                When(transaction__type_transaction='ENTREE', then='transaction__montant'),
                When(transaction__type_transaction='SORTIE', then=-F('transaction__montant')),
                default=0,
                output_field=DecimalField()
            )
        ) + F('solde_initial')
    ).filter(solde__gt=0)
    
    # Passifs (comptes avec solde négatif)
    passifs = Compte.objects.filter(
        entreprise=entreprise
    ).annotate(
        solde=Sum(
            Case(
                When(transaction__type_transaction='ENTREE', then='transaction__montant'),
                When(transaction__type_transaction='SORTIE', then=-F('transaction__montant')),
                default=0,
                output_field=DecimalField()
            )
        ) + F('solde_initial')
    ).filter(solde__lt=0).annotate(
        solde_abs=Abs(F('solde'))
    )
    
    total_actifs = actifs.aggregate(total=Sum('solde'))['total'] or 0
    total_passifs = - (passifs.aggregate(total=Sum('solde'))['total'] or 0)
    
    return render(request, 'tresorerie/rapports/bilan.html', {
        'actifs': actifs,
        'passifs': passifs,
        'total_actifs': total_actifs,
        'total_passifs': total_passifs,
        'date_fin': date_fin
    })
    


import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from django.contrib.auth.decorators import login_required
from openpyxl.styles import Font, PatternFill

@login_required
def export_journal_excel(request):
    entreprise = request.user.parametre
    
    # Récupérer les mêmes filtres que dans journal_caisse
    date_debut = request.GET.get('date_debut', (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d'))
    date_fin = request.GET.get('date_fin', timezone.now().strftime('%Y-%m-%d'))
    compte_id = request.GET.get('compte')
    
    transactions = Transaction.objects.filter(
        compte__entreprise=entreprise,
        date_transaction__date__gte=date_debut,
        date_transaction__date__lte=date_fin
    ).order_by('-date_transaction')
    
    if compte_id:
        transactions = transactions.filter(compte_id=compte_id)
    
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal de Caisse"
    
    # En-têtes
    headers = [
        "Date", 
        "Compte", 
        "Type", 
        "Catégorie", 
        "Description", 
        "Mode Paiement", 
        "Montant"
    ]
    ws.append(headers)
    
    # Données
    for t in transactions:
        ws.append([
            t.date_transaction.strftime("%d/%m/%Y %H:%M"),
            t.compte.nom,
            t.get_type_transaction_display(),
            t.categorie.nom if t.categorie else "-",
            t.description,
            t.get_mode_paiement_display(),
            t.montant
        ])
    
    # Style des en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format de la date
    for row in ws.iter_rows(min_row=2):
        row[0].number_format = 'DD/MM/YYYY HH:MM'
    
    # Format monétaire
    for row in ws.iter_rows(min_row=2):
        row[6].number_format = '#,##0.00'
    
    # Ajouter les totaux
    ws.append([])  # Ligne vide
    
    # Calcul des totaux
    total_entrees = sum(t.montant for t in transactions if t.type_transaction == 'ENTREE')
    total_sorties = sum(t.montant for t in transactions if t.type_transaction == 'SORTIE')
    solde_net = total_entrees - total_sorties
    
    # Ajout des totaux
    ws.append(["", "", "", "", "", "Total Entrées:", total_entrees])
    ws.append(["", "", "", "", "", "Total Sorties:", total_sorties])
    ws.append(["", "", "", "", "", "Solde Net:", solde_net])
    
    # Style des totaux
    for row in ws.iter_rows(min_row=ws.max_row-2, max_row=ws.max_row):
        for cell in row[5:]:
            cell.font = Font(bold=True)
        row[-1].number_format = '#,##0.00'
    
    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"journal_caisse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return responses





from django.shortcuts import render
from django.db.models import Sum, Count
from datetime import datetime, timedelta

def dashboard_tresorerie(request):
    # Paramètre de période
    jours = int(request.GET.get('jours', 30))
    date_debut = datetime.now() - timedelta(days=jours)
    
    # 1. Flux de trésorerie (entrées/sorties par jour)
    transactions = Transaction.objects.filter(
        date_transaction__gte=date_debut
    ).values('date_transaction__date').annotate(
        entieres=Sum('montant', filter=Q(type_transaction='ENTREE')),
        sorties=Sum('montant', filter=Q(type_transaction='SORTIE'))
    ).order_by('date_transaction__date')
    
    dates_flux = []
    entieres_flux = []
    sorties_flux = []
    
    for t in transactions:
        dates_flux.append(t['date_transaction__date'].strftime('%d/%m'))
        entieres_flux.append(float(t['entieres'] or 0))
        sorties_flux.append(float(t['sorties'] or 0))
    
    # 2. Répartition des dépenses par catégorie
    categories_depenses = CategorieTransaction.objects.filter(
        type_categorie='DEP'
    ).annotate(
        total=Sum('transaction__montant', filter=Q(
            transaction__type_transaction='SORTIE',
            transaction__date_transaction__gte=date_debut
        ))
    ).exclude(total=None).order_by('-total')
    
    labels_depenses = []
    data_depenses = []
    colors_depenses = []
    color_palette = [
        '#e74a3b', '#f6c23e', '#36b9cc', '#858796', 
        '#1cc88a', '#4e73df', '#5a5c69', '#fd7e14'
    ]
    
    for i, categorie in enumerate(categories_depenses):
        labels_depenses.append(categorie.nom)
        data_depenses.append(float(categorie.total))
        colors_depenses.append(color_palette[i % len(color_palette)])
    
    # 3. Soldes des comptes
    comptes = Compte.objects.all().annotate(
        solde=Sum(
            Case(
                When(transaction__type_transaction='ENTREE', then='transaction__montant'),
                When(transaction__type_transaction='SORTIE', then=-F('transaction__montant')),
                default=0,
                output_field=DecimalField()
            )
        ) + F('solde_initial')
    ).order_by('-solde')
    
    labels_comptes = [c.nom for c in comptes]
    data_comptes = [float(c.solde) for c in comptes]
    
    # 4. Répartition des moyens de paiement
    modes_paiement = Transaction.objects.filter(
        date_transaction__gte=date_debut
    ).values('mode_paiement').annotate(
        total=Sum('montant'),
        count=Count('id')
    ).order_by('-total')
    
    labels_paiements = []
    data_paiements = []
    colors_paiements = []
    mode_colors = {
        'ESP': '#4e73df',  # Espèces - bleu
        'MM': '#1cc88a',   # Mobile Money - vert
        'CHQ': '#f6c23e',  # Chèque - jaune
        'VIR': '#36b9cc',  # Virement - cyan
        'CRD': '#e74a3b'   # Carte - rouge
    }
    
    for mode in modes_paiement:
        labels_paiements.append(dict(Transaction.MODE_PAIEMENT).get(mode['mode_paiement']))
        data_paiements.append(float(mode['total']))
        colors_paiements.append(mode_colors.get(mode['mode_paiement'], '#858796'))
    
    return render(request, 'tresorerie/dashboard.html', {
        'dates_flux': dates_flux,
        'entieres_flux': entieres_flux,
        'sorties_flux': sorties_flux,
        'categories_depenses': zip(categories_depenses, colors_depenses),
        'labels_depenses': labels_depenses,
        'data_depenses': data_depenses,
        'colors_depenses': colors_depenses,
        'labels_comptes': labels_comptes,
        'data_comptes': data_comptes,
        'labels_paiements': labels_paiements,
        'data_paiements': data_paiements,
        'colors_paiements': colors_paiements,
        'modes_paiement': [{'mode': m[1], 'color': mode_colors.get(m[0], '#858796')} 
                          for m in Transaction.MODE_PAIEMENT]
    })
    
    
def liste_paiements(request):
    paiements = Paiement.objects.all().order_by('-date').select_related('facture', 'facture__commande')
    for paiement in paiements:
        paiement.transaction = Transaction.objects.filter(
            commande=paiement.facture.commande,
            montant=paiement.montant
        ).first()
    return render(request, 'tresorerie/paiements/liste.html', {'paiements': paiements})

def detail_paiement(request, pk):
    paiement = get_object_or_404(Paiement, pk=pk)
    transaction = paiement.transaction_set.first()
    return render(request, 'tresorerie/paiements/detail.html', {
        'paiement': paiement,
        'transaction': transaction
    })
from django.shortcuts import render, redirect, get_object_or_404


def gestion_stock_et_produit(request):
    return render(request,'Gestion des produits et stock/liste.html')

def venteModule(request):
    return render(request,'Vente au comptoir ou en ligne/liste.html')



# views.py
from django.shortcuts import render
from .models import  Parametre, TauxChange
from decimal import Decimal, InvalidOperation

def get_devise_affichee():
    """Retourne la devise principale depuis les paramètres."""
    parametre = Parametre.objects.first()
    return parametre.devise_principale if parametre else 'CDF'


def get_taux(source, cible):
    """Retourne le taux de conversion entre deux devises."""
    if source == cible:
        return Decimal('1.0')

    try:
        # Taux direct
        taux = (
            TauxChange.objects
            .filter(devise_source=source, devise_cible=cible)
            .order_by('-date_mise_a_jour')
            .first()
        )
        if taux:
            return taux.taux

        # Taux inverse si le direct n'existe pas
        taux_inverse = (
            TauxChange.objects
            .filter(devise_source=cible, devise_cible=source)
            .order_by('-date_mise_a_jour')
            .first()
        )
        if taux_inverse and taux_inverse.taux != 0:
            return Decimal('1.0') / taux_inverse.taux

    except (TauxChange.DoesNotExist, InvalidOperation):
        pass

    return Decimal('1.0')  # Valeur de secours


@require_POST
@login_required
def set_devise(request):
    devise = request.POST.get('devise')
    print(f"Devise reçue: {devise}")  # Debug
    try:
        parametres = Parametre.objects.get(user=request.user)
        if devise in parametres.devises_acceptees:
            request.session['devise_affichee'] = devise
            print("Devise enregistrée en session")  # Debug
            return JsonResponse({'status': 'success'})
    except Parametre.DoesNotExist:
        if devise in ["USD", "CDF"]:  # Simplifié
            request.session['devise_affichee'] = devise
            return JsonResponse({'status': 'success'})
    
    return JsonResponse({'status': 'error', 'message': 'Devise non autorisée'}, status=400)





@login_required
def obtenir_taux_change(request):
    devise_cible = request.GET.get('devise')
    parametres = Parametre.objects.first()

    if not parametres or not devise_cible:
        return JsonResponse({'status': 'error', 'message': 'Paramètres manquants'})

    try:
        devise_source = parametres.devise_principale
        taux = TauxChange.get_taux(devise_source, devise_cible)

        if taux is None:
            return JsonResponse({'status': 'error', 'message': f"Aucun taux de change disponible pour {devise_source} vers {devise_cible}"})

        return JsonResponse({'status': 'success', 'taux': str(taux)})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})





    
    
    
    
    
    #Module tresorerie
    
def Tresorerie(request):
    return render(request,"tresorerie/module_tresorerie.html")

import requests
class MettreAJourTauxChange(View):
    def get(self, request):
        parametres = Parametre.objects.first()
        if not parametres or not parametres.devises_acceptees:
            return JsonResponse({'status': 'error', 'message': 'Configuration manquante'})
        
        try:
            # Exemple avec une API fictive - à remplacer par une vraie API
            response = requests.get(
                f"https://api.exemple.com/taux?devises={','.join(parametres.devises_acceptees)}"
            )
            data = response.json()
            
            for devise in parametres.devises_acceptees:
                if devise == parametres.devise_principale:
                    continue
                
                taux, created = TauxChange.objects.update_or_create(
                    devise_source=devise,
                    devise_cible=parametres.devise_principale,
                    defaults={
                        'taux': Decimal(str(data['rates'][devise])),
                        'est_manuel': False
                    }
                )
            
            return JsonResponse({
                'status': 'success',
                'message': f'Taux mis à jour le {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
        
        
        
        
        
from django.shortcuts import render, redirect
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy
from .models import Parametre, TauxChange
from .forms import DeviseForm, TauxChangeForm
import requests
from decimal import Decimal
from django.http import JsonResponse

class ListeDevisesView(ListView):
    model = TauxChange
    template_name = 'tresorerie/devises/liste.html'
    context_object_name = 'taux_list'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['parametres'] = Parametre.objects.first()
        return context

def ajouter_devise(request):
    parametres = Parametre.objects.first()
    if request.method == 'POST':
        form = DeviseForm(request.POST, instance=parametres)
        if form.is_valid():
            form.save()
            messages.success(request, "Devises mises à jour avec succès")
            return redirect('liste_devises')
    else:
        form = DeviseForm(instance=parametres)
    
    return render(request, 'tresorerie/devises/ajouter_devise.html', {'form': form})

from django.urls import reverse_lazy

def ajouter_taux(request):
    # Récupérer les devises disponibles
    parametres = Parametre.objects.first()
    devises_disponibles = []
    if parametres:
        devises_disponibles = sorted(set(
            parametres.devises_acceptees + [parametres.devise_principale]
        ))

    if request.method == 'POST':
        # Récupérer les données du formulaire
        devise_source = request.POST.get('devise_source')
        devise_cible = request.POST.get('devise_cible')
        taux_str = request.POST.get('taux')

        # Validation manuelle
        errors = []
        if not devise_source or not devise_cible or not taux_str:
            errors.append("Tous les champs sont obligatoires")
        
        if devise_source == devise_cible:
            errors.append("Les devises source et cible doivent être différentes")

        try:
            taux = Decimal(taux_str)
            if taux <= 0:
                errors.append("Le taux doit être positif")
        except:
            errors.append("Le taux doit être un nombre valide")

        # Vérifier si le taux existe déjà
        if TauxChange.objects.filter(
            devise_source=devise_source,
            devise_cible=devise_cible
        ).exists():
            errors.append("Ce taux de change existe déjà")

        if not errors:
            # Créer et sauvegarder le nouveau taux
            TauxChange.objects.create(
                devise_source=devise_source,
                devise_cible=devise_cible,
                taux=taux,
                est_manuel=True
            )
            messages.success(
                request,
                f"Taux de change {devise_source}/{devise_cible} ajouté avec succès"
            )
            return redirect(reverse_lazy('liste_devises'))

        # Si erreurs, les afficher
        for error in errors:
            messages.error(request, error)

    context = {
        'devises_disponibles': devises_disponibles,
        'page_title': "Ajouter un Taux de Change",
        'active_menu': 'tresorerie'
    }
    return render(request, 'tresorerie/devises/ajouter_taux.html', context)


from django.core.paginator import Paginator

def historique_taux(request):
    """
    Affiche l'historique des modifications des taux de change
    avec pagination et filtres possibles.
    """
    # Récupérer les paramètres de filtrage
    devise_source = request.GET.get('devise_source')
    devise_cible = request.GET.get('devise_cible')
    
    # Construire la requête de base
    queryset = TauxChange.objects.all().order_by('-date_mise_a_jour')
    
    # Appliquer les filtres
    if devise_source:
        queryset = queryset.filter(devise_source=devise_source)
    if devise_cible:
        queryset = queryset.filter(devise_cible=devise_cible)
    
    # Pagination (10 éléments par page)
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Récupérer les devises disponibles pour les filtres
    parametres = Parametre.objects.first()
    devises_disponibles = []
    if parametres:
        devises_disponibles = parametres.devises_acceptees + [parametres.devise_principale]
    
    context = {
        'page_obj': page_obj,
        'devises_disponibles': sorted(set(devises_disponibles)),
        'selected_devise_source': devise_source,
        'selected_devise_cible': devise_cible,
        'page_title': "Historique des Taux de Change",
        'active_menu': 'tresorerie'
    }
    
    return render(request, 'tresorerie/devises/historique_taux.html', context)
    
    
    
    


from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.cache import cache
from .models import Parametre, TauxChange
import requests
import json
from decimal import Decimal

def get_available_currencies(parametres):
    """
    Récupère les devises disponibles depuis l'API ou le cache
    """
    if not parametres.openexchangerates_app_id:
        return {}

    cache_key = f'oer_currencies_{parametres.user.id}'
    currencies = cache.get(cache_key)
    
    if not currencies:
        try:
            response = requests.get(
                'https://openexchangerates.org/api/currencies.json',
                params={'app_id': parametres.openexchangerates_app_id},
                timeout=5
            )
            currencies = response.json()
            cache.set(cache_key, currencies, timeout=86400)  # Cache 24h
        except Exception:
            currencies = {}

    return currencies


#modif taux
def modifier_taux(request, taux_id):
    taux = get_object_or_404(TauxChange, id=taux_id)
    parametres = Parametre.objects.first()  # Récupération des paramètres
    
    if request.method == 'POST':
        # Récupération des données du formulaire
        devise_source = request.POST.get('devise_source')
        devise_cible = request.POST.get('devise_cible')
        taux_value = request.POST.get('taux')
        
        # Validation basique
        errors = []
        if not devise_source or not devise_cible or not taux_value:
            errors.append("Tous les champs sont obligatoires")
        
        if devise_source == devise_cible:
            errors.append("Les devises source et cible doivent être différentes")

        try:
            taux_decimal = Decimal(taux_value)
            if taux_decimal <= 0:
                errors.append("Le taux doit être positif")
        except:
            errors.append("Le taux doit être un nombre valide")

        if not errors:
            try:
                taux.devise_source = devise_source
                taux.devise_cible = devise_cible
                taux.taux = taux_decimal
                taux.save()
                messages.success(request, "Taux mis à jour avec succès")
                return redirect('liste_devises')
            except Exception as e:
                messages.error(request, f"Erreur: {str(e)}")
        else:
            for error in errors:
                messages.error(request, error)
    
    # Préparer les devises disponibles
    devises_disponibles = []
    if parametres:
        devises_disponibles = sorted(set(
            parametres.devises_acceptees + [parametres.devise_principale]
        ))
    
    context = {
        'taux': taux,
        'devises_disponibles': devises_disponibles,
        'page_title': "Modifier Taux de Change",
        'parametres': parametres  # Ajout des paramètres au contexte
    }
    return render(request, 'tresorerie/devises/modifier_taux.html', context)


def supprimer_taux(request, taux_id):
    taux = get_object_or_404(TauxChange, id=taux_id)
    
    if request.method == 'POST':
        taux.delete()
        messages.success(request, "Taux supprimé avec succès")
        return redirect('liste_devises')
    
    context = {
        'taux': taux,
        'page_title': "Confirmer la suppression"
    }
    return render(request, 'tresorerie/devises/confirmer_suppression.html', context)


def obtenir_devises_disponibles(request):
    if request.method == 'GET':
        try:
            parametres = Parametre.objects.first()
            if not parametres or not parametres.openexchangerates_app_id:
                return JsonResponse({'status': 'error', 'message': 'Configuration manquante'})
            
            response = requests.get(
                f"https://openexchangerates.org/api/currencies.json",
                params={'app_id': parametres.openexchangerates_app_id}
            )
            
            if response.status_code == 200:
                return JsonResponse({
                    'status': 'success',
                    'devises': response.json()
                })
            else:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Erreur lors de la récupération des devises'
                })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée'})



@csrf_exempt  # nécessaire si tu n'utilises pas {% csrf_token %}, mais tu l'utilises donc à retirer si non requis
def changer_devise(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nouvelle_devise = data.get('devise')

            if nouvelle_devise in ['USD', 'EUR', 'CDF']:  # sécurise les entrées
                parametre = Parametre.objects.first()
                if parametre:
                    parametre.devise_principale = nouvelle_devise
                    parametre.save()
                else:
                    Parametre.objects.create(devise_principale=nouvelle_devise)

                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'message': 'Devise invalide'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})



def liste_devises(request):
    try:
        parametres, _ = Parametre.objects.get_or_create(
            user=request.user,
            defaults={
                "nom_societe": "Ma Société",
                "devise_principale": "USD",
                "devises_acceptees": ["USD", "EUR", "CDF"],
                "openexchangerates_app_id": "",
                "openexchangerates_base_devise": "USD"
            }
        )
    except Exception as e:
        messages.error(request, str(e))
        return redirect('home')

    taux_list = TauxChange.objects.all()

    context = {
        'parametres': parametres,
        'taux_list': taux_list,
        'page_title': "Gestion des Devises",
        'devises_disponibles': json.dumps(get_available_currencies(parametres))
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'tresorerie/devises/partials/rates_table.html', context)

    return render(request, 'tresorerie/devises/liste_devises.html', context)


from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.utils import timezone
from decimal import Decimal
import requests
from .models import Parametre, TauxChange

@require_http_methods(["POST"])
def maj_taux_auto(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'Non authentifié'}, status=403)

    try:
        parametres = Parametre.objects.get(user=request.user)
        if not parametres.openexchangerates_app_id:
            return JsonResponse({'status': 'error', 'message': 'Clé API non définie'}, status=400)

        response = requests.get(
            'https://openexchangerates.org/api/latest.json',
            params={
                'app_id': parametres.openexchangerates_app_id,
                'base': parametres.openexchangerates_base_devise,
                'symbols': ','.join(set(parametres.devises_acceptees + [parametres.devise_principale]))
            },
            timeout=10
        )

        if response.status_code != 200:
            return JsonResponse({'status': 'error', 'message': 'Erreur API'}, status=400)

        data = response.json()

        # Si on veut juste tester l’API sans modifier la base
        if request.POST.get("test_only"):
            return JsonResponse({'status': 'success', 'message': 'Connexion API réussie'})

        base_rate = data['rates'].get(parametres.devise_principale, 1)
        updated_count = 0

        for devise, taux in data['rates'].items():
            if devise == parametres.devise_principale:
                continue

            taux_converti = Decimal(str(base_rate / taux)) \
                if parametres.openexchangerates_base_devise != parametres.devise_principale \
                else Decimal(str(taux))

            TauxChange.objects.update_or_create(
                devise_source=devise,
                devise_cible=parametres.devise_principale,
                defaults={'taux': taux_converti, 'est_manuel': False}
            )
            updated_count += 1

        return JsonResponse({
            'status': 'success',
            'message': f"{updated_count} taux mis à jour",
            'last_update': timezone.now().isoformat()
        })

    except requests.Timeout:
        return JsonResponse({'status': 'error', 'message': 'Délai API dépassé'}, status=408)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def modifier_taux(request, taux_id):
    taux = get_object_or_404(TauxChange, id=taux_id)
    parametres = Parametre.objects.first()

    if request.method == 'POST':
        devise_source = request.POST.get('devise_source')
        devise_cible = request.POST.get('devise_cible')
        taux_value = request.POST.get('taux')

        errors = []
        if not devise_source or not devise_cible or not taux_value:
            errors.append("Tous les champs sont obligatoires")
        if devise_source == devise_cible:
            errors.append("Les devises doivent être différentes")

        try:
            taux_decimal = Decimal(taux_value)
            if taux_decimal <= 0:
                errors.append("Le taux doit être positif")
        except:
            errors.append("Taux invalide")

        if not errors:
            taux.devise_source = devise_source
            taux.devise_cible = devise_cible
            taux.taux = taux_decimal
            taux.est_manuel = True
            taux.save()
            messages.success(request, "Taux mis à jour avec succès")
            return redirect('liste_devises')
        else:
            for err in errors:
                messages.error(request, err)

    context = {
        'taux': taux,
        'devises_disponibles': sorted(set(parametres.devises_acceptees + [parametres.devise_principale])),
        'page_title': "Modifier Taux de Change"
    }
    return render(request, 'tresorerie/devises/modifier_taux.html', context)

def supprimer_taux(request, taux_id):
    taux = get_object_or_404(TauxChange, id=taux_id)
    if request.method == 'POST':
        taux.delete()
        messages.success(request, "Taux supprimé avec succès")
        return redirect('liste_devises')
    return render(request, 'tresorerie/devises/confirmer_suppression.html', {'taux': taux})




#reception de stocks
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.forms import formset_factory
from django.http import JsonResponse
from decimal import Decimal
from .models import Parametre, TauxChange


    
# views.py (ou wherever se trouve ta vue manual_backup)
import os
import json
import io
import zipfile
from datetime import datetime
from django.conf import settings
from django.http import JsonResponse, HttpResponseForbidden, FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.management import call_command
from django.contrib.auth.decorators import login_required
import logging
logger = logging.getLogger(__name__)

def get_backup_dir(entreprise):
    """Retourne le chemin du dossier de backup spécifique à l'entreprise"""
    try:
        # Vérifie que l'entreprise a un ID valide
        if not entreprise or not entreprise.id:
            raise ValueError("Entreprise ou ID d'entreprise invalide")
            
        # Utilise BACKUP_DIR des settings ou un dossier par défaut
        backup_base_dir = getattr(settings, 'BACKUP_DIR', os.path.join(settings.BASE_DIR, 'backups'))
        entreprise_dir = os.path.join(backup_base_dir, str(entreprise.id))
        
        # Crée le dossier s'il n'existe pas
        os.makedirs(entreprise_dir, exist_ok=True)
        
        # Vérifie les permissions d'écriture
        if not os.access(entreprise_dir, os.W_OK):
            raise PermissionError(f"Pas de permission d'écriture dans {entreprise_dir}")
            
        return entreprise_dir
        
    except Exception as e:
        logger.error(f"Erreur dans get_backup_dir: {str(e)}")
        raise

@csrf_exempt
@login_required
def manual_backup(request):
    """Sauvegarde manuelle des données avec isolation SaaS"""
    try:
        # Vérification de l'entreprise
        if not hasattr(request, 'entreprise') or not request.entreprise:
            logger.error("Entreprise non définie dans la requête")
            return JsonResponse({
                'status': 'error', 
                'message': 'Entreprise non définie'
            }, status=400)

        if request.method != 'POST':
            return JsonResponse({
                'error': 'Méthode non autorisée'
            }, status=405)

        # Configuration du chemin de sauvegarde
        backup_dir = get_backup_dir(request.entreprise)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_{request.entreprise.id}_{timestamp}.zip"
        backup_path = os.path.join(backup_dir, backup_filename)

        logger.info(f"Début de sauvegarde pour l'entreprise {request.entreprise.id}")

        # 1. Génération du dumpdata
        json_buffer = io.StringIO()
        try:
            call_command(
                'dumpdata',
                '--natural-foreign',
                '--natural-primary',
                '--indent', '2',
                '--exclude', 'contenttypes',
                '--exclude', 'auth.permission',
                '--exclude', 'admin.logentry',
                '--exclude', 'sessions.session',
                stdout=json_buffer
            )
            json_data = json_buffer.getvalue()
            
            if not json_data.strip():
                raise ValueError("Le dumpdata a retourné une chaîne vide")
                
        except Exception as e:
            logger.error(f"Erreur lors de dumpdata: {str(e)}")
            raise ValueError(f"Échec de la génération des données: {str(e)}")

        # 2. Création de l'archive ZIP
        try:
            with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                zip_file.writestr('backup.json', json_data.encode('utf-8'))
            
            # Vérification que le fichier a bien été créé
            if not os.path.exists(backup_path):
                raise IOError("Le fichier de sauvegarde n'a pas été créé")
                
            file_size = os.path.getsize(backup_path)
            logger.info(f"Sauvegarde créée: {backup_path} ({file_size} octets)")
            
            return JsonResponse({
                'status': 'success',
                'message': 'Sauvegarde créée avec succès',
                'filename': backup_filename,
                'path': backup_path,
                'size': file_size,
                'timestamp': timestamp
            })
            
        except Exception as e:
            logger.error(f"Erreur lors de la création du ZIP: {str(e)}")
            # Nettoyage en cas d'erreur
            if os.path.exists(backup_path):
                os.remove(backup_path)
            raise ValueError(f"Échec de la création de l'archive: {str(e)}")

    except Exception as e:
        logger.exception(f"Erreur critique lors de la sauvegarde: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Échec de la sauvegarde',
            'detail': str(e),
            'entreprise_id': getattr(request.entreprise, 'id', None)
        }, status=500)
logger = logging.getLogger(__name__)

def get_backup_dir(entreprise):
    """Retourne le chemin du dossier de backup"""
    return os.path.join(settings.BACKUP_ROOT, f"entreprise_{entreprise.id}")

def is_object_allowed(obj, entreprise):
    """Vérifie si l'objet appartient bien à l'entreprise"""
    return obj.get('entreprise_id', None) == entreprise.id

@csrf_exempt
@login_required
def restore_backup(request):
    """Restauration avec vérification SaaS"""
    if not hasattr(request.user, 'entreprise') or not request.user.entreprise:
        return JsonResponse({'status': 'error', 'message': 'Entreprise non définie'}, status=400)

    if request.method == 'POST':
        filename = request.POST.get('filename')
        if not filename:
            return JsonResponse({'status': 'error', 'message': 'Nom de fichier manquant'}, status=400)

        # Vérification sécurité du nom de fichier
        if not filename.endswith('.zip') or '/' in filename or '\\' in filename:
            return JsonResponse({'status': 'error', 'message': 'Nom de fichier invalide'}, status=400)

        # Vérification que le backup appartient bien à l'entreprise
        if not filename.startswith(f"backup_{request.user.entreprise.id}_"):
            return JsonResponse({'status': 'error', 'message': 'Backup non autorisé'}, status=403)

        backup_dir = get_backup_dir(request.user.entreprise)
        backup_path = os.path.join(backup_dir, filename)
        
        if not os.path.exists(backup_path):
            return JsonResponse({'status': 'error', 'message': 'Fichier introuvable'}, status=404)

        try:
            # Vérification supplémentaire du contenu avant restauration
            with zipfile.ZipFile(backup_path, 'r') as zip_file:
                if 'backup.json' not in zip_file.namelist():
                    return JsonResponse({'status': 'error', 'message': 'Format de backup invalide'}, status=400)
                
                with zip_file.open('backup.json') as f:
                    data = json.load(f)
                    if not all(is_object_allowed(obj, request.user.entreprise) for obj in data):
                        return JsonResponse({'status': 'error', 'message': 'Contenu non autorisé'}, status=403)

            # Restauration propre
            temp_dir = os.path.join(backup_dir, 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            with zipfile.ZipFile(backup_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            call_command('loaddata', os.path.join(temp_dir, 'backup.json'))
            
            # Nettoyage
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)
            
            return JsonResponse({'status': 'success', 'message': 'Sauvegarde restaurée avec succès'})
        except Exception as e:
            logger.error(f"Erreur restauration entreprise {request.user.entreprise.id}: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'Erreur lors de la restauration'}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée'}, status=405)
def is_object_allowed(obj, entreprise):
    """Vérifie si l'objet appartient bien à l'entreprise"""
    # Implémentez cette logique selon vos modèles
    if 'entreprise' in obj['fields']:
        return str(obj['fields']['entreprise']) == str(entreprise.id)
    return True  # ou False selon votre politique

@login_required
def backup_management(request):
    """Gestion des sauvegardes avec isolation SaaS"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        return HttpResponseForbidden("Entreprise non définie")

    # Seuls les admin SaaS ou admin entreprise peuvent accéder
    if not (request.user.is_superuser or request.user.has_perm('saas.manage_backups')):
        return HttpResponseForbidden("Accès refusé")

    try:
        backup_dir = get_backup_dir(request.entreprise)
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        backups = []
        for filename in os.listdir(backup_dir):
            if filename.endswith('.zip') and filename.startswith(f"backup_{request.entreprise.id}_"):
                filepath = os.path.join(backup_dir, filename)
                stat = os.stat(filepath)

                backups.append({
                    'filename': filename,
                    'path': filepath,
                    'size': stat.st_size,
                    'date': stat.st_mtime,
                    'human_size': sizeof_fmt(stat.st_size),
                    'formatted_date': datetime.fromtimestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M")
                })

        backups.sort(key=lambda x: x['date'], reverse=True)

        return render(request, 'backup/backup_management.html', {
            'backups': backups,
            'backup_dir': backup_dir,
            'title': f'Gestion des sauvegardes - {request.entreprise.nom}',
            'entreprise': request.entreprise
        })

    except Exception as e:
        logger.error(f"Erreur backup_management entreprise {request.entreprise.id}: {str(e)}")
        return render(request, 'backup/backup_management.html', {
            'backups': [],
            'error': str(e),
            'entreprise': request.entreprise
        })

@login_required
def download_backup(request, filename):
    """Téléchargement avec vérification SaaS"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        return JsonResponse({'status': 'error', 'message': 'Entreprise non définie'}, status=400)

    # Vérification que le backup appartient bien à l'entreprise
    if not filename.startswith(f"backup_{request.entreprise.id}_"):
        return JsonResponse({'status': 'error', 'message': 'Backup non autorisé'}, status=403)

    backup_path = os.path.join(get_backup_dir(request.entreprise), filename)
    if not os.path.exists(backup_path):
        return JsonResponse({'status': 'error', 'message': 'Fichier non trouvé'}, status=404)

    response = FileResponse(open(backup_path, 'rb'))
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def delete_backup(request):
    """Suppression avec vérification SaaS"""
    if not hasattr(request, 'entreprise') or not request.entreprise:
        return JsonResponse({'status': 'error', 'message': 'Entreprise non définie'}, status=400)

    filename = request.GET.get('filename')
    if not filename:
        return JsonResponse({'status': 'error', 'message': 'Nom de fichier manquant'}, status=400)

    # Vérification que le backup appartient bien à l'entreprise
    if not filename.startswith(f"backup_{request.entreprise.id}_"):
        return JsonResponse({'status': 'error', 'message': 'Backup non autorisé'}, status=403)

    backup_path = os.path.join(get_backup_dir(request.entreprise), filename)
    if not os.path.exists(backup_path):
        return JsonResponse({'status': 'error', 'message': 'Fichier non trouvé'}, status=404)

    try:
        os.remove(backup_path)
        return JsonResponse({'status': 'success', 'message': 'Sauvegarde supprimée'})
    except Exception as e:
        logger.error(f"Erreur suppression backup {filename} entreprise {request.entreprise.id}: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def sizeof_fmt(num, suffix='o'):
    """Convertit une taille en format lisible (Ko, Mo, Go, etc.)"""
    for unit in ['','K','M','G','T','P','E','Z']:
        if abs(num) < 1024.0:
            return "%3.1f %s%s" % (num, unit, suffix)
        num /= 1024.0
    return "%.1f %s%s" % (num, 'Y', suffix)